from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

from actions._read_only_common import bulletize_text, summarize_text
from actions._write_common import artifact_payload, ensure_allowed_output_path, normalize_source_context
from runtime.capability_registry import SafeCapabilityError


def _workspace_root() -> Path:
    from actions._read_only_common import workspace_root

    return workspace_root()


def _safe_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float)):
        return value
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


_COLUMN_ALIAS_GROUPS = (
    {"title", "baslik", "konu", "name", "ad"},
    {"url", "href", "link", "adres", "sourceurl", "kaynakurl"},
    {"summary", "ozet", "snippet", "description", "aciklama", "text", "icerik"},
    {"source", "kaynak", "origin"},
    {"date", "tarih", "datetime", "timestamp"},
    {"amount", "tutar", "value", "deger", "price", "fiyat"},
)


def _normalized_key(value: Any) -> str:
    folded = unicodedata.normalize(
        "NFKD",
        str(value or "").casefold().replace("ı", "i"),
    )
    return "".join(char for char in folded if char.isalnum())


def _column_candidates(column: str) -> set[str]:
    normalized = _normalized_key(column)
    for group in _COLUMN_ALIAS_GROUPS:
        if normalized in group:
            return group
    return {normalized}


def _row_value(row: dict[Any, Any], column: str) -> Any:
    if column in row:
        return row[column]
    by_key = {
        _normalized_key(key): value
        for key, value in row.items()
        if _normalized_key(key)
    }
    for candidate in _column_candidates(column):
        if candidate in by_key:
            return by_key[candidate]
    return None


def _normalized_columns(columns: list[str] | None) -> list[str]:
    normalized: list[str] = []
    used: set[str] = set()
    for index, item in enumerate(columns or [], start=1):
        base = (" ".join(str(item or "").split()) or f"Column {index}")[:120]
        candidate = base
        suffix = 2
        key = _normalized_key(candidate)
        while not key or key in used:
            suffix_text = f" {suffix}"
            candidate = f"{base[: 120 - len(suffix_text)]}{suffix_text}"
            suffix += 1
            key = _normalized_key(candidate)
        used.add(key)
        normalized.append(candidate)
    return normalized


def _safe_sheet_title(value: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", " ", str(value or ""))
    cleaned = " ".join(cleaned.split()).strip(" '")
    return (cleaned or "Elyan Sheet")[:31]


def _inferred_columns(rows: Any) -> list[str]:
    if not isinstance(rows, list):
        return []
    columns: list[str] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        for key in row:
            name = str(key).strip()
            if name and name not in columns:
                columns.append(name)
            if len(columns) >= 128:
                return columns
    if columns:
        return columns
    matrix_width = max(
        (
            len(row)
            if isinstance(row, (list, tuple))
            else 1
            if row is not None
            else 0
        )
        for row in rows
    ) if rows else 0
    if matrix_width:
        return [f"Column {index}" for index in range(1, min(matrix_width, 128) + 1)]
    return columns


def _normalize_rows(rows: Any, columns: list[str] | None = None) -> list[list[Any]]:
    normalized: list[list[Any]] = []
    if not isinstance(rows, list):
        return normalized
    target_width = len(columns or [])
    for row in rows:
        if isinstance(row, dict):
            values = [_row_value(row, column) for column in columns] if columns else list(row.values())
        elif isinstance(row, (list, tuple)):
            values = list(row)
        elif row is not None:
            values = [row]
        else:
            continue
        if target_width:
            values = values[:target_width] + [None] * max(0, target_width - len(values))
        normalized.append([_safe_cell_value(value) for value in values])
    return normalized


def _first_text(mapping: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = mapping.get(key)
        if value not in (None, "", [], {}):
            text = str(value).strip()
            if text:
                return text
    return ""


def _first_list(mapping: dict[str, Any], *keys: str) -> list[Any] | None:
    for key in keys:
        value = mapping.get(key)
        if isinstance(value, list):
            return value
    return None


def _coerce_spreadsheet_inputs(
    *,
    prompt: str,
    title: str,
    columns: list[str] | None,
    rows: list[Any] | None,
    kwargs: dict[str, Any],
) -> tuple[str, str, list[str] | None, list[Any] | None]:
    prompt = str(prompt or "").strip() or _first_text(kwargs, "content", "summary", "description", "text")
    title = str(title or "").strip() or _first_text(kwargs, "name", "sheetTitle", "sheet_title")
    columns = columns if isinstance(columns, list) else _first_list(kwargs, "headers", "fields")
    rows = rows if isinstance(rows, list) else _first_list(kwargs, "data", "items", "records", "values")
    table = None
    for key in ("table", "worksheet", "sheet", "spreadsheet"):
        value = kwargs.get(key)
        if isinstance(value, dict):
            table = value
            break
    sheets = kwargs.get("sheets") if isinstance(kwargs.get("sheets"), list) else kwargs.get("worksheets")
    if table is None and isinstance(sheets, list) and sheets and isinstance(sheets[0], dict):
        table = sheets[0]
    if isinstance(table, dict):
        prompt = prompt or _first_text(table, "prompt", "content", "summary", "description", "text")
        title = title or _first_text(table, "title", "name", "sheetTitle", "sheet_title")
        columns = columns or _first_list(table, "columns", "headers", "fields")
        rows = rows or _first_list(table, "rows", "data", "items", "records", "values")
    if columns is None and isinstance(rows, list) and rows:
        first = rows[0]
        if isinstance(first, list) and len(first) > 1 and all(not isinstance(item, (list, dict)) for item in first):
            columns = [str(item) for item in first]
            rows = rows[1:]
    return prompt, title, columns, rows


def spreadsheet_write(
    prompt: str = "",
    output_path: str = "",
    title: str = "",
    columns: list[str] | None = None,
    rows: list[Any] | None = None,
    source_context: str = "",
    overwrite: bool = False,
    **kwargs: Any,
) -> dict[str, Any]:
    from openpyxl import Workbook  # type: ignore[reportMissingImports]
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    prompt, title, columns, rows = _coerce_spreadsheet_inputs(
        prompt=prompt,
        title=title,
        columns=columns,
        rows=rows,
        kwargs=kwargs,
    )
    if not str(output_path or "").strip():
        output_path = str(kwargs.get("outputPath", "") or kwargs.get("output_path", "") or "")
    if not str(source_context or "").strip():
        source_context = str(kwargs.get("sourceContext", "") or kwargs.get("source_context", "") or "")

    resolved_output = ensure_allowed_output_path(
        output_path,
        extension=".xlsx",
        overwrite=overwrite,
        hint=title or prompt or "elyan-sheet",
        root_resolver=_workspace_root,
    )
    seed_text = str(source_context or prompt or "").strip()
    if not seed_text and not rows:
        raise SafeCapabilityError("INVALID_ARGUMENT", "Çalışma sayfası oluşturmak için içerik veya satır verisi gerekli.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_title(title)

    normalized_columns = _normalized_columns(columns)
    if not normalized_columns:
        normalized_columns = _normalized_columns(_inferred_columns(rows))
    normalized_rows = _normalize_rows(rows, normalized_columns)
    written_rows = list(normalized_rows)
    if normalized_columns:
        sheet.append(normalized_columns)
    if normalized_rows:
        for row in normalized_rows:
            sheet.append(row)
    else:
        header = normalized_columns or ["Content"]
        if not normalized_columns:
            sheet.append(header)
        written_rows = [] if not normalized_columns else [[bullet] for bullet in bulletize_text(seed_text, limit=24)]
        for row in written_rows:
            sheet.append(row)

    if sheet.max_row >= 1 and sheet.max_column >= 1:
        header_fill = PatternFill("solid", fgColor="243447")
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        if normalized_columns and sheet.max_row > 1:
            table = Table(displayName="ElyanData", ref=sheet.dimensions)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        numeric_names = {"tutar", "amount", "fiyat", "price", "gelir", "gider", "deger", "değer", "value"}
        for column_index in range(1, sheet.max_column + 1):
            header = str(sheet.cell(1, column_index).value or "").strip().casefold()
            values = [sheet.cell(row_index, column_index).value for row_index in range(1, sheet.max_row + 1)]
            width = min(48, max(10, max((len(str(value)) for value in values if value is not None), default=8) + 2))
            sheet.column_dimensions[get_column_letter(column_index)].width = width
            if header in numeric_names:
                for row_index in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row_index, column_index)
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.number_format = "#,##0.00" if isinstance(cell.value, float) else "#,##0"

    workbook.save(str(resolved_output))
    return {
        "text": f"XLSX oluşturuldu: {resolved_output.name}",
        "result": {
            "kind": "spreadsheet_write",
            "sourceContext": normalize_source_context(seed_text or "structured_rows"),
            "outputPath": str(resolved_output),
            "contentType": artifact_payload(resolved_output)["contentType"],
            "created": True,
            "title": sheet.title,
            "summary": summarize_text(seed_text or "Spreadsheet created.", max_chars=260),
            "columns": normalized_columns,
            "rows": written_rows[:12],
            "rowCount": len(written_rows),
        },
        "artifacts": [artifact_payload(resolved_output)],
    }
