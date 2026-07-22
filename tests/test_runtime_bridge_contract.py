from __future__ import annotations

import json
import threading
import time
from pathlib import Path

import pytest

from actions import browser_session
from runtime import bridge, state_store, task_router
from runtime.backend_client import BackendResult
from runtime.capability_registry import SafeCapabilityError

VALID_DEVICE_ID = "11111111-1111-4111-8111-111111111111"
VALID_CONNECTION_ID = "22222222-2222-4222-8222-222222222222"
VALID_CHAT_SESSION_ID = "33333333-3333-4333-8333-333333333333"
VALID_DEVICE_SECRET = "device-secret-123456"


@pytest.fixture(autouse=True)
def _deterministic_ollama_tags(monkeypatch: pytest.MonkeyPatch) -> None:
    """Testler canlı ollama'ya problamasın: yeni sözleşmede ollama ancak model
    GERÇEKTEN kuruluysa aday olur; testlerde kurulu modelleri cache'e tohumla."""
    monkeypatch.setattr(
        bridge,
        "_OLLAMA_TAGS_CACHE",
        {"at": float("inf"), "names": ["llama3.2:3b", "llama3.1:8b", "qwen2.5:7b"]},
    )


def _isolate_state(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    state_path = tmp_path / "elyan_state.json"
    monkeypatch.setattr(state_store, "CONFIG_DIR", tmp_path)
    monkeypatch.setattr(state_store, "STATE_PATH", state_path)
    monkeypatch.setattr(state_store, "LEGACY_STATE_PATH", tmp_path / "legacy.json")


def _arm_device_identity() -> None:
    """P0 güven zinciri: atanmış görev yürütmesi cihaz kimliği + sırrı ister."""
    state_store.update_state(
        {"runtime": {"deviceId": VALID_DEVICE_ID, "deviceSecret": VALID_DEVICE_SECRET}}
    )


def _work_order_envelope(
    prompt: str,
    capabilities: list[str],
    steps: list[dict] | None = None,
) -> dict:
    if steps is None:
        # Capability listesi sıralı bir boru hattıdır (draft, research çıktısını
        # tüketir) — P2 scheduler bağımsız read'leri paralelleştirdiğinden örtük
        # sıra açık dependsOn zincirine çevrilir.
        steps = []
        for index, capability in enumerate(capabilities):
            step = {"id": f"step_{index + 1}", "capability": capability, "description": capability, "args": {}}
            if index > 0:
                step["dependsOn"] = [f"step_{index}"]
            steps.append(step)
    return {
        "schema": "elyan.desktop_work_order.v1",
        "source": "mobile_chat_dispatch",
        "goal": {"kind": "computer_task", "summary": prompt, "language": "tr", "sourceTextHash": "a" * 24},
        "entities": [],
        "constraints": [],
        "requiredCapabilities": list(capabilities),
        "localContextNeeded": [],
        "expectedOutputs": [{"kind": "chat_result", "format": "elyan_blocks.v2", "required": True}],
        "verificationRules": [
            {"id": "runtime_completed", "description": "Runtime completes.", "evidence": "runtime_status"},
        ],
        "execution": {"mode": "cowork_dispatch", "approvalPolicy": "capability_policy", "maxSteps": 8},
        "planPreview": {"summary": prompt, "privacyClass": "local_private", "steps": steps},
    }


def _trusted_task(
    task: dict,
    *,
    capabilities: list[str],
    steps: list[dict] | None = None,
    user_id: str = "user-tests",
) -> dict:
    """Atanmış görevi P0 WorkOrder güven bağı için gerekli alanlarla donat."""
    payload = task.setdefault("payload", {})
    prompt = str(payload.get("prompt") or task.get("title") or "")
    payload.setdefault("desktopWorkOrder", _work_order_envelope(prompt, capabilities, steps))
    task.setdefault("userId", user_id)
    task.setdefault("targetDeviceId", VALID_DEVICE_ID)
    return task


def _write_native_desktop_snapshot(tmp_path: Path) -> Path:
    path = tmp_path / "desktop-runtime.json"
    path.write_text(
        json.dumps(
            {
                "available": True,
                "source": "native_addon",
                "collectedAt": "2026-06-03T12:00:00Z",
                "platform": "darwin",
                "osPermissionModel": "macos_privacy_tcc",
                "processInspectionAvailable": True,
                "activeWindowAvailable": True,
                "permissionProbeAvailable": True,
                "globalShortcutsAvailable": True,
                "screenCaptureAvailable": True,
                "permissions": {
                    "accessibility": {"required": True, "granted": None, "status": "required"},
                    "screenRecording": {"required": True, "granted": None, "status": "required"},
                    "inputMonitoring": {"required": True, "granted": None, "status": "required"},
                    "automation": {"required": True, "granted": None, "status": "required"},
                },
                "processes": {"available": True, "total": 1, "items": [{"pid": 101, "name": "Finder"}]},
                "activeWindow": {
                    "available": True,
                    "appName": "Finder",
                    "windowTitle": "Desktop",
                    "processId": 101,
                },
                "lastErrorCode": "",
            }
        ),
        encoding="utf-8",
    )
    return path


def test_user_auth_ready_accepts_refresh_token_only(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    # No accessToken, only refreshToken — should still be considered ready
    state_store.update_state({"account": {"refreshToken": "rt_abc123", "accessToken": ""}})
    assert runtime._user_auth_ready() is True


def test_user_auth_ready_false_with_no_tokens(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.update_state({"account": {"refreshToken": "", "accessToken": ""}})
    assert runtime._user_auth_ready() is False


def test_envelope_result_contains_conversation_state(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    def fake_route(_state: dict, _conversation: list, text: str, **_kwargs: object) -> dict:
        return {"ok": True, "content": f"Yanıt: {text}", "provider": "test"}

    monkeypatch.setattr(bridge, "_route_chat", fake_route)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_test",
            "taskId": "task_test",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "merhaba"},
        }
    )

    assert response["ok"] is True
    result = response["result"]
    assert result["assistantMessage"] == "Yanıt: merhaba"
    assert result["state"]["conversation"]["activeId"]
    assert result["conversations"][0]["messageCount"] == 2


def test_signed_in_new_chat_stays_local_blank_until_first_message(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state({"account": {"accessToken": "user-token"}})

    class FakeBackend:
        configured = True
        loopback = False

        def chat_session_create(self, _payload: dict[str, object]) -> BackendResult:
            raise AssertionError("blank chats must not be persisted")

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    response = runtime.create_conversation()

    assert response["ok"] is True
    assert response["conversationId"] == ""
    assert response["state"]["conversation"]["activeId"] == ""
    assert response["state"]["conversation"]["items"] == []


def test_chat_messages_include_current_user_text() -> None:
    messages = bridge._chat_messages(
        [{"role": "assistant", "text": "Hazırım."}],
        "Sistem bilgisini göster",
    )

    assert messages[-1] == {"role": "user", "content": "Sistem bilgisini göster"}


def test_conversation_send_failure_still_returns_visible_state(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    monkeypatch.setattr(
        bridge,
        "_route_chat",
        lambda _state, _conversation, _text, **_kwargs: {"ok": False, "error": "ollama_model_missing"},
    )
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_error",
            "taskId": "task_error",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "merhaba"},
        }
    )

    assert response["ok"] is True
    result = response["result"]
    assert result["chatOk"] is False
    assert result["assistantMessage"] == "Yerel model seçilmemiş."
    assert result["state"]["conversation"]["activeId"]
    assert result["conversations"][0]["messageCount"] == 2


def test_model_name_active_provider_is_normalized() -> None:
    state = state_store._ensure_defaults(
        {
            "providers": {
                "active": "llama3.2:3b",
                "ollama": {"defaultModel": "llama3.1:8b"},
                "local": {"defaultModel": ""},
            }
        }
    )

    assert state["providers"]["active"] == "ollama"
    assert state["providers"]["ollama"]["defaultModel"] == "llama3.2:3b"
    assert state["providers"]["local"]["defaultModel"] == "llama3.2:3b"


def test_bootstrap_includes_brain_profile_surface(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state({"account": {"accessToken": "user-token"}})

    class FakeBackend:
        configured = True
        loopback = False

        def auth_me(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_me", status_code=200, data={"user": {"email": "user@example.com"}})

        def mobile_bootstrap(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_bootstrap", status_code=200, data={"devices": []})

        def health(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_health", status_code=200, data={"dependencies": {}})

        def brain_profile(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_brain",
                status_code=200,
                data={"chat": {"brainProfilePath": "/v1/brain/profile"}},
            )

    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "controlPlane": {
                "health": {
                    "ok": True,
                    "agent": {"chatReady": True, "serverBrainReady": True},
                }
            }
        }
    )
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    payload = runtime.bootstrap()

    assert payload["backend"]["brainProfile"]["ok"] is True
    assert payload["backend"]["brainProfile"]["data"]["chat"]["brainProfilePath"] == "/v1/brain/profile"


def test_bootstrap_reconnects_process_transport_when_persisted_runtime_is_ready(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "runtime": {
                "runtimeToken": "persisted-runtime-token",
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "ready": True,
                "lifecycleState": "ready",
                "websocketConnected": True,
            }
        }
    )
    runtime = bridge.RuntimeBridge()
    connect_calls: list[bool] = []
    monkeypatch.setattr(runtime, "_runtime_backend_snapshot", lambda: {"ok": True})
    monkeypatch.setattr(
        runtime,
        "_connect_runtime_transport",
        lambda: connect_calls.append(True) or (False, None),
    )
    monkeypatch.setattr(runtime, "status", lambda: {"ok": True})

    payload = runtime.bootstrap()

    assert payload["runtime"]["ok"] is True
    assert connect_calls == [True]


def test_planning_envelope_includes_native_desktop_truth_and_intelligence(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    # Eski serbest metin prompt yerine yapılandırılmış zarf (elyan.plan.v2):
    # masaüstü canlı durumu ve öğrenilmiş rota geçmişi VERİ olarak taşınmalı.
    _isolate_state(monkeypatch, tmp_path)
    monkeypatch.setattr(
        bridge,
        "_desktop_native_snapshot_payload",
        lambda: {
            "available": True,
            "platform": "darwin",
            "source": "native_addon",
            "activeWindow": {"appName": "Finder", "windowTitle": "Desktop"},
            "processes": {"total": 1},
            "processInspectionAvailable": True,
            "activeWindowAvailable": True,
            "permissionProbeAvailable": True,
            "operator": {
                "available": True,
                "screenObservationReady": True,
                "accessibilityReady": True,
                "inputControlReady": True,
            },
            "permissions": {"accessibility": {"status": "granted"}},
        },
    )
    state_store.record_recent_route(query="selam", intent="chat", capability="open_app", confidence=0.9)
    state_store.record_recent_route(query="hava", intent="weather", capability="get_weather", confidence=0.92)
    state_store.record_recent_route(
        query="belge ozetle",
        intent="document_read",
        capability="document_read",
        confidence=0.94,
    )

    envelope = bridge._build_structured_planning_request(state_store.snapshot(), "pencereyi aç")

    assert envelope["contract"] == "elyan.plan.v2"
    desktop = envelope["context"]["desktop"]
    assert desktop["activeWindow"]["appName"] == "Finder"
    assert desktop["operator"]["accessibilityReady"] is True
    assert desktop["permissions"]["accessibility"] == "granted"
    recent = envelope["context"]["recentIntents"]
    assert any(item.get("capability") == "get_weather" for item in recent)
    assert all(isinstance(item, dict) for item in recent)
    # Araç kataloğu JSON Schema parametreli olmalı; düz metin talimat yok.
    assert isinstance(envelope["toolCatalog"], list) and envelope["toolCatalog"]
    assert envelope["responseSchema"]["properties"]["contract"]["const"] == "elyan.plan.v2"


def test_runtime_status_includes_operator_state(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "operator": {
                "activeRunId": "oprun_status",
                "status": "waiting_approval",
                "abortRequested": False,
                "abortReason": "",
                "currentStepIndex": 1,
                "lastObservationId": "obs_status",
                "lastStopReason": "verification_failed",
            }
        }
    )

    runtime = bridge.RuntimeBridge()
    payload = runtime.status()

    assert payload["operatorStatus"]["activeRunId"] == ""
    assert payload["operatorStatus"]["status"] == "stopped"
    assert payload["operatorStatus"]["lastStopReason"] == "runtime_restarted"


def test_runtime_status_includes_full_desktop_native_snapshot(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    snapshot_path = _write_native_desktop_snapshot(tmp_path)
    monkeypatch.setenv("ELYAN_DESKTOP_NATIVE_STATE_PATH", str(snapshot_path))

    runtime = bridge.RuntimeBridge()
    payload = runtime.status()

    assert payload["desktopNative"]["available"] is True
    assert payload["desktopNative"]["platform"] == "darwin"
    assert payload["desktopNative"]["activeWindow"]["appName"] == "Finder"
    assert payload["desktopNative"]["nativeReadiness"]["runtimeReady"] is True
    assert payload["agentStatus"]["nativeReadiness"]["runtimeReady"] is True
    assert payload["agentStatus"]["degradationReasons"] == []
    assert payload["desktopNativeStatus"]["available"] is True


def test_pending_plan_persists_agent_plan(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(bridge, "route_text_to_tool", lambda _text, **_kwargs: None)
    monkeypatch.setattr(
        bridge,
        "_semantic_route",
        lambda _state, _conversation, _text, **_kwargs: {
            "intent": "document_write",
            "capability": "document_write",
            "args": {"outputPath": "/tmp/elyan-note.txt", "title": "Not"},
            "confidence": 0.97,
            "requiresConfirmation": True,
            "isMultiStep": True,
            "privacyClass": "local_private",
            "planPreview": {
                "summary": "Not oluşturulacak.",
                "steps": [
                    {"capability": "retrieve_context", "args": {"query": "not"}},
                    {"capability": "document_write", "args": {"outputPath": "/tmp/elyan-note.txt"}},
                ],
                "agentPlan": {
                    "summary": "Not oluşturulacak.",
                    "stepCount": 2,
                    "capabilities": ["retrieve_context", "document_write"],
                    "stepRoles": [
                        {"capability": "retrieve_context", "role": "observer"},
                        {"capability": "document_write", "role": "writer"},
                    ],
                    "agentRoles": ["planner", "observer", "writer"],
                    "roleBreakdown": [
                        {"role": "observer", "stepCount": 1, "capabilities": ["retrieve_context"]},
                        {"role": "writer", "stepCount": 1, "capabilities": ["document_write"]},
                    ],
                    "laneCount": 3,
                    "executionStrategy": "multi_lane",
                },
            },
            "provider": "ollama",
        },
    )

    response = runtime.handle(
        {
            "id": "req_pending_agent_plan",
            "taskId": "task_pending_agent_plan",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "not hazırla"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["needsConfirmation"] is True
    pending_plan = state_store.get_pending_plan(response["result"]["pendingPlanId"])
    assert pending_plan is not None
    assert pending_plan["agentPlan"]["executionStrategy"] == "multi_lane"
    assert pending_plan["stepCount"] == 2
    assert pending_plan["agentRoles"] == ["planner", "observer", "writer"]


def test_execute_assigned_runtime_task_reports_local_result(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []
            self.heartbeats: list[dict] = []

        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        _trusted_task(
                            {
                                "id": "task-1",
                                "title": "Test görevi",
                                "payload": {"prompt": "merhaba"},
                            },
                            capabilities=["retrieve_context"],
                        )
                    ]
                },
            )

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

        def heartbeat(self, payload: dict) -> BackendResult:
            self.heartbeats.append(payload)
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    fake_backend = FakeBackend()
    monkeypatch.setattr(
        bridge,
        "_route_chat",
        lambda _state, _conversation, text, **_kwargs: {"ok": True, "content": f"Yanıt: {text}", "provider": "test"},
    )
    runtime = bridge.RuntimeBridge()
    _arm_device_identity()
    runtime.backend = fake_backend  # type: ignore[assignment]

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"][0]["ok"] is True
    assert result["executions"][0]["status"] == "completed"
    assert fake_backend.status_updates[0][1]["status"] == "running"
    completed = fake_backend.status_updates[-1][1]
    assert completed["status"] == "completed"
    assert completed["result"]["assistantMessage"]
    assert completed["artifacts"][0]["kind"] == "summary"


def test_mobile_assigned_task_reuses_backend_chat_session_in_worker(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    routed_session_ids: list[str] = []

    monkeypatch.setattr(
        runtime,
        "_execute_deterministic_remote_task",
        lambda *_args, **_kwargs: None,
    )

    def fake_send_conversation(*_args: object, **_kwargs: object) -> dict:
        import importlib

        active_bridge = importlib.import_module("runtime.bridge")
        routed_session_ids.append(active_bridge._ACTIVE_DISPATCH_SESSION_ID.get())
        return {"ok": True, "assistantMessage": "Devam edildi."}

    monkeypatch.setattr(runtime, "send_conversation", fake_send_conversation)

    result = runtime.remote_task_runner._execute_local_with_timeout(
        {
            "payload": {
                "metadata": {"chat": {"sessionId": VALID_CHAT_SESSION_ID}},
            },
        },
        "önceki konuşmaya devam et",
        "Devam",
        task_id="task-mobile-session",
    )

    assert result["ok"] is True
    assert routed_session_ids == [VALID_CHAT_SESSION_ID]
    import importlib

    active_bridge = importlib.import_module("runtime.bridge")
    assert active_bridge._ACTIVE_DISPATCH_SESSION_ID.get() == ""


def test_quantum_capabilities_are_registered() -> None:
    capabilities = bridge.capability_names()

    assert "quantum_model_problem" in capabilities
    assert "quantum_run_experiment" in capabilities
    assert "quantum_compare_classical" in capabilities
    assert "quantum_generate_report" in capabilities


def test_execute_assigned_quantum_task_uses_deterministic_pipeline(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    from actions import quantum as quantum_actions

    monkeypatch.setattr(quantum_actions, "_qiskit_available", lambda: True)

    class FakeBackend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []
            self.heartbeats: list[dict] = []

        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        _trusted_task({
                            "id": "task-quantum",
                            "title": "Quantum demo",
                            "requestedCapabilities": [
                                "quantum_model_problem",
                                "quantum_run_experiment",
                                "quantum_compare_classical",
                                "quantum_generate_report",
                            ],
                            "payload": {
                                "prompt": "İki değişkenli QUBO problemi oluştur, QAOA ile simüle et ve klasik çözümle karşılaştırıp raporla.",
                                "metadata": {
                                    "routeDecision": {
                                        "route": "desktop_runtime",
                                        "mode": "mixed_task",
                                        "capabilities": [
                                            "quantum_model_problem",
                                            "quantum_run_experiment",
                                            "quantum_compare_classical",
                                            "quantum_generate_report",
                                        ],
                                        "privacyClass": "public_text",
                                        "requiresApproval": False,
                                        "reason": "quantum test",
                                    }
                                },
                            },
                        }, capabilities=[
                            "quantum_model_problem",
                            "quantum_run_experiment",
                            "quantum_compare_classical",
                            "quantum_generate_report",
                        ], steps=[])
                    ]
                },
            )

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

        def heartbeat(self, payload: dict) -> BackendResult:
            self.heartbeats.append(payload)
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "controlPlane": {
                "health": {
                    "ok": True,
                    "agent": {"chatReady": True, "serverBrainReady": True},
                }
            }
        }
    )
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    _arm_device_identity()

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"][0]["status"] == "completed"
    completed = runtime.backend.status_updates[-1][1]  # type: ignore[attr-defined]
    assert completed["status"] == "completed"
    assert completed["result"]["quantum"]["mode"] == "hybrid"
    assert completed["result"]["structuredResult"]["kind"] == "quantum_report"
    assert any(artifact["contentType"] == "text/markdown" for artifact in completed["artifacts"])


def test_execute_assigned_quantum_task_fails_safely_without_qiskit(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    from actions import quantum as quantum_actions

    monkeypatch.setattr(quantum_actions, "_qiskit_available", lambda: False)

    class FakeBackend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []
            self.heartbeats: list[dict] = []

        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        _trusted_task({
                            "id": "task-quantum-missing-dep",
                            "title": "Quantum demo",
                            "requestedCapabilities": [
                                "quantum_model_problem",
                                "quantum_run_experiment",
                                "quantum_compare_classical",
                                "quantum_generate_report",
                            ],
                            "payload": {
                                "prompt": "QUBO modelle ve QAOA çalıştır.",
                                "metadata": {
                                    "routeDecision": {
                                        "route": "desktop_runtime",
                                        "mode": "mixed_task",
                                        "capabilities": [
                                            "quantum_model_problem",
                                            "quantum_run_experiment",
                                            "quantum_compare_classical",
                                            "quantum_generate_report",
                                        ],
                                        "privacyClass": "public_text",
                                        "requiresApproval": False,
                                    }
                                },
                            },
                        }, capabilities=[
                            "quantum_model_problem",
                            "quantum_run_experiment",
                            "quantum_compare_classical",
                            "quantum_generate_report",
                        ], steps=[])
                    ]
                },
            )

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

        def heartbeat(self, payload: dict) -> BackendResult:
            self.heartbeats.append(payload)
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "controlPlane": {
                "health": {
                    "ok": True,
                    "agent": {"chatReady": True, "serverBrainReady": True},
                }
            }
        }
    )
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    _arm_device_identity()

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"][0]["status"] == "failed"
    failed = runtime.backend.status_updates[-1][1]  # type: ignore[attr-defined]
    assert failed["status"] == "failed"
    assert failed["error"] == "QUANTUM_DEPENDENCY_UNAVAILABLE"
    assert "Qiskit" in failed["result"]["assistantMessage"]


def test_runtime_quantum_execution_capability_requires_qiskit_aer(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    original_module_available = bridge._module_available

    def fake_module_available(module_name: str) -> bool:
        if module_name in {"qiskit", "qiskit_aer"}:
            return False
        return original_module_available(module_name)

    monkeypatch.setattr(bridge, "_module_available", fake_module_available)
    runtime = bridge.RuntimeBridge()

    capabilities = runtime._runtime_heartbeat_payload("online")["capabilities"]
    status = runtime.status()

    assert "quantum_model_problem" in capabilities
    assert "quantum_run_experiment" not in capabilities
    assert "quantum_compare_classical" in capabilities
    assert "quantum_generate_report" in capabilities
    assert status["dependencyStatus"]["qiskit"]["available"] is False
    assert status["dependencyStatus"]["qiskit_aer"]["available"] is False


def test_execute_assigned_runtime_task_waits_for_approval_without_terminal_report(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    # Plan modu (composer'dan planMode=true): zararsız görev bile plan
    # önizlemesi + onay ile ilerler; onaysız yürütülmez.
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []
            self.heartbeats: list[dict] = []

        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        _trusted_task(
                            {
                                "id": "task-waiting",
                                "title": "Takvim görevi",
                                "status": "queued",
                                "payload": {
                                    "prompt": "takvime cuma 14:00 ürün toplantısı ekle",
                                    "metadata": {"planMode": True},
                                },
                            },
                            capabilities=["add_calendar_event"],
                        )
                    ]
                },
            )

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

        def heartbeat(self, payload: dict) -> BackendResult:
            self.heartbeats.append(payload)
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "controlPlane": {
                "health": {
                    "ok": True,
                    "agent": {"chatReady": True, "serverBrainReady": True},
                }
            }
        }
    )
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(
        runtime,
        "send_conversation",
        lambda *_args, **_kwargs: {
            "ok": True,
            "chatOk": True,
            "assistantMessage": "Takvime ürün toplantısı eklenecek.",
            "provider": "local_planner",
            "toolEvents": [],
            "conversationId": "conv_waiting",
            "needsConfirmation": True,
            "pendingPlanId": "plan_waiting",
            "planPreview": {
                "summary": "Takvime 'ürün toplantısı' eklenecek.",
                "steps": [{"capability": "add_calendar_event", "description": "Takvim etkinliği oluştur"}],
            },
        },
    )

    _arm_device_identity()

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"][0]["status"] == "waiting_approval"
    statuses = [payload["status"] for _, payload in runtime.backend.status_updates]  # type: ignore[attr-defined]
    assert statuses[0] == "running"
    assert statuses[-1] == "waiting_approval"
    assert runtime.backend.status_updates[-1][1]["approvalRequest"]["summary"]  # type: ignore[attr-defined]
    assert runtime.backend.status_updates[-1][1]["approvalRequest"]["manualApprovalRequired"] is True  # type: ignore[attr-defined]
    link = state_store.get_remote_task_link("task-waiting")
    assert link is not None
    assert link["pendingPlanId"] == "plan_waiting"


def test_approval_request_payload_includes_email_context() -> None:
    runtime = bridge.RuntimeBridge()

    payload = runtime._approval_request_payload(
        {
            "assistantMessage": "Mail taslağı hazır.",
            "planPreview": {
                "summary": "Atatürk hakkında kaynaklar toparlandı ve mail gönderimi için taslak hazırlandı.",
                "steps": [
                    {"capability": "web_research", "description": "Atatürk hakkında web araştırması yapılacak."},
                    {"capability": "email_draft", "description": "ali@example.com için e-posta taslağı hazırlanacak."},
                ],
            },
            "structuredResult": {
                "kind": "email_send",
                "to": ["ali@example.com"],
                "subject": "Atatürk hakkında notlar",
                "body": "Merhaba,\n\nKısa özet burada yer alıyor.\n\nİyi çalışmalar.",
                "provider": "google",
            },
        }
    )

    assert payload["title"] == "Mail gönderilsin mi?"
    assert payload["summary"].startswith("Atatürk hakkında kaynaklar")
    assert "Alıcı: ali@example.com" in payload["message"]
    assert "Konu: Atatürk hakkında notlar" in payload["message"]
    assert "Özet: Merhaba," in payload["message"]
    assert payload["kind"] == "email_send"
    assert payload["to"] == ["ali@example.com"]
    assert payload["recipientCount"] == 1
    assert payload["subject"] == "Atatürk hakkında notlar"
    assert payload["bodyPreview"].startswith("Merhaba,")
    assert payload["provider"] == "google"
    assert payload["capability"] == "email_send"
    assert payload["permission"] == "side_effect"
    assert payload["idempotency"] == "non_idempotent"
    assert payload["confirmLabel"] == "Onayla"
    assert payload["rejectLabel"] == "Reddet"


def test_approval_request_payload_classifies_only_registered_safe_writes_as_idempotent() -> None:
    runtime = bridge.RuntimeBridge()

    safe = runtime._approval_request_payload(
        {
            "assistantMessage": "Belge hazır.",
            "planPreview": {
                "summary": "Yeni görev belgesi oluşturulacak.",
                "steps": [
                    {"capability": "web_research", "description": "Kaynakları oku."},
                    {"capability": "document_write", "description": "Belgeyi görev çıktısı olarak oluştur."},
                ],
            },
        }
    )
    unknown = runtime._approval_request_payload(
        {
            "assistantMessage": "İşlem hazır.",
            "planPreview": {
                "summary": "Sınıflandırılmamış işlem.",
                "steps": [{"capability": "unknown.write", "description": "Bilinmeyen işlem."}],
            },
        }
    )
    generated_image = runtime._approval_request_payload(
        {
            "assistantMessage": "Görsel üretilecek.",
            "planPreview": {
                "summary": "Yeni görsel üretilecek.",
                "steps": [{"capability": "image_generate", "description": "Görsel üret."}],
            },
        }
    )
    generated_chart = runtime._approval_request_payload(
        {
            "assistantMessage": "Grafik üretilecek.",
            "planPreview": {
                "summary": "Yeni grafik üretilecek.",
                "steps": [{"capability": "chart_generate", "description": "Grafik üret."}],
            },
        }
    )
    browser_control = runtime._approval_request_payload(
        {
            "assistantMessage": "Tarayıcı kontrol edilecek.",
            "planPreview": {
                "summary": "Tarayıcı kontrol edilecek.",
                "steps": [{"capability": "browser_control", "description": "Tarayıcıyı kontrol et."}],
            },
        }
    )
    destructive_overwrite = runtime._approval_request_payload(
        {
            "assistantMessage": "Belge güncellenecek.",
            "planPreview": {
                "summary": "Mevcut belgenin üzerine yazılacak.",
                "steps": [{
                    "capability": "document_write",
                    "description": "Mevcut belgeyi değiştir.",
                    "args": {"outputPath": "notes.docx", "overwrite": True},
                }],
            },
        }
    )

    assert safe["permission"] == "write"
    assert safe["idempotency"] == "idempotent_write"
    assert safe["capability"] == "document_write"
    assert unknown["permission"] == "side_effect"
    assert unknown["idempotency"] == "non_idempotent"
    assert generated_image["permission"] == "side_effect"
    assert generated_image["idempotency"] == "non_idempotent"
    assert generated_chart["permission"] == "side_effect"
    assert generated_chart["idempotency"] == "non_idempotent"
    assert browser_control["permission"] == "side_effect"
    assert browser_control["idempotency"] == "non_idempotent"
    assert destructive_overwrite["permission"] == "side_effect"
    assert destructive_overwrite["idempotency"] == "non_idempotent"
    assert destructive_overwrite["steps"][0]["overwrite"] is True


def test_remote_research_email_send_uses_backend_route_decision_without_replanning(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    calls: list[tuple[str, dict]] = []

    def fake_run_capability(capability: str, args: dict, _state: dict) -> dict:
        calls.append((capability, dict(args)))
        if capability == "web_research":
            return {
                "ok": True,
                "tool": capability,
                "output": "Web araştırması tamamlandı.",
                "result": {
                    "kind": "web_research",
                    "summary": "Atatürk hakkında güvenilir kaynak özeti.",
                    "sources": [{"title": "Kaynak", "url": "https://example.com"}],
                },
                "artifacts": [],
                "error": None,
            }
        if capability == "email_draft":
            return {
                "ok": True,
                "tool": capability,
                "output": "E-posta taslağı hazırlandı.",
                "result": {
                    "kind": "email_draft",
                    "to": ["ali@example.com"],
                    "subject": "Atatürk hakkında notlar",
                    "body": "Merhaba,\n\nAtatürk hakkında güvenilir kaynak özeti.\n\nİyi çalışmalar.",
                },
                "artifacts": [],
                "error": None,
            }
        raise AssertionError(f"unexpected capability before approval: {capability}")

    class FakeBackend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []
            self.heartbeats: list[dict] = []

        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        _trusted_task({
                            "id": "task-email-send",
                            "title": "Atatürk araştırması",
                            "status": "queued",
                            "requestedCapabilities": ["web_research", "email_draft", "email_send"],
                            "payload": {
                                "prompt": "Atatürk hakkında araştırma yap ve ali@example.com adresine mail gönder",
                                "metadata": {
                                    "routeDecision": {
                                        "route": "desktop_runtime",
                                        "mode": "mixed_task",
                                        "capabilities": ["web_research", "email_draft", "email_send"],
                                        "privacyClass": "side_effect",
                                        "requiresApproval": True,
                                        "reason": "desktop task",
                                    }
                                },
                            },
                        }, capabilities=["web_research", "email_draft", "email_send"])
                    ]
                },
            )

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

        def heartbeat(self, payload: dict) -> BackendResult:
            self.heartbeats.append(payload)
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(bridge, "run_capability", fake_run_capability)
    monkeypatch.setattr(
        bridge,
        "route_text_to_tool",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("routeDecision should skip local replanning")),
    )
    monkeypatch.setattr(
        runtime,
        "send_conversation",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("remote task should not replan")),
    )
    _arm_device_identity()

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"][0]["status"] == "waiting_approval"
    assert [name for name, _args in calls] == ["web_research", "email_draft"]
    # Canlı adım-adım ilerleme ara 'running' güncellemeleri enjekte edebilir;
    # sözleşme, ilk (yol açan) running'i ve terminal waiting_approval'ı içermeleridir.
    statuses = [payload["status"] for _, payload in runtime.backend.status_updates]  # type: ignore[attr-defined]
    assert statuses[0] == "running"
    assert statuses[-1] == "waiting_approval"
    waiting = runtime.backend.status_updates[-1][1]  # type: ignore[attr-defined]
    assert waiting["status"] == "waiting_approval"
    assert waiting["approvalRequest"]["kind"] == "email_send"
    assert waiting["approvalRequest"]["to"] == ["ali@example.com"]
    assert waiting["approvalRequest"]["subject"] == "Atatürk hakkında notlar"
    assert waiting["approvalRequest"]["bodyPreview"].startswith("Merhaba,")
    link = state_store.get_remote_task_link("task-email-send")
    assert link is not None
    plan = state_store.get_pending_plan(link["pendingPlanId"])
    assert plan is not None
    assert plan["steps"][0]["capability"] == "email_send"
    assert plan["steps"][0]["args"]["body"].startswith("Merhaba,")


def test_execute_assigned_runtime_task_uses_explicit_route_steps_without_replanning(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    calls: list[tuple[str, dict]] = []

    def fake_run_capability(capability: str, args: dict, _state: dict) -> dict:
        calls.append((capability, dict(args)))
        return {
            "ok": True,
            "tool": capability,
            "output": f"{capability}:ok",
            "result": {"kind": capability, "receivedArgs": dict(args)},
            "artifacts": [],
            "error": None,
        }

    class FakeBackend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []
            self.heartbeats: list[dict] = []

        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        _trusted_task({
                            "id": "task-explicit-route",
                            "title": "Desktop planlı görev",
                            "status": "queued",
                            "payload": {
                                "prompt": "Cihaz durumunu ve çalışma bağlamını kontrol et",
                                "metadata": {
                                    "routeDecision": {
                                        "route": "desktop_runtime",
                                        "mode": "mixed_task",
                                        "privacyClass": "local_private",
                                        "reason": "Hazırlık ve cihaz kontrolü desktop üzerinde yapılacak.",
                                        "steps": [
                                            {
                                                "capability": "retrieve_context",
                                                "args": {"query": "son görev bağlamı"},
                                                "description": "Önce yerel bağlam kontrol edilecek.",
                                            },
                                            {
                                                "capability": "sys_info",
                                                "args": {"query": "battery cpu memory"},
                                                "description": "Ardından cihaz durumu okunacak.",
                                            },
                                        ],
                                    }
                                },
                            },
                        }, capabilities=["retrieve_context", "sys_info"], steps=[
                            {
                                "id": "step_1",
                                "capability": "retrieve_context",
                                "description": "Önce yerel bağlam kontrol edilecek.",
                                "args": {"query": "son görev bağlamı"},
                            },
                            {
                                "id": "step_2",
                                "capability": "sys_info",
                                "description": "Ardından cihaz durumu okunacak.",
                                "args": {"query": "battery cpu memory"},
                            },
                        ])
                    ]
                },
            )

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

        def heartbeat(self, payload: dict) -> BackendResult:
            self.heartbeats.append(payload)
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(bridge, "run_capability", fake_run_capability)
    monkeypatch.setattr(
        bridge,
        "route_text_to_tool",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("explicit route steps should skip local replanning")),
    )
    monkeypatch.setattr(
        runtime,
        "send_conversation",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("explicit route steps should not fall back to chat")),
    )
    _arm_device_identity()

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"][0]["status"] == "completed"
    # P2 scheduler: bağımsız read adımları sınırlı paralel koşabilir; çağrı
    # sırası değil, kümenin tamamının yürütülmesi sözleşmedir.
    assert sorted(name for name, _args in calls) == ["retrieve_context", "sys_info"]
    running = runtime.backend.status_updates[0][1]  # type: ignore[attr-defined]
    completed = runtime.backend.status_updates[-1][1]  # type: ignore[attr-defined]
    assert running["planPreview"]["agentPlan"]["stepCount"] == 2
    assert running["result"]["executionTrace"]["status"] == "running"
    assert completed["result"]["planPreview"]["agentPlan"]["executionStrategy"] == "balanced"
    assert sorted(completed["result"]["planPreview"]["agentPlan"]["capabilities"]) == ["retrieve_context", "sys_info"]
    assert completed["result"]["executionTrace"]["status"] == "completed"
    inbox_item = state_store.get_task_inbox_item("task-explicit-route")
    assert inbox_item is not None
    assert inbox_item["executionTrace"]["status"] == "completed"


def test_typed_work_order_executes_without_redundant_route_decision(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    prompt = "Hesap Makinesi uygulamasını aç ve açıldığını doğrula."
    payload = {
        "prompt": "Masaüstü cowork görevi — Hesap Makinesi'ni aç",
        "metadata": {
            "desktopDispatch": True,
            "desktopFullAuthorityEnabled": True,
        },
        "desktopWorkOrder": {
            "schema": "elyan.desktop_work_order.v1",
            "source": "mobile_chat_dispatch",
            "goal": {
                "kind": "computer_task",
                "summary": "Masaüstü cowork görevi — Hesap Makinesi'ni aç",
                "language": "tr",
                "sourceTextHash": "a" * 24,
            },
            "entities": [{"type": "topic", "value": prompt}],
            "constraints": [],
            "requiredCapabilities": ["open_app"],
            "localContextNeeded": [],
            "expectedOutputs": [{"kind": "chat_result", "format": "elyan_blocks.v2", "required": True}],
            "verificationRules": [
                {"id": "runtime_completed", "description": "Runtime completes.", "evidence": "runtime_status"},
                {"id": "tool_result", "description": "Tool succeeds.", "evidence": "tool_result"},
            ],
            "execution": {"mode": "cowork_dispatch", "approvalPolicy": "capability_policy", "maxSteps": 8},
            "planPreview": {
                "summary": "Hesap Makinesi açılacak.",
                "privacyClass": "local_private",
                "steps": [
                    {
                        "id": "step_open_app",
                        "capability": "open_app",
                        "description": "Hesap Makinesi açılacak.",
                        "args": {"app_name": "Hesap Makinesi"},
                    }
                ],
            },
        },
    }
    task = {
        "id": "task-work-order-only",
        "title": "Masaüstü cowork görevi",
        "status": "queued",
        "requestedCapabilities": ["open_app"],
        "payload": payload,
    }

    preview = runtime._remote_task_running_plan_preview(task, prompt, payload)
    result = runtime._execute_deterministic_remote_task(task, prompt, task["title"])

    assert preview["steps"][0]["capability"] == "open_app"
    assert result is not None
    assert result["needsConfirmation"] is True
    pending = state_store.get_pending_plan(result["pendingPlanId"])
    assert pending is not None
    assert pending["query"] == prompt
    assert pending["steps"][0]["capability"] == "open_app"
    assert pending["steps"][0]["args"]["app_name"] == "Hesap Makinesi"


def test_dispatch_prefers_local_deterministic_route_over_llm_planner(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Bariz tek-adım iş ("masaüstündeki dosyaları listele") backend geniş
    capability önerse ve server_brain hazır olsa bile LLM planlayıcıya
    DELEGE EDİLMEZ — yerel yüksek-güven rota (directory_tree) kazanır.
    (Canlı arıza: LLM operator.run seçti → doğrulama hatası → onay çıkmazı.)"""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(bridge, "_server_brain_ready", lambda _state: True)
    monkeypatch.setattr(
        runtime,
        "send_conversation",
        lambda *_a, **_k: (_ for _ in ()).throw(AssertionError("yerel rota varken LLM'e delege edilmemeli")),
    )
    prompt = "Masaüstündeki dosyaları listele"
    task = {
        "id": "task-local-route-wins",
        "title": "Masaüstü cowork görevi",
        "status": "queued",
        "payload": {
            "prompt": prompt,
            "metadata": {
                    "routeDecision": {
                        "route": "desktop_runtime",
                        "capabilities": ["desktop_operator.run", "directory_tree"],
                        "reason": "Masaüstü görevi.",
                }
            },
        },
    }

    result = runtime._execute_deterministic_remote_task(task, prompt, task["title"])

    assert result is not None, "deterministik yol LLM'e düşmemeli"
    dumped = json.dumps(result, ensure_ascii=False, default=str)
    assert "directory_tree" in dumped
    assert "desktop_operator" not in dumped


def test_dispatch_generic_explicit_operator_fallback_yields_to_safe_local_route(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Gerçek mobil dispatch zarfındaki jenerik tek operator adımı, bariz
    ve onaysız yerel directory_tree rotasını ezmemeli."""
    _isolate_state(monkeypatch, tmp_path)
    _arm_device_identity()
    (tmp_path / "rapor.txt").write_text("yerel", encoding="utf-8")
    monkeypatch.setattr(task_router, "_resolve_location_path", lambda _text: str(tmp_path))
    runtime = bridge.RuntimeBridge()
    prompt = "Masaüstündeki dosyaları listele"
    task = _trusted_task({
        "id": "task-generic-explicit-operator",
        "title": prompt,
        "status": "queued",
        "payload": {
            "prompt": prompt,
            "metadata": {
                "desktopDispatch": True,
                "chatSurface": "mobile",
                "routeDecision": {
                    "route": "desktop_runtime",
                    "capabilities": ["desktop_operator.run"],
                    "reason": "Kullanıcı dispatch butonu ile bu görevi masaüstüne yönlendirdi.",
                    "planPreview": {
                        "summary": "Kullanıcı dispatch butonu ile bu görevi masaüstüne yönlendirdi.",
                        "steps": [
                            {
                                "id": "step_1",
                                "capability": "desktop_operator.run",
                                "description": "Masaüstü görevi yürütülecek.",
                                "args": {},
                            }
                        ],
                    },
                },
            },
        },
    }, capabilities=["desktop_operator.run", "directory_tree"], steps=[
        {
            "id": "step_1",
            "capability": "desktop_operator.run",
            "description": "Masaüstü görevi yürütülecek.",
            "args": {},
        }
    ])

    # Doğrudan çağrı P0 güven bağlamını kendisi kurmaz; üretimdeki gibi
    # cihazda mühürlenmiş v2 iş emriyle (salt-okunur ikame kapsamda) sar.
    from runtime.desktop_work_order import validate_payload as _validate_payload
    from runtime.execution_trust import prepare_work_order_v2 as _prepare_v2

    raw_order = _validate_payload(task["payload"]).work_order
    sealed_order = _prepare_v2(task, raw_order, prompt=prompt, state=state_store.snapshot())
    trust_token = runtime._begin_trusted_work_order(sealed_order, "dispatch")
    try:
        result = runtime._execute_deterministic_remote_task(task, prompt, task["title"])
    finally:
        runtime._end_trusted_work_order(trust_token)

    assert result is not None
    assert result["chatOk"] is True
    assert result["needsConfirmation"] is False
    dumped = json.dumps(result, ensure_ascii=False, default=str)
    assert "directory_tree" in dumped
    assert "rapor.txt" in dumped
    assert "desktop_operator.run" not in dumped


def test_assigned_dispatch_generic_operator_completes_via_directory_tree_lifecycle(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Canlı backend zarfının tam runner yolu: jenerik explicit operator
    queued/running sonrası onaysız directory_tree ile completed olmalı."""
    _isolate_state(monkeypatch, tmp_path)
    _arm_device_identity()
    (tmp_path / "dispatch-proof.txt").write_text("ok", encoding="utf-8")
    monkeypatch.setattr(task_router, "_resolve_location_path", lambda _text: str(tmp_path))
    prompt = "Masaüstündeki dosyaları listele"
    task = _trusted_task({
        "id": "task-generic-operator-lifecycle",
        "title": prompt,
        "status": "queued",
        "payload": {
            "prompt": prompt,
            "metadata": {
                "desktopDispatch": True,
                "chatSurface": "mobile",
                "routeDecision": {
                    "route": "desktop_runtime",
                    "capabilities": ["desktop_operator.run"],
                    "reason": "Kullanıcı dispatch butonu ile bu görevi masaüstüne yönlendirdi.",
                    "planPreview": {
                        "summary": "Kullanıcı dispatch butonu ile bu görevi masaüstüne yönlendirdi.",
                        "steps": [
                            {
                                "id": "step_1",
                                "capability": "desktop_operator.run",
                                "description": "Masaüstü görevi yürütülecek.",
                                "args": {},
                            }
                        ],
                    },
                },
            },
        },
    }, capabilities=["desktop_operator.run", "directory_tree"], steps=[
        {
            "id": "step_1",
            "capability": "desktop_operator.run",
            "description": "Masaüstü görevi yürütülecek.",
            "args": {},
        }
    ])

    class _Backend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []
            self.heartbeats: list[dict] = []

        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_dispatch_lifecycle",
                status_code=200,
                data={"tasks": [task]},
            )

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

        def heartbeat(self, payload: dict) -> BackendResult:
            self.heartbeats.append(payload)
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    runtime = bridge.RuntimeBridge()
    runtime.backend = _Backend()  # type: ignore[assignment]
    monkeypatch.setattr(
        runtime,
        "send_conversation",
        lambda *_a, **_k: (_ for _ in ()).throw(AssertionError("güvenli yerel rota LLM'e düşmemeli")),
    )

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"][0]["status"] == "completed"
    statuses = [payload["status"] for _, payload in runtime.backend.status_updates]  # type: ignore[attr-defined]
    assert statuses[0] == "running"
    assert statuses[-1] == "completed"
    assert "waiting_approval" not in statuses
    dumped = json.dumps(runtime.backend.status_updates, ensure_ascii=False, default=str)  # type: ignore[attr-defined]
    assert "directory_tree" in dumped
    assert "dispatch-proof.txt" in dumped
    assert "desktop_operator.run" not in dumped


def test_dispatch_keeps_specific_explicit_operator_step(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Spesifik GUI hedefi/argümanı taşıyan gerçek explicit operator planı
    jenerik fallback sayılmamalı."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    prompt = "Masaüstündeki dosyaları listele"
    route_decision = {
        "route": "desktop_runtime",
        "capabilities": ["desktop_operator.run"],
        "planPreview": {
            "summary": "Finder içinde belirli GUI hedefi kullanılacak.",
            "steps": [
                {
                    "capability": "desktop_operator.run",
                    "args": {
                        "goal": "Finder penceresinde Sıralama menüsünü aç.",
                        "action": "run",
                        "targetApp": "Finder",
                    },
                    "description": "Finder GUI hedefi yürütülecek.",
                }
            ],
        },
    }
    task = {
        "id": "task-specific-explicit-operator",
        "title": prompt,
        "payload": {"prompt": prompt, "metadata": {"routeDecision": route_decision}},
    }

    steps, _preview = runtime._remote_task_steps_from_route(
        task,
        prompt,
        {"desktop_operator.run"},
        route_decision,
    )

    assert [step["capability"] for step in steps] == ["desktop_operator.run"]
    assert steps[0]["args"]["targetApp"] == "Finder"


def test_generic_operator_joker_yields_to_safe_local_route(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Backend'in jenerik operator jokerini yüksek-güvenli yerel rota ezer.

    Eski sözleşme jokerin aynen yürümesini bekliyordu; canlı arıza bunun tam
    tersini kanıtladı: "masaüstündeki dosyaları listele" operator'a kör hedef
    olarak gidip doğrulama+onay çıkmazı üretiyordu. Salt-okunur/zararsız yerel
    eşleşme (directory_tree) artık kazanır (_SAFE_LOCAL_OVERRIDE)."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    prompt = "Masaüstündeki dosyaları listele"
    decision = {
        "route": "desktop_runtime",
        "capabilities": ["desktop_operator.run"],
        "planPreview": {
            "summary": prompt,
            "steps": [
                {
                    "id": "step_1",
                    "capability": "desktop_operator.run",
                    "args": {},
                    "description": prompt,
                }
            ],
        },
    }
    task = {
        "id": "scope-bound",
        "payload": {"prompt": prompt, "metadata": {"routeDecision": decision}},
    }

    steps, _preview = runtime._remote_task_steps_from_route(
        task,
        prompt,
        {"desktop_operator.run"},
        decision,
    )

    assert [step["capability"] for step in steps] == ["directory_tree"]


def test_screen_glance_dispatch_overrides_operator_joker(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Canlı arıza: "Ekranda ne var" operator jokerine gidip 'Operator
    doğrulaması başarısız oldu.' üretiyordu — analyze_screen kazanmalı."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    prompt = "Ekranda ne var"
    decision = {
        "route": "desktop_runtime",
        "capabilities": ["desktop_operator.run"],
        "planPreview": {
            "summary": prompt,
            "steps": [
                {"id": "step_1", "capability": "desktop_operator.run", "args": {"goal": prompt}, "description": prompt}
            ],
        },
    }
    task = {"id": "glance", "payload": {"prompt": prompt, "metadata": {"routeDecision": decision}}}

    steps, _preview = runtime._remote_task_steps_from_route(
        task, prompt, {"desktop_operator.run"}, decision
    )
    assert [step["capability"] for step in steps] == ["analyze_screen"]


def test_new_tab_dispatch_overrides_blind_browser_search(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Canlı arıza: "Chrome dan yeni sekme aç" backend'in kör 'search' adımıyla
    Google'da 'yeni sekme' aratıyordu — action=new_tab kazanmalı."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    prompt = "Chrome dan yeni sekme aç"
    decision = {
        "route": "desktop_runtime",
        "capabilities": ["browser_control"],
        "planPreview": {
            "summary": prompt,
            "steps": [
                {"id": "step_1", "capability": "browser_control", "args": {"action": "search", "query": "yeni sekme"}, "description": prompt}
            ],
        },
    }
    task = {"id": "newtab", "payload": {"prompt": prompt, "metadata": {"routeDecision": decision}}}

    steps, _preview = runtime._remote_task_steps_from_route(
        task, prompt, {"browser_control"}, decision
    )
    assert len(steps) == 1
    assert steps[0]["capability"] == "browser_control"
    assert steps[0]["args"]["action"] == "new_tab"


def test_remote_browser_agent_route_requires_approval(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Browser agent tıklama/yazma/indirme yapabildiği için mobil dispatch
    planında doğrudan çalışmamalı; waiting-approval planı üretmeli."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    prompt = "İnternetten Example Domain başlığını bul ve söyle"
    task = {
        "id": "task-browser-agent-approval",
        "title": prompt,
        "payload": {
            "prompt": prompt,
            "metadata": {
                "desktopDispatch": True,
                "routeDecision": {
                    "route": "desktop_runtime",
                    "capabilities": ["browser_control"],
                    "reason": "Tarayıcı görevi.",
                },
            },
        },
    }

    result = runtime._execute_deterministic_remote_task(task, prompt, task["title"])

    assert result is not None
    assert result["needsConfirmation"] is True
    pending = state_store.get_pending_plan(result["pendingPlanId"])
    assert pending is not None
    assert [step["capability"] for step in pending["steps"]] == ["browser_agent.run"]


def test_browser_snapshot_never_reads_form_values() -> None:
    scripts: list[str] = []

    class _Page:
        url = "https://example.test/account"

        @staticmethod
        def title() -> str:
            return "Hesap"

        @staticmethod
        def evaluate(script: str, _limit: int) -> list:
            scripts.append(script)
            return []

    result = browser_session._SessionThread()._snapshot(_Page(), {"limit": 10})

    assert result["elements"] == []
    assert scripts and "el.value" not in scripts[0]
    assert "isFormControl" in scripts[0]


def test_browser_session_blocks_sensitive_fields_and_irreversible_clicks() -> None:
    class _Locator:
        def __init__(self, attrs: dict[str, str], text: str = "") -> None:
            self.attrs = attrs
            self.text = text
            self.filled = False
            self.clicked = False

        def get_attribute(self, name: str) -> str:
            return self.attrs.get(name, "")

        @property
        def first(self) -> "_Locator":
            return self

        def inner_text(self, timeout: int = 0) -> str:
            return self.text

        def fill(self, _value: str, timeout: int = 0) -> None:
            self.filled = True

        def click(self, timeout: int = 0) -> None:
            self.clicked = True

    class _Page:
        def __init__(self, locator: _Locator) -> None:
            self._target = locator

        def locator(self, _selector: str) -> _Locator:
            return self._target

    session = browser_session._SessionThread()
    card = _Locator({"type": "tel", "autocomplete": "cc-number"})
    with pytest.raises(SafeCapabilityError) as field_error:
        session._type(_Page(card), {"selector": "#card", "value": "4111111111111111"})
    assert field_error.value.code == "SENSITIVE_FIELD_BLOCKED"
    assert card.filled is False

    purchase = _Locator({"type": "button"}, "Satın al")
    with pytest.raises(SafeCapabilityError) as click_error:
        session._click(_Page(purchase), {"selector": "#purchase"})
    assert click_error.value.code == "SENSITIVE_ACTION_BLOCKED"
    assert purchase.clicked is False


def test_remote_task_runner_adds_canonical_run_payload(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []

        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        _trusted_task({
                            "id": "task-canonical",
                            "title": "Canonical payload",
                            "status": "queued",
                            "payload": {
                                "prompt": "Cihaz durumunu kontrol et",
                                "metadata": {
                                    "routeDecision": {
                                        "route": "desktop_runtime",
                                        "steps": [
                                            {
                                                "capability": "sys_info",
                                                "args": {"query": "battery"},
                                                "description": "Cihaz durumu okunacak.",
                                            }
                                        ],
                                    }
                                },
                            },
                        }, capabilities=["sys_info"], steps=[
                            {
                                "id": "step_1",
                                "capability": "sys_info",
                                "description": "Cihaz durumu okunacak.",
                                "args": {"query": "battery"},
                            }
                        ])
                    ]
                },
            )

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

        def heartbeat(self, payload: dict) -> BackendResult:
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    def fake_run_capability(capability: str, args: dict, _state: dict) -> dict:
        return {
            "ok": True,
            "tool": capability,
            "output": "sys_info:ok",
            "result": {"kind": capability, "receivedArgs": dict(args)},
            "artifacts": [],
            "error": None,
        }

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(bridge, "run_capability", fake_run_capability)
    _arm_device_identity()

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"][0]["status"] == "completed"
    running = runtime.backend.status_updates[0][1]  # type: ignore[attr-defined]
    completed = runtime.backend.status_updates[-1][1]  # type: ignore[attr-defined]
    assert running["result"]["runnerVersion"] == "remote_task_orchestrator_v1"
    assert running["result"]["taskRunId"].startswith("run_")
    assert running["result"]["capabilityReadiness"][0]["capability"] == "sys_info"
    assert completed["result"]["taskRunId"] == running["result"]["taskRunId"]
    assert completed["result"]["executionTrace"]["runnerVersion"] == "remote_task_orchestrator_v1"
    inbox_item = state_store.get_task_inbox_item("task-canonical")
    assert inbox_item is not None
    assert inbox_item["taskRunId"] == running["result"]["taskRunId"]
    assert inbox_item["capabilityReadiness"][0]["capability"] == "sys_info"


def test_remote_task_runner_readiness_failure_reports_safe_terminal(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    from runtime import remote_task_runner

    class FakeBackend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []

        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        _trusted_task({
                            "id": "task-readiness-fail",
                            "title": "Readiness fail",
                            "status": "queued",
                            "payload": {
                                "prompt": "Belge hazırla",
                                "metadata": {
                                    "routeDecision": {
                                        "route": "desktop_runtime",
                                        "steps": [
                                            {
                                                "capability": "document_write",
                                                "args": {"title": "Rapor"},
                                                "description": "DOCX üretilecek.",
                                            }
                                        ],
                                    }
                                },
                            },
                        }, capabilities=["document_write"], steps=[
                            {
                                "id": "step_1",
                                "capability": "document_write",
                                "description": "DOCX üretilecek.",
                                "args": {"title": "Rapor"},
                            }
                        ])
                    ]
                },
            )

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(
        remote_task_runner,
        "capability_readiness",
        lambda name, **_kwargs: {
            "ready": False,
            "available": False,
            "dependencyReady": False,
            "platformSupported": True,
            "permissionClass": "approval_required",
            "errorCode": "DEPENDENCY_UNAVAILABLE",
            "missingDependencies": ["python-docx"],
            "degradationReason": "dependency_unavailable",
        },
    )
    monkeypatch.setattr(
        bridge,
        "run_capability",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("readiness failure must not execute")),
    )
    _arm_device_identity()

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"][0]["status"] == "failed"
    terminal = runtime.backend.status_updates[-1][1]  # type: ignore[attr-defined]
    assert terminal["status"] == "failed"
    assert terminal["error"] == "DEPENDENCY_UNAVAILABLE"
    assert terminal["result"]["runnerVersion"] == "remote_task_orchestrator_v1"
    assert terminal["result"]["capabilityReadiness"][0]["missingDependencies"] == ["python-docx"]


def test_explicit_route_side_effect_waits_for_approval_before_execution(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    # Doğrudan mod bile olsa geri-alınamaz/dışa dönük (blocklist) adım
    # (email_send) açık onay ister; onaysız yürütülmez.
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []
            self.heartbeats: list[dict] = []

        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        _trusted_task({
                            "id": "task-write-approval",
                            "title": "Mail gönder",
                            "status": "queued",
                            "payload": {
                                "prompt": "ali@example.com'a toplantı notlarını mail at",
                                "metadata": {
                                    "routeDecision": {
                                        "route": "desktop_runtime",
                                        "mode": "mixed_task",
                                        "privacyClass": "local_private",
                                        "reason": "Mail gönderimi açık onay gerektirir.",
                                        "steps": [
                                            {
                                                "capability": "email_send",
                                                "args": {"to": "ali@example.com", "subject": "Toplantı notları", "body": "Notlar ekte."},
                                                "description": "E-posta gönderilecek.",
                                            }
                                        ],
                                    }
                                },
                            },
                        }, capabilities=["email_send"], steps=[
                            {
                                "id": "step_1",
                                "capability": "email_send",
                                "description": "E-posta gönderilecek.",
                                "args": {"to": "ali@example.com", "subject": "Toplantı notları", "body": "Notlar ekte."},
                            }
                        ])
                    ]
                },
            )

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

        def heartbeat(self, payload: dict) -> BackendResult:
            self.heartbeats.append(payload)
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(
        bridge,
        "run_capability",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("side-effect step must wait for approval")),
    )
    _arm_device_identity()

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"][0]["status"] == "waiting_approval"
    waiting = runtime.backend.status_updates[-1][1]  # type: ignore[attr-defined]
    assert waiting["status"] == "waiting_approval"
    assert waiting["approvalRequest"]["kind"] == "email_send"
    assert waiting["result"]["executionTrace"]["status"] == "waiting_approval"
    link = state_store.get_remote_task_link("task-write-approval")
    assert link is not None
    plan = state_store.get_pending_plan(link["pendingPlanId"])
    assert plan is not None
    assert plan["steps"][0]["capability"] == "email_send"


def test_canonical_capability_preserves_dotted_desktop_operator_names() -> None:
    assert bridge._canonical_capability_name("desktop_operator.run") == "desktop_operator.run"
    assert bridge._canonical_capability_name("desktop.operator.execute_action") == "desktop_operator.execute_action"


def test_remote_task_trace_keeps_stable_card_and_inline_approval_step() -> None:
    runtime = bridge.RuntimeBridge()
    trace = runtime._remote_task_trace_payload(
        {
            "summary": "Toplantı ve e-posta",
            "steps": [
                {"id": "step_1", "capability": "calendar_create", "description": "Toplantıyı oluştur"},
                {"id": "step_2", "capability": "email_send", "description": "E-postayı gönder"},
            ],
        },
        status="waiting_approval",
        task_id="task-card-1",
    )

    assert trace["stableBlockId"] == "task_trace_task-card-1"
    assert any(step["status"] == "waiting_approval" for step in trace["steps"])


def test_execute_assigned_runtime_task_skips_duplicate_inflight_delivery(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        {
                            "id": "task-dup",
                            "title": "Tek görev",
                            "status": "queued",
                            "payload": {"prompt": "merhaba"},
                        }
                    ]
                },
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    runtime._assigned_task_inflight.add("task-dup")

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"] == [{"taskId": "task-dup", "ok": True, "status": "skipped_duplicate"}]


def test_execute_assigned_runtime_task_skips_recent_terminal_replay(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        {
                            "id": "task-terminal",
                            "title": "Tekrar gelen görev",
                            "status": "queued",
                            "payload": {"prompt": "merhaba"},
                        }
                    ]
                },
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    runtime._remember_terminal_assigned_task("task-terminal")
    monkeypatch.setattr(
        runtime,
        "_execute_runtime_task",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("terminal replay should not execute")),
    )

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"] == [
        {"taskId": "task-terminal", "ok": True, "status": "skipped_recent_terminal"}
    ]


def test_terminal_task_report_is_remembered_for_replay_suppression(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    monkeypatch.setattr(runtime, "_send_runtime_socket_message", lambda _payload: False)
    monkeypatch.setattr(
        runtime.backend,
        "runtime_task_status",
        lambda task_id, payload: bridge.BackendResult(
            ok=True,
            request_id="req_status",
            status_code=200,
            data={"ok": True, "taskId": task_id},
        ),
    )
    monkeypatch.setattr(runtime, "_resync_terminal_remote_task", lambda _task_id: None)

    report = runtime._report_runtime_task_status(
        "task-terminal",
        {"status": "completed", "summary": "Hazır."},
    )

    assert report is not None
    assert report.ok is True
    assert runtime._is_recent_terminal_assigned_task("task-terminal") is True


def test_execute_assigned_runtime_tasks_reconciles_stale_active_item_to_unknown(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.upsert_task_inbox_item(
        {
            "id": "task-stale",
            "title": "Eski görev",
            "status": "running",
            "summary": "devam ediyor",
        }
    )

    class FakeBackend:
        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={"tasks": []},
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    item = state_store.get_task_inbox_item("task-stale")
    assert item is not None
    assert item["status"] == "unknown"
    assert item["lastVerifiedAt"]


def test_execute_assigned_runtime_tasks_preserves_waiting_approval_link_on_reconcile(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.upsert_task_inbox_item(
        {
            "id": "task-wait",
            "title": "Onay bekleyen görev",
            "status": "waiting_approval",
            "summary": "onay gerekiyor",
        }
    )
    state_store.save_remote_task_link(
        "task-wait",
        "plan-1",
        "conv-1",
        status="waiting_approval",
    )

    class FakeBackend:
        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={"tasks": []},
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    item = state_store.get_task_inbox_item("task-wait")
    assert item is not None
    assert item["status"] == "waiting_approval"
    link = state_store.get_remote_task_link("task-wait")
    assert link is not None
    assert link["pendingPlanId"] == "plan-1"


def test_runtime_websocket_mode_uses_poll_only_for_fallback_recovery(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    runtime._runtime_ws_connected = True
    runtime._last_assigned_task_fetch_at = time.monotonic()
    monkeypatch.setattr(runtime, "_assigned_task_poll_fallback_seconds", lambda: 12.0)

    assert runtime._should_poll_assigned_tasks() is False

    runtime._request_assigned_task_fetch()

    assert runtime._should_poll_assigned_tasks() is True


def test_runtime_websocket_mode_periodically_polls_assigned_tasks_as_delivery_fallback(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    runtime._runtime_ws_connected = True
    monkeypatch.setattr(runtime, "_assigned_task_poll_fallback_seconds", lambda: 12.0)

    runtime._last_assigned_task_fetch_at = time.monotonic() - 11.0

    assert runtime._should_poll_assigned_tasks() is False

    runtime._last_assigned_task_fetch_at = time.monotonic() - 12.1

    assert runtime._should_poll_assigned_tasks() is True


def test_remote_task_explicit_computer_control_maps_to_desktop_operator(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    steps, preview = runtime._remote_task_explicit_steps_from_route(
        {"id": "task-computer-control", "payload": {}},
        "Safariyi aç",
        {
            "route": "desktop_runtime",
            "planPreview": {
                "steps": [
                    {
                        "capability": "computer_control",
                        "args": {"prompt": "Safariyi aç"},
                        "description": "Bilgisayar kontrolü çalışacak.",
                    }
                ]
            },
        },
    )

    assert steps
    assert steps[0]["capability"] == "desktop_operator.run"
    assert steps[0]["args"]["goal"] == "Safariyi aç"
    assert preview["agentPlan"]["capabilities"] == ["desktop_operator.run"]


def test_execute_assigned_runtime_task_fails_closed_on_capability_mismatch(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "capabilities": ["runtime.status"],
            }
        }
    )

    class FakeBackend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []

        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        _trusted_task({
                            "id": "task-cap-mismatch",
                            "title": "Araştırma yap",
                            "targetDeviceId": VALID_DEVICE_ID,
                            "requestedCapabilities": ["web_research"],
                            "payload": {"prompt": "Elyan hakkında araştırma yap"},
                        }, capabilities=["web_research"])
                    ]
                },
            )

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

    fake_backend = FakeBackend()
    runtime = bridge.RuntimeBridge()
    runtime.backend = fake_backend  # type: ignore[assignment]

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"][0]["status"] == "failed_closed"
    assert fake_backend.status_updates[0][0] == "task-cap-mismatch"
    assert fake_backend.status_updates[0][1]["status"] == "failed"
    assert fake_backend.status_updates[0][1]["error"] == "runtime_capability_mismatch"


def test_execute_assigned_runtime_task_applies_rejected_approval_resolution(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    conversation = state_store.create_conversation("")
    plan = state_store.save_pending_plan(
        {
            "id": "plan_reject",
            "conversationId": conversation["id"],
            "query": "mail gönder",
            "intent": "email_send",
            "capability": "email_send",
            "confidence": 0.91,
            "steps": [],
        }
    )
    state_store.save_remote_task_link(
        "task-rejected",
        plan["id"],
        conversation["id"],
        title="Mail görevi",
        status="waiting_approval",
    )

    class FakeBackend:
        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        {
                            "id": "task-rejected",
                            "title": "Mail görevi",
                            "status": "waiting_approval",
                            "approvalRequest": {
                                "kind": "email_send",
                                "resolution": {
                                    "approved": False,
                                    "rejected": True,
                                    "status": "rejected",
                                },
                            },
                        }
                    ]
                },
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"] == [{"taskId": "task-rejected", "ok": True, "status": "canceled"}]
    assert state_store.get_remote_task_link("task-rejected") is None
    assert state_store.get_pending_plan(plan["id"]) is None
    item = state_store.get_task_inbox_item("task-rejected")
    assert item is not None
    assert item["status"] == "canceled"


def test_execute_assigned_runtime_task_canceled_terminal_clears_local_waiting_state(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    conversation = state_store.create_conversation("")
    plan = state_store.save_pending_plan(
        {
            "id": "plan_terminal_cancel",
            "conversationId": conversation["id"],
            "query": "takvim etkinliği oluştur",
            "intent": "add_calendar_event",
            "capability": "add_calendar_event",
            "confidence": 0.82,
            "steps": [],
        }
    )
    state_store.save_remote_task_link(
        "task-terminal-cancel",
        plan["id"],
        conversation["id"],
        title="Takvim görevi",
        status="waiting_approval",
    )

    class FakeBackend:
        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        {
                            "id": "task-terminal-cancel",
                            "title": "Takvim görevi",
                            "status": "canceled",
                            "approvalRequest": {},
                        }
                    ]
                },
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"] == [{"taskId": "task-terminal-cancel", "ok": True, "status": "skipped_terminal"}]
    assert state_store.get_remote_task_link("task-terminal-cancel") is None
    assert state_store.get_pending_plan(plan["id"]) is None
    item = state_store.get_task_inbox_item("task-terminal-cancel")
    assert item is not None
    assert item["status"] == "canceled"


def test_execute_assigned_runtime_task_fails_closed_without_personal_action_permission(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []
            self.heartbeats: list[dict] = []

        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        _trusted_task(
                            {
                                "id": "task-no-permission",
                                "title": "Takvim görevi",
                                "status": "queued",
                                "payload": {"prompt": "takvime cuma 14:00 ürün toplantısı ekle"},
                            },
                            capabilities=["add_calendar_event"],
                        )
                    ]
                },
            )

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

        def heartbeat(self, payload: dict) -> BackendResult:
            self.heartbeats.append(payload)
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    runtime = bridge.RuntimeBridge()
    _arm_device_identity()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"][0]["status"] == "failed"
    statuses = [payload["status"] for _, payload in runtime.backend.status_updates]  # type: ignore[attr-defined]
    assert statuses[0] == "running"
    assert statuses[-1] == "failed"
    assert runtime.backend.status_updates[-1][1]["error"] == "PERMISSION_REQUIRED"  # type: ignore[attr-defined]
    assert state_store.get_remote_task_link("task-no-permission") is None


def test_runtime_task_approval_resume_reports_terminal_result(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []
            self.artifact_updates: list[tuple[str, list[dict]]] = []
            self.heartbeats: list[dict] = []

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

        def runtime_task_artifacts(self, task_id: str, payload: dict) -> BackendResult:
            artifacts = payload.get("artifacts", []) if isinstance(payload, dict) else []
            self.artifact_updates.append((task_id, artifacts))
            return BackendResult(ok=True, request_id="req_artifacts", status_code=200, data={"ok": True})

        def heartbeat(self, payload: dict) -> BackendResult:
            self.heartbeats.append(payload)
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    conversation = state_store.create_conversation("")
    plan = state_store.save_pending_plan(
        {
            "id": "plan_approve",
            "conversationId": conversation["id"],
            "query": "takvime cuma 14:00 ürün toplantısı ekle",
            "intent": "add_calendar_event",
            "capability": "add_calendar_event",
            "confidence": 0.84,
            "steps": [
                {
                    "capability": "add_calendar_event",
                    "args": {"title": "Ürün toplantısı"},
                }
            ],
        }
    )
    state_store.save_remote_task_link(
        "task-approve",
        plan["id"],
        conversation["id"],
        title="Takvim görevi",
    )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    observed_access: list[dict] = []

    def confirm_with_task_access(*_args: object, **_kwargs: object) -> dict:
        observed_access.append(runtime._access_status())
        return {
            "chatOk": True,
            "assistantMessage": "add_calendar_event:Ürün toplantısı",
            "provider": "local_tool",
            "toolEvents": [],
            "conversationId": conversation["id"],
            "structuredResult": {"kind": "add_calendar_event"},
            "artifacts": [],
        }

    monkeypatch.setattr(runtime, "confirm_conversation_plan", confirm_with_task_access)

    result = runtime._resume_remote_task_after_approval("task-approve", True)

    assert result["ok"] is True
    assert [payload["status"] for _, payload in runtime.backend.status_updates] == [  # type: ignore[attr-defined]
        "running",
        "completed",
    ]
    assert runtime.backend.artifact_updates[0][1][0]["kind"] == "summary"  # type: ignore[attr-defined]
    assert state_store.get_remote_task_link("task-approve") is None
    assert state_store.get_pending_plan(plan["id"]) is None
    inbox_item = state_store.get_task_inbox_item("task-approve")
    assert inbox_item is not None
    assert inbox_item["status"] == "completed"
    assert observed_access[0]["fullAccessSession"]["scope"] == "task"
    assert observed_access[0]["fullAccessSession"]["taskId"] == "task-approve"
    assert observed_access[0]["effectivePermissions"]["allow_computer_control"] is True
    assert runtime._access_status()["fullAccessSession"]["enabled"] is False


def test_execute_assigned_runtime_tasks_skips_recent_terminal_duplicate_approved_resume(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []

        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        {
                            "id": "task-approve",
                            "title": "Takvim görevi",
                            "status": "waiting_approval",
                            "approvalRequest": {
                                "resolution": {
                                    "approved": True,
                                    "status": "approved",
                                }
                            },
                        }
                    ]
                },
            )

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

        def runtime_task_artifacts(self, task_id: str, payload: dict) -> BackendResult:
            return BackendResult(ok=True, request_id="req_artifacts", status_code=200, data={"ok": True})

        def heartbeat(self, payload: dict) -> BackendResult:
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    conversation = state_store.create_conversation("")
    plan = state_store.save_pending_plan(
        {
            "id": "plan_approve_duplicate",
            "conversationId": conversation["id"],
            "query": "takvime cuma 14:00 ürün toplantısı ekle",
            "intent": "add_calendar_event",
            "capability": "add_calendar_event",
            "confidence": 0.84,
            "steps": [
                {
                    "capability": "add_calendar_event",
                    "args": {"title": "Ürün toplantısı"},
                }
            ],
        }
    )
    state_store.save_remote_task_link(
        "task-approve",
        plan["id"],
        conversation["id"],
        title="Takvim görevi",
    )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(
        runtime,
        "confirm_conversation_plan",
        lambda *_args, **_kwargs: {
            "chatOk": True,
            "assistantMessage": "add_calendar_event:Ürün toplantısı",
            "provider": "local_tool",
            "toolEvents": [],
            "conversationId": conversation["id"],
            "structuredResult": {"kind": "add_calendar_event"},
            "artifacts": [],
        },
    )

    first = runtime.execute_assigned_runtime_tasks()
    second = runtime.execute_assigned_runtime_tasks()

    assert first["ok"] is True
    assert first["executions"][0]["status"] == "completed"
    assert second["ok"] is True
    assert second["executions"] == [
        {"taskId": "task-approve", "ok": True, "status": "skipped_recent_terminal"}
    ]
    assert [payload["status"] for _, payload in runtime.backend.status_updates] == [  # type: ignore[attr-defined]
        "running",
        "completed",
    ]


def test_runtime_task_approval_missing_link_reports_safe_failure(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    result = runtime._resume_remote_task_after_approval("task-missing-link", True)

    assert result["status"] == "failed"
    assert runtime.backend.status_updates[0][1]["status"] == "failed"  # type: ignore[attr-defined]
    assert runtime.backend.status_updates[0][1]["error"] == "pending_link_missing"  # type: ignore[attr-defined]
    inbox_item = state_store.get_task_inbox_item("task-missing-link")
    assert inbox_item is not None
    assert inbox_item["status"] == "failed"


def test_runtime_task_cancel_cleans_pending_local_plan(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    conversation = state_store.create_conversation("")
    plan = state_store.save_pending_plan(
        {
            "id": "plan_cancelled",
            "conversationId": conversation["id"],
            "query": "takvime cuma 14:00 ürün toplantısı ekle",
            "intent": "add_calendar_event",
            "capability": "add_calendar_event",
            "confidence": 0.84,
            "steps": [],
        }
    )
    state_store.save_remote_task_link(
        "task-cancelled",
        plan["id"],
        conversation["id"],
        title="Takvim görevi",
    )
    runtime = bridge.RuntimeBridge()

    runtime._handle_runtime_ws_message(
        json.dumps({"type": "task.cancel", "taskId": "task-cancelled"})
    )

    assert state_store.get_pending_plan(plan["id"]) is None
    assert state_store.get_remote_task_link("task-cancelled") is None
    active = state_store.get_conversation(conversation["id"])
    assert active is not None
    assert active["messages"][-1]["text"] == "İşlem iptal edildi."


def test_runtime_ws_dispatch_skips_duplicate_without_second_ack(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    ack_payloads: list[dict[str, object]] = []
    thread_starts: list[tuple[object, tuple[object, ...]]] = []

    def fake_send_runtime_socket_message(payload: dict[str, object]) -> bool:
        ack_payloads.append(payload)
        return True

    class FakeThread:
        def __init__(self, target: object, args: tuple[object, ...], name: str, daemon: bool) -> None:
            assert name == "elyan-runtime-task-dispatch"
            assert daemon is True
            thread_starts.append((target, args))

        def start(self) -> None:
            return None

    monkeypatch.setattr(runtime, "_send_runtime_socket_message", fake_send_runtime_socket_message)
    monkeypatch.setattr(bridge.threading, "Thread", FakeThread)

    payload = json.dumps(
        {
            "type": "task.dispatch",
            "task": {"id": "task-ws-dup", "title": "Websocket gorevi"},
            "leaseId": "lease-ws-1",
        }
    )

    runtime._handle_runtime_ws_message(payload)
    runtime._handle_runtime_ws_message(payload)

    assert len(ack_payloads) == 1
    assert ack_payloads[0]["type"] == "task.ack"
    assert ack_payloads[0]["taskId"] == "task-ws-dup"
    assert ack_payloads[0]["leaseId"] == "lease-ws-1"
    assert isinstance(ack_payloads[0]["acceptedAt"], str)
    assert len(thread_starts) == 1


def test_runtime_ws_invalid_message_log_contains_safe_diagnostics(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    runtime._handle_runtime_ws_message("{invalid")

    output = capsys.readouterr().err
    assert "runtime ws_message_invalid" in output
    assert "error=JSONDecodeError" in output
    assert "length=8" in output
    assert "digest=" in output


def test_runtime_ws_accepts_utf8_binary_json_frame(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    runtime._handle_runtime_ws_message(
        b'{"type":"error","message":"binary frame decoded"}'
    )

    assert runtime._runtime_ws_last_error == "binary frame decoded"


def test_handle_normalizes_nested_backend_error_without_crashing(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    class FakeBackend:
        def pairing_create_session(self, _payload: dict[str, object]) -> BackendResult:
            return BackendResult(
                ok=False,
                request_id="req_pairing_unauthorized",
                status_code=401,
                data={"error": {"code": "UNAUTHORIZED", "message": "Giriş gerekli."}},
                error={"code": "UNAUTHORIZED", "message": "Giriş gerekli."},
            )

    runtime.backend = FakeBackend()  # type: ignore[assignment]

    response = runtime.handle(
        {
            "id": "req_pairing",
            "taskId": "task_pairing",
            "capability": "pairing.create_session",
            "payload": {"deviceLabel": "Elyan", "platform": "macos"},
        }
    )

    assert response["ok"] is False
    assert response["error"] == {
        "code": "UNAUTHORIZED",
        "message": "Giriş gerekli.",
    }


def test_runtime_ws_dispatch_persists_local_acceptance_before_ack(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    ack_payloads: list[dict[str, object]] = []
    thread_starts: list[tuple[object, tuple[object, ...]]] = []

    def fake_send_runtime_socket_message(payload: dict[str, object]) -> bool:
        if payload.get("type") == "task.ack":
            link = state_store.get_runtime_dispatch_link("task-ws-persist")
            assert link is not None
            assert link["leaseId"] == "lease-ws-persist"
            assert link["status"] == "accepted"
            assert link["executionState"] == "accepted"
            assert link["dispatchAckAt"] == ""
            assert payload.get("acceptedAt") == link["acceptedAt"]
        ack_payloads.append(payload)
        return True

    class FakeThread:
        def __init__(self, target: object, args: tuple[object, ...], name: str, daemon: bool) -> None:
            assert name == "elyan-runtime-task-dispatch"
            assert daemon is True
            thread_starts.append((target, args))

        def start(self) -> None:
            return None

    monkeypatch.setattr(runtime, "_send_runtime_socket_message", fake_send_runtime_socket_message)
    monkeypatch.setattr(bridge.threading, "Thread", FakeThread)

    runtime._handle_runtime_ws_message(
        json.dumps(
            {
                "type": "task.dispatch",
                "task": {"id": "task-ws-persist", "title": "Kalici kabul"},
                "leaseId": "lease-ws-persist",
            }
        )
    )

    assert len(ack_payloads) == 1
    assert ack_payloads[0]["type"] == "task.ack"
    assert ack_payloads[0]["taskId"] == "task-ws-persist"
    assert ack_payloads[0]["leaseId"] == "lease-ws-persist"
    assert isinstance(ack_payloads[0]["acceptedAt"], str)
    link = state_store.get_runtime_dispatch_link("task-ws-persist")
    assert link is not None
    assert link["status"] == "acked"
    assert link["executionState"] == "acked"
    assert link["dispatchAckAt"] == ack_payloads[0]["acceptedAt"]
    assert len(thread_starts) == 1


def test_runtime_ws_dispatch_skips_recent_terminal_without_ack(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    ack_payloads: list[dict[str, object]] = []
    thread_starts: list[tuple[object, tuple[object, ...]]] = []

    def fake_send_runtime_socket_message(payload: dict[str, object]) -> bool:
        ack_payloads.append(payload)
        return True

    class FakeThread:
        def __init__(self, target: object, args: tuple[object, ...], name: str, daemon: bool) -> None:
            thread_starts.append((target, args))

        def start(self) -> None:
            return None

    runtime._remember_terminal_assigned_task("task-ws-terminal")
    monkeypatch.setattr(runtime, "_send_runtime_socket_message", fake_send_runtime_socket_message)
    monkeypatch.setattr(bridge.threading, "Thread", FakeThread)

    runtime._handle_runtime_ws_message(
        json.dumps(
            {
                "type": "task.dispatch",
                "task": {"id": "task-ws-terminal", "title": "Bitmis gorev"},
                "leaseId": "lease-ws-terminal",
            }
        )
    )

    assert ack_payloads == []
    assert thread_starts == []


def test_backend_task_wrappers_sync_task_inbox_state(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        def tasks_list(self, *, limit: int = 20, target_device_id: str = "", status: str = "") -> BackendResult:
            assert limit == 5
            assert target_device_id == ""
            assert status == ""
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        {
                            "id": "task-list-1",
                            "title": "Mobil görev",
                            "status": "waiting_approval",
                            "summary": "Onay bekleniyor.",
                            "targetDeviceId": VALID_DEVICE_ID,
                            "updatedAt": "2026-05-21T19:00:00Z",
                        }
                    ]
                },
                x_request_id="req-server-tasks",
            )

        def task_detail(self, task_id: str) -> BackendResult:
            assert task_id == "task-list-1"
            return BackendResult(
                ok=True,
                request_id="req_detail",
                status_code=200,
                data={
                    "task": {
                        "id": "task-list-1",
                        "title": "Mobil görev",
                        "status": "waiting_approval",
                        "summary": "Onay bekleniyor.",
                        "targetDeviceId": VALID_DEVICE_ID,
                        "updatedAt": "2026-05-21T19:00:00Z",
                    },
                    "events": [],
                    "artifacts": [{"kind": "summary", "name": "elyan-result.txt"}],
                },
            )

        def task_approval(self, task_id: str, approved: bool, notes: str | None = None) -> BackendResult:
            assert task_id == "task-list-1"
            assert approved is True
            assert notes is None
            return BackendResult(
                ok=True,
                request_id="req_approval",
                status_code=200,
                data={
                    "task": {
                        "id": "task-list-1",
                        "title": "Mobil görev",
                        "status": "waiting_approval",
                        "summary": "Onay işlendi.",
                        "targetDeviceId": VALID_DEVICE_ID,
                        "updatedAt": "2026-05-21T19:02:00Z",
                    }
                },
                x_request_id="req-server-approval",
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    listed = runtime.backend_tasks_list({"limit": 5, "hydrateDetails": True})
    approved = runtime.backend_task_approval("task-list-1", True)
    listed_envelope = runtime.handle(
        {
            "id": "req_tasks_envelope",
            "taskId": "task_tasks_envelope",
            "capability": "backend.tasks.list",
            "payload": {"limit": 5, "hydrateDetails": True},
        }
    )
    approved_envelope = runtime.handle(
        {
            "id": "req_approval_envelope",
            "taskId": "task_approval_envelope",
            "capability": "backend.tasks.approval",
            "payload": {"taskId": "task-list-1", "approved": True},
        }
    )

    assert listed["ok"] is True
    assert listed["result"]["xRequestId"] == "req-server-tasks"
    assert approved["ok"] is True
    assert approved["result"]["xRequestId"] == "req-server-approval"
    assert listed_envelope["ok"] is True
    assert listed_envelope["capability"] == "backend.tasks.list"
    assert listed_envelope["result"]["ok"] is True
    assert approved_envelope["ok"] is True
    assert approved_envelope["capability"] == "backend.tasks.approval"
    assert approved_envelope["result"]["ok"] is True
    inbox_item = state_store.get_task_inbox_item("task-list-1")
    assert inbox_item is not None
    assert inbox_item["artifactCount"] == 1


def test_integration_app_commands_use_bridge_and_disconnect_requires_confirmation(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    calls: list[tuple[str, str]] = []

    class FakeBackend:
        def integration_apps(self) -> BackendResult:
            calls.append(("apps", ""))
            return BackendResult(
                ok=True,
                request_id="req_apps",
                status_code=200,
                data={"apps": [{"id": "gmail", "displayName": "Gmail"}]},
            )

        def start_integration_app_oauth(self, app_id: str) -> BackendResult:
            calls.append(("connect", app_id))
            return BackendResult(
                ok=True,
                request_id="req_connect",
                status_code=200,
                data={"authUrl": "https://accounts.google.com/o/oauth2/v2/auth?state=safe"},
            )

        def disconnect_integration_app(self, app_id: str) -> BackendResult:
            calls.append(("disconnect", app_id))
            return BackendResult(
                ok=True,
                request_id="req_disconnect",
                status_code=200,
                data={"connected": False},
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    apps = runtime.handle(
        {"id": "req_apps", "taskId": "cli", "capability": "backend.integrations.apps", "payload": {}}
    )
    connect = runtime.handle(
        {
            "id": "req_connect",
            "taskId": "cli",
            "capability": "backend.integrations.oauth_start",
            "payload": {"appId": "gmail"},
        }
    )
    denied_disconnect = runtime.handle(
        {
            "id": "req_disconnect_denied",
            "taskId": "cli",
            "capability": "backend.integrations.disconnect",
            "payload": {"appId": "gmail"},
        }
    )
    confirmed_disconnect = runtime.handle(
        {
            "id": "req_disconnect",
            "taskId": "cli",
            "capability": "backend.integrations.disconnect",
            "payload": {"appId": "gmail", "_confirmed": True},
        }
    )
    invalid_app = runtime.handle(
        {
            "id": "req_invalid_app",
            "taskId": "cli",
            "capability": "backend.integrations.oauth_start",
            "payload": {"appId": "../gmail"},
        }
    )

    assert apps["ok"] is True
    assert connect["ok"] is True
    assert denied_disconnect["ok"] is False
    assert denied_disconnect["error"]["code"] == "PERMISSION_REQUIRED"
    assert confirmed_disconnect["ok"] is True
    assert invalid_app["ok"] is False
    assert invalid_app["error"]["code"] == "INTEGRATION_APP_INVALID"
    assert calls == [("apps", ""), ("connect", "gmail"), ("disconnect", "gmail")]
    advertised = bridge._runtime_advertised_capabilities()
    assert "backend.integrations.apps" in advertised
    assert "backend.integrations.oauth_start" in advertised
    assert "backend.integrations.disconnect" in advertised


def test_remote_mcp_unauthorized_forces_one_single_flight_registration_and_retry(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "runtimeToken": "stale-runtime-token",
                "ready": True,
            }
        }
    )
    calls = {"inventory": 0, "register": 0}

    class FakeBackend:
        def runtime_mcp_connections(self) -> BackendResult:
            calls["inventory"] += 1
            if calls["inventory"] == 1:
                return BackendResult(
                    ok=False,
                    request_id="req_mcp_unauthorized",
                    status_code=401,
                    data=None,
                    error="runtime_unauthorized",
                )
            return BackendResult(
                ok=True,
                request_id="req_mcp_retry",
                status_code=200,
                data={"servers": [], "revision": "rev_recovered"},
            )

        def register_runtime(self, payload: dict[str, object]) -> BackendResult:
            calls["register"] += 1
            assert payload["deviceId"] == VALID_DEVICE_ID
            return BackendResult(
                ok=True,
                request_id="req_register",
                status_code=200,
                data={
                    "runtime": {"deviceId": VALID_DEVICE_ID, "connectionId": VALID_CONNECTION_ID},
                    "tokens": {"accessToken": "fresh-runtime-token"},
                },
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_connect_runtime_transport", lambda: (False, None))
    monkeypatch.setattr(runtime, "_prime_runtime_task_delivery", lambda: None)

    result = runtime._remote_mcp_servers()

    assert result["errorCode"] == ""
    assert result["revision"] == "rev_recovered"
    assert calls == {"inventory": 2, "register": 1}
    assert state_store.snapshot()["runtime"]["runtimeToken"] == "fresh-runtime-token"


def test_remote_mcp_auth_retry_is_bounded_when_second_attempt_fails(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "runtimeToken": "stale-runtime-token",
                "ready": True,
            }
        }
    )
    calls = {"inventory": 0, "register": 0}

    class FakeBackend:
        def runtime_mcp_connections(self) -> BackendResult:
            calls["inventory"] += 1
            return BackendResult(
                ok=False,
                request_id=f"req_mcp_{calls['inventory']}",
                status_code=403,
                data=None,
                error="runtime_forbidden",
            )

        def register_runtime(self, _payload: dict[str, object]) -> BackendResult:
            calls["register"] += 1
            return BackendResult(
                ok=True,
                request_id="req_register",
                status_code=200,
                data={
                    "runtime": {"deviceId": VALID_DEVICE_ID, "connectionId": VALID_CONNECTION_ID},
                    "tokens": {"accessToken": "fresh-runtime-token"},
                },
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_connect_runtime_transport", lambda: (False, None))
    monkeypatch.setattr(runtime, "_prime_runtime_task_delivery", lambda: None)

    result = runtime._remote_mcp_servers()

    assert result["errorCode"] == "MCP_CONTROL_PLANE_AUTH_REQUIRED"
    assert calls == {"inventory": 2, "register": 1}


def test_runtime_tasks_execute_assigned_is_available_through_handle(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    def fake_execute(limit: int = 1) -> dict[str, object]:
        assert limit == 3
        return {"ok": True, "executions": [{"taskId": "task-1", "ok": True, "status": "completed"}], "fetched": 1}

    monkeypatch.setattr(runtime, "execute_assigned_runtime_tasks", fake_execute)

    response = runtime.handle(
        {
            "id": "req_execute_assigned",
            "taskId": "task_execute_assigned",
            "capability": "runtime.tasks.execute_assigned",
            "payload": {"limit": 3},
        }
    )

    assert response["ok"] is True
    assert response["capability"] == "runtime.tasks.execute_assigned"
    assert response["result"]["ok"] is True
    assert response["result"]["fetched"] == 1


def test_backend_auth_login_hydrates_truth_surfaces(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        def __init__(self) -> None:
            self.calls: list[str] = []
            self.configured = True
            self.loopback = False

        def auth_login(self, _email: str, _password: str) -> BackendResult:
            self.calls.append("login")
            return BackendResult(
                ok=True,
                request_id="req_login",
                status_code=200,
                data={
                    "user": {"email": "user@example.com", "displayName": "Emre"},
                    "subscription": {
                        "planCode": "free",
                        "status": "free",
                        "aiCreditsMonthly": 0,
                        "taskLimitMonthly": 10,
                        "periodEndsAt": "",
                    },
                    "tokens": {"accessToken": "user-token", "refreshToken": "refresh-token"},
                },
            )

        def auth_me(self) -> BackendResult:
            self.calls.append("auth_me")
            return BackendResult(ok=True, request_id="req_me", status_code=200, data={"user": {"email": "user@example.com"}})

        def mobile_bootstrap(self) -> BackendResult:
            self.calls.append("mobile_bootstrap")
            return BackendResult(ok=True, request_id="req_bootstrap", status_code=200, data={"devices": []})

        def health(self) -> BackendResult:
            self.calls.append("health")
            return BackendResult(
                ok=True,
                request_id="req_health",
                status_code=200,
                data={"dependencies": {"billing": {"status": "degraded", "checkoutEnabled": False}}},
            )

        def brain_profile(self) -> BackendResult:
            self.calls.append("brain_profile")
            return BackendResult(
                ok=True,
                request_id="req_brain_profile",
                status_code=200,
                data={"chat": {"localProviderHint": "ollama"}},
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    response = runtime.backend_auth_login({"email": "user@example.com", "password": "secret"})

    assert response["ok"] is True
    assert response["hydrationOk"] is True
    assert response["authMe"]["ok"] is True
    assert response["mobileBootstrap"]["ok"] is True
    assert response["health"]["ok"] is True
    assert response["brainProfile"]["ok"] is True
    assert response["runtimeSession"]["error"] == "runtime_token_missing"
    assert response["runtime"]["ok"] is True
    assert runtime.backend.calls == [  # type: ignore[attr-defined]
        "login",
        "auth_me",
        "mobile_bootstrap",
        "health",
        "brain_profile",
    ]


def test_backend_auth_login_stays_ok_when_bootstrap_is_partial(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        configured = True
        loopback = False

        def auth_login(self, _email: str, _password: str) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_login",
                status_code=200,
                data={
                    "user": {"email": "user@example.com", "displayName": "Emre"},
                    "subscription": {
                        "planCode": "free",
                        "status": "free",
                        "aiCreditsMonthly": 0,
                        "taskLimitMonthly": 10,
                        "periodEndsAt": "",
                    },
                    "tokens": {"accessToken": "user-token", "refreshToken": "refresh-token"},
                },
            )

        def auth_me(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_me", status_code=200, data={"user": {"email": "user@example.com"}})

        def mobile_bootstrap(self) -> BackendResult:
            return BackendResult(
                ok=False,
                request_id="req_bootstrap",
                status_code=503,
                data={"error": "bootstrap_unavailable"},
                error="bootstrap_unavailable",
            )

        def health(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_health",
                status_code=200,
                data={"dependencies": {"billing": {"status": "degraded"}}},
            )

        def brain_profile(self) -> BackendResult:
            return BackendResult(
                ok=False,
                request_id="req_brain_profile",
                status_code=503,
                data={"error": "brain_unavailable"},
                error="brain_unavailable",
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    response = runtime.backend_auth_login({"email": "user@example.com", "password": "secret"})

    assert response["ok"] is True
    assert response["hydrationOk"] is False
    assert response["result"]["ok"] is True
    assert response["authMe"]["ok"] is True
    assert response["mobileBootstrap"]["ok"] is False
    assert response["health"]["ok"] is True
    assert response["brainProfile"]["ok"] is False


def test_bootstrap_and_truth_refresh_share_canonical_backend_truth_surfaces(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        configured = True
        loopback = False

        def auth_me(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_me",
                status_code=200,
                data={
                    "user": {"id": "user-1", "email": "user@example.com", "displayName": "Emre"},
                    "subscription": {"planCode": "pro", "status": "active"},
                    "usage": {"dailyRemaining": 5, "weeklyRemaining": 25, "serverBrainAllowed": True},
                },
            )

        def mobile_bootstrap(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_bootstrap",
                status_code=200,
                data={"devices": [{"id": "mobile-1", "type": "mobile", "isActive": True}]},
            )

        def health(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_health",
                status_code=200,
                data={"network": {"externalClientsCanReachAdvertisedBaseUrl": True}},
            )

        def brain_profile(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_brain_profile",
                status_code=200,
                data={"chat": {"serverBrainName": "Elyan"}, "bridge": {"serverBrainReady": True}},
            )

        def runtime_session(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_runtime_session",
                status_code=200,
                data={
                    "readiness": {"targetStatus": "ready", "canReceiveTasks": True},
                    "connection": {"status": "online"},
                },
            )

    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "account": {"accessToken": "user-token"},
            "runtime": {
                "runtimeToken": "runtime-token",
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
            },
        }
    )
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    bootstrap = runtime.bootstrap()
    truth = runtime.backend_truth_refresh()

    for key in ("authMe", "mobileBootstrap", "health", "brainProfile", "runtimeSession"):
        assert bootstrap["backend"][key] == truth[key]
        assert bootstrap["backend"]["controlPlane"][key] == truth["controlPlane"][key]


def test_backend_auth_oauth_login_hydrates_truth_surfaces(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        def __init__(self) -> None:
            self.calls: list[str] = []
            self.configured = True
            self.loopback = False

        def auth_oauth_login(
            self,
            provider: str,
            id_token: str,
            *,
            email: str | None = None,
            display_name: str | None = None,
            authorization_code: str | None = None,
        ) -> BackendResult:
            self.calls.append(f"oauth:{provider}")
            assert provider == "apple"
            assert id_token == "apple-id-token"
            assert email == "user@example.com"
            assert display_name == "Emre"
            assert authorization_code == "auth-code"
            return BackendResult(
                ok=True,
                request_id="req_login",
                status_code=200,
                data={
                    "user": {"email": "user@example.com", "displayName": "Emre"},
                    "subscription": {
                        "planCode": "pro",
                        "status": "active",
                        "aiCreditsMonthly": 2500,
                        "taskLimitMonthly": 250,
                        "periodEndsAt": "",
                        "brainProfile": {
                            "chat": {
                                "tier": "premium",
                                "reasoningMultiplier": 5,
                            }
                        },
                    },
                    "tokens": {"accessToken": "user-token", "refreshToken": "refresh-token"},
                },
            )

        def auth_me(self) -> BackendResult:
            self.calls.append("auth_me")
            return BackendResult(ok=True, request_id="req_me", status_code=200, data={"user": {"email": "user@example.com"}})

        def mobile_bootstrap(self) -> BackendResult:
            self.calls.append("mobile_bootstrap")
            return BackendResult(ok=True, request_id="req_bootstrap", status_code=200, data={"devices": []})

        def health(self) -> BackendResult:
            self.calls.append("health")
            return BackendResult(
                ok=True,
                request_id="req_health",
                status_code=200,
                data={"dependencies": {"billing": {"status": "degraded", "checkoutEnabled": False}}},
            )

        def brain_profile(self) -> BackendResult:
            self.calls.append("brain_profile")
            return BackendResult(
                ok=True,
                request_id="req_brain_profile",
                status_code=200,
                data={"chat": {"localProviderHint": "ollama"}},
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    response = runtime.backend_auth_oauth_login(
        {
            "provider": "apple",
            "idToken": "apple-id-token",
            "email": "user@example.com",
            "displayName": "Emre",
            "authorizationCode": "auth-code",
        }
    )

    assert response["ok"] is True
    assert response["hydrationOk"] is True
    assert response["authMe"]["ok"] is True
    assert response["brainProfile"]["ok"] is True
    assert response["runtime"]["ok"] is True
    assert response["state"]["billing"]["planCode"] == "pro"
    assert response["state"]["billing"]["brainProfile"]["chat"]["reasoningMultiplier"] == 5
    assert response["state"]["account"]["subscription"]["planCode"] == "pro"
    assert runtime.backend.calls == [  # type: ignore[attr-defined]
        "oauth:apple",
        "auth_me",
        "mobile_bootstrap",
        "health",
        "brain_profile",
    ]


def test_backend_auth_login_applies_ollama_hint_only_when_local_models_are_empty(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        configured = True
        loopback = False

        def auth_login(self, _email: str, _password: str) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_login",
                status_code=200,
                data={
                    "user": {"email": "user@example.com"},
                    "tokens": {"accessToken": "user-token", "refreshToken": "refresh-token"},
                },
            )

        def auth_me(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_me", status_code=200, data={"user": {"email": "user@example.com"}})

        def mobile_bootstrap(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_bootstrap", status_code=200, data={"devices": []})

        def health(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_health", status_code=200, data={"dependencies": {}})

        def brain_profile(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_brain",
                status_code=200,
                data={"chat": {"localProviderHint": "ollama"}},
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(
        runtime,
        "local_models_status",
        lambda: {
            "models": {
                "models": [
                    {"name": "llama3.2:3b"},
                    {"name": "qwen2.5:7b"},
                ]
            }
        },
    )

    response = runtime.backend_auth_login({"email": "user@example.com", "password": "secret"})

    assert response["ok"] is True
    state = response["state"]
    assert state["providers"]["active"] == "ollama"
    assert state["providers"]["local"]["defaultModel"] == "llama3.2:3b"
    assert state["providers"]["ollama"]["defaultModel"] == "llama3.2:3b"


def test_backend_auth_login_preserves_existing_provider_when_local_model_is_set(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "providers": {
                "active": "openai",
                "local": {"defaultModel": "llama3.1:8b"},
                "ollama": {"defaultModel": "llama3.1:8b"},
            }
        }
    )

    class FakeBackend:
        configured = True
        loopback = False

        def auth_login(self, _email: str, _password: str) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_login",
                status_code=200,
                data={
                    "user": {"email": "user@example.com"},
                    "tokens": {"accessToken": "user-token", "refreshToken": "refresh-token"},
                },
            )

        def auth_me(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_me", status_code=200, data={"user": {"email": "user@example.com"}})

        def mobile_bootstrap(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_bootstrap", status_code=200, data={"devices": []})

        def health(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_health", status_code=200, data={"dependencies": {}})

        def brain_profile(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_brain",
                status_code=200,
                data={"chat": {"localProviderHint": "ollama"}},
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(
        runtime,
        "local_models_status",
        lambda: {"models": {"models": [{"name": "llama3.2:3b"}]}},
    )

    response = runtime.backend_auth_login({"email": "user@example.com", "password": "secret"})

    assert response["ok"] is True
    state = response["state"]
    assert state["providers"]["active"] == "openai"
    assert state["providers"]["local"]["defaultModel"] == "llama3.1:8b"
    assert state["providers"]["ollama"]["defaultModel"] == "llama3.1:8b"


def test_backend_auth_logout_returns_cleared_state(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "account": {"accessToken": "user-token", "refreshToken": "refresh-token"},
            "runtime": {"deviceId": VALID_DEVICE_ID, "deviceSecret": VALID_DEVICE_SECRET, "ready": True},
        }
    )

    class FakeBackend:
        configured = True
        loopback = False

        def auth_logout(self) -> BackendResult:
            state_store.update_state(
                {
                    "account": {"accessToken": "", "refreshToken": "", "email": ""},
                    "runtime": {"ready": False, "lifecycleState": "offline"},
                }
            )
            return BackendResult(ok=True, request_id="req_logout", status_code=200, data={"ok": True})

        def health(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_health",
                status_code=200,
                data={"dependencies": {"billing": {"status": "degraded"}}},
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    response = runtime.backend_auth_logout()

    assert response["ok"] is True
    assert response["health"]["ok"] is True
    assert response["runtimeSession"]["error"] == "logged_out"
    assert response["state"]["account"]["accessToken"] == ""
    assert response["runtime"]["runtimeLifecycleState"] == "offline"


def test_backend_auth_sync_session_hydrates_truth_surfaces(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        configured = True
        loopback = False

        def auth_me(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_me",
                status_code=200,
                data={"user": {"email": "user@example.com", "displayName": "Emre"}},
            )

        def mobile_bootstrap(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_bootstrap", status_code=200, data={"devices": []})

        def health(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_health", status_code=200, data={"dependencies": {}})

        def brain_profile(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_brain",
                status_code=200,
                data={"chat": {"serverBrainName": "Elyan"}},
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    response = runtime.backend_auth_sync_session(
        {
            "signedIn": True,
            "id": "user-1",
            "email": "user@example.com",
            "displayName": "Emre",
            "accessToken": "user-token",
            "refreshToken": "refresh-token",
        }
    )

    assert response["ok"] is True
    assert response["signedIn"] is True
    assert response["authMe"]["ok"] is True
    assert response["mobileBootstrap"]["ok"] is True
    assert response["brainProfile"]["ok"] is True
    state = response["state"]
    assert state["account"]["accessToken"] == "user-token"
    assert state["account"]["refreshToken"] == "refresh-token"
    assert state["account"]["email"] == "user@example.com"
    assert state["account"]["displayName"] == "Emre"


def test_backend_auth_sync_session_clears_local_truth_without_network_logout(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "account": {
                "accessToken": "user-token",
                "refreshToken": "refresh-token",
                "email": "user@example.com",
                "displayName": "Emre",
                "subscription": {"planCode": "solo"},
            },
            "controlPlane": {
                "authMe": {"ok": True},
                "mobileBootstrap": {"ok": True},
                "brainProfile": {"ok": True},
                "runtimeSession": {"ok": True},
            },
        }
    )

    class FakeBackend:
        configured = True
        loopback = False

        def health(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_health", status_code=200, data={"dependencies": {}})

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    response = runtime.backend_auth_sync_session({"signedIn": False})

    assert response["ok"] is True
    assert response["signedIn"] is False
    assert response["authMe"]["error"] == "logged_out"
    assert response["state"]["account"]["accessToken"] == ""
    assert response["state"]["account"]["refreshToken"] == ""
    assert response["state"]["controlPlane"]["authMe"] is None
    assert response["state"]["controlPlane"]["mobileBootstrap"] is None


def test_runtime_handle_sanitizes_state_tokens_for_transport(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "account": {"accessToken": "user-token", "refreshToken": "refresh-token"},
            "runtime": {
                "runtimeToken": "runtime-token",
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "connectionId": "conn-1",
            },
            "pairing": {
                "pairingToken": "pair-token",
                "lastSessionId": "session-1",
            },
        }
    )

    runtime = bridge.RuntimeBridge()
    response = runtime.handle({"id": "req_state", "taskId": "task_state", "capability": "state.get", "payload": {}})

    assert response["ok"] is True
    state = response["result"]["state"]
    assert state["account"]["accessToken"] == ""
    assert state["account"]["refreshToken"] == ""
    assert state["runtime"]["runtimeToken"] == ""
    assert state["runtime"]["deviceSecret"] == ""
    assert state["runtime"]["connectionId"] == ""
    assert state["pairing"]["pairingToken"] == ""
    assert state["pairing"]["lastSessionId"] == ""


def test_runtime_handle_sanitizes_nested_auth_tokens_from_backend_payload(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    class FakeBackend:
        configured = True
        loopback = False

        def auth_login(self, _email: str, _password: str) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_login",
                status_code=200,
                data={
                    "user": {"email": "user@example.com"},
                    "tokens": {"accessToken": "user-token", "refreshToken": "refresh-token"},
                },
            )

        def auth_me(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_me", status_code=200, data={"user": {"email": "user@example.com"}})

        def mobile_bootstrap(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_bootstrap", status_code=200, data={"devices": []})

        def health(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_health", status_code=200, data={"dependencies": {}})

        def brain_profile(self) -> BackendResult:
            return BackendResult(ok=True, request_id="req_brain", status_code=200, data={"chat": {"localProviderHint": "ollama"}})

    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "controlPlane": {
                "health": {
                    "ok": True,
                    "agent": {"chatReady": True, "serverBrainReady": True},
                }
            }
        }
    )
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    response = runtime.handle(
        {
            "id": "req_auth_login",
            "taskId": "task_auth_login",
            "capability": "backend.auth_login",
            "payload": {"email": "user@example.com", "password": "secret"},
        }
    )

    assert response["ok"] is True
    result = response["result"]
    assert result["result"]["data"]["tokens"]["accessToken"] == ""
    assert result["result"]["data"]["tokens"]["refreshToken"] == ""
    assert result["state"]["account"]["accessToken"] == ""
    assert result["state"]["account"]["refreshToken"] == ""
    snapshot = state_store.snapshot()
    assert snapshot["account"]["accessToken"] == "user-token"
    assert snapshot["account"]["refreshToken"] == "refresh-token"


def test_ensure_runtime_registered_uses_heartbeat_fallback_when_websocket_unavailable(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    original_module_available = bridge._module_available

    def fake_module_available(module_name: str) -> bool:
        if module_name in {"qiskit", "qiskit_aer"}:
            return True
        return original_module_available(module_name)

    monkeypatch.setattr(bridge, "_module_available", fake_module_available)
    state_store.update_state(
        {
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
            }
        }
    )

    class FakeBackend:
        configured = True
        loopback = False

        def register_runtime(self, payload: dict[str, object]) -> BackendResult:
            assert payload["deviceId"] == VALID_DEVICE_ID
            assert payload["deviceSecret"] == VALID_DEVICE_SECRET
            capabilities = payload.get("capabilities")
            assert isinstance(capabilities, list)
            assert "quantum_run_experiment" in capabilities
            return BackendResult(
                ok=True,
                request_id="req_register",
                status_code=200,
                data={
                    "runtime": {"deviceId": VALID_DEVICE_ID, "connectionId": VALID_CONNECTION_ID},
                    "tokens": {"accessToken": "runtime-token"},
                },
                x_request_id="req-server-register",
            )

        def heartbeat(self, payload: dict[str, object]) -> BackendResult:
            assert payload["status"] == "online"
            capabilities = payload.get("capabilities")
            assert isinstance(capabilities, list)
            assert "quantum_model_problem" in capabilities
            assert "quantum_run_experiment" in capabilities
            assert "quantum_compare_classical" in capabilities
            assert "quantum_generate_report" in capabilities
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_start_runtime_websocket_if_needed", lambda: False)
    monkeypatch.setattr(runtime, "_start_task_relay_if_ready", lambda: None)

    response = runtime.ensure_runtime_registered()

    snapshot = state_store.snapshot()
    assert response["ok"] is True
    assert response["transport"]["mode"] == "heartbeat"
    assert snapshot["runtime"]["runtimeToken"] == "runtime-token"
    assert snapshot["runtime"]["ready"] is True


def test_runtime_transport_heartbeats_until_websocket_is_actually_open(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    heartbeat_calls: list[str] = []
    relay_started = {"value": False}
    monkeypatch.setattr(runtime, "_start_runtime_websocket_if_needed", lambda: True)
    monkeypatch.setattr(
        runtime,
        "_send_backend_runtime_heartbeat",
        lambda status: heartbeat_calls.append(status)
        or BackendResult(ok=True, request_id="req-heartbeat", status_code=200, data={"ok": True}),
    )
    monkeypatch.setattr(runtime, "_start_task_relay_if_ready", lambda: relay_started.__setitem__("value", True))

    connected, heartbeat = runtime._connect_runtime_transport()

    assert connected is False
    assert heartbeat is not None and heartbeat.ok is True
    assert heartbeat_calls == ["online"]
    assert relay_started["value"] is True


def test_runtime_transport_does_not_overwrite_websocket_opened_during_heartbeat(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "runtime": {
                "runtimeToken": "runtime-token",
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
            }
        }
    )
    runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(runtime, "_start_runtime_websocket_if_needed", lambda: True)

    def heartbeat_after_socket_open(_status: str) -> BackendResult:
        runtime._runtime_ws_connected = True
        return BackendResult(ok=True, request_id="req-heartbeat", status_code=200, data={"ok": True})

    monkeypatch.setattr(runtime, "_send_backend_runtime_heartbeat", heartbeat_after_socket_open)
    monkeypatch.setattr(runtime, "_start_task_relay_if_ready", lambda: None)

    runtime._connect_runtime_transport()

    snapshot = state_store.snapshot()
    assert snapshot["runtime"]["websocketConnected"] is True
    assert snapshot["runtime"]["lifecycleState"] == "ready"


def test_ensure_runtime_registered_rejects_invalid_runtime_identity_without_backend_call(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "pairing": {
                "desktopDeviceId": "11111111-1111-4111-8111-111111111111",
                "lastSessionStatus": "claimed",
            },
            "runtime": {
                "deviceId": "desktop-ext-1",
                "deviceSecret": "short-secret",
            },
        }
    )

    class FakeBackend:
        def register_runtime(self, _payload: dict[str, object]) -> BackendResult:
            raise AssertionError("register_runtime should not be called")

        def runtime_register_identity_error(self) -> dict[str, str] | None:
            return {
                "code": "RUNTIME_REGISTER_INVALID_IDENTITY",
                "message": "Runtime kimliği geçersiz.",
            }

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_start_runtime_websocket_if_needed", lambda: False)
    monkeypatch.setattr(runtime, "_start_task_relay_if_ready", lambda: None)

    response = runtime.ensure_runtime_registered()

    snapshot = state_store.snapshot()
    assert response == {
        "ok": False,
        "error": {
            "code": "RUNTIME_REGISTER_INVALID_IDENTITY",
            "message": "Runtime kimliği geçersiz.",
        },
    }
    assert snapshot["runtime"]["lifecycleState"] == "offline"
    assert snapshot["runtime"]["lastErrorCode"] == "runtime_register_invalid_identity"
    assert snapshot["runtime"]["deviceId"] == ""
    assert snapshot["runtime"]["deviceSecret"] == ""
    assert snapshot["pairing"]["lastSessionId"] == ""


def test_runtime_register_uses_state_identity_and_advertised_capabilities(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    original_module_available = bridge._module_available

    def fake_module_available(module_name: str) -> bool:
        if module_name in {"qiskit", "qiskit_aer"}:
            return True
        return original_module_available(module_name)

    monkeypatch.setattr(bridge, "_module_available", fake_module_available)
    state_store.update_state(
        {
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
            }
        }
    )

    captured: dict[str, object] = {}

    class FakeBackend:
        configured = True
        loopback = False

        def register_runtime(self, payload: dict[str, object]) -> BackendResult:
            captured["payload"] = dict(payload)
            return BackendResult(
                ok=True,
                request_id="req_register",
                status_code=200,
                data={
                    "runtime": {"deviceId": VALID_DEVICE_ID, "connectionId": VALID_CONNECTION_ID},
                    "tokens": {"accessToken": "runtime-token"},
                },
                x_request_id="req-server-register",
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_connect_runtime_transport", lambda: (False, None))
    monkeypatch.setattr(runtime, "_prime_runtime_task_delivery", lambda: None)

    response = runtime.register_runtime(
        {
            "deviceId": "ignored",
            "deviceSecret": "ignored",
            "capabilities": ["bogus"],
        }
    )

    snapshot = state_store.snapshot()
    payload = captured["payload"]
    assert response["ok"] is True
    assert isinstance(payload, dict)
    assert payload["deviceId"] == VALID_DEVICE_ID
    assert payload["deviceSecret"] == VALID_DEVICE_SECRET
    assert isinstance(payload["capabilities"], list)
    assert "quantum_run_experiment" in payload["capabilities"]
    assert snapshot["runtime"]["runtimeToken"] == "runtime-token"
    assert snapshot["runtime"]["lifecycleState"] == "runtime_connecting"


def test_explicit_runtime_register_and_constructor_retry_are_single_flight(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
            }
        }
    )

    retry_checked = threading.Event()
    release_retry = threading.Event()
    original_should_continue = bridge.RuntimeBridge._runtime_register_retry_should_continue

    def controlled_should_continue(
        runtime: bridge.RuntimeBridge,
        target: dict[str, str],
        *,
        generation: int,
    ) -> bool:
        should_continue = original_should_continue(runtime, target, generation=generation)
        if threading.current_thread().name == "elyan-runtime-register-retry" and not retry_checked.is_set():
            retry_checked.set()
            release_retry.wait(timeout=2)
        return should_continue

    monkeypatch.setattr(
        bridge.RuntimeBridge,
        "_runtime_register_retry_should_continue",
        controlled_should_continue,
    )

    register_calls = 0

    class FakeBackend:
        configured = True
        loopback = False

        def register_runtime(self, _payload: dict[str, object]) -> BackendResult:
            nonlocal register_calls
            register_calls += 1
            return BackendResult(
                ok=True,
                request_id="req_register_single_flight",
                status_code=200,
                data={
                    "runtime": {"deviceId": VALID_DEVICE_ID, "connectionId": VALID_CONNECTION_ID},
                    "tokens": {"accessToken": "runtime-token"},
                },
                x_request_id="req-server-register-single-flight",
            )

    runtime = bridge.RuntimeBridge()
    assert retry_checked.wait(timeout=1)
    retry_thread = runtime._runtime_register_retry_thread
    assert retry_thread is not None
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_connect_runtime_transport", lambda: (False, None))
    monkeypatch.setattr(runtime, "_prime_runtime_task_delivery", lambda: None)

    try:
        response = runtime.register_runtime({})
    finally:
        release_retry.set()
    retry_thread.join(timeout=2)

    snapshot = state_store.snapshot()
    assert response["ok"] is True
    assert register_calls == 1
    assert retry_thread.is_alive() is False
    assert snapshot["runtime"]["runtimeToken"] == "runtime-token"
    assert snapshot["runtime"]["lifecycleState"] == "runtime_connecting"


def test_runtime_register_rejects_invalid_identity_without_backend_call(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "pairing": {
                "desktopDeviceId": VALID_DEVICE_ID,
                "lastSessionStatus": "claimed",
            },
            "runtime": {
                "deviceId": "desktop-ext-1",
                "deviceSecret": "short-secret",
            },
        }
    )

    class FakeBackend:
        def register_runtime(self, _payload: dict[str, object]) -> BackendResult:
            raise AssertionError("register_runtime should not be called")

        def runtime_register_identity_error(self) -> dict[str, str] | None:
            return {
                "code": "RUNTIME_REGISTER_INVALID_IDENTITY",
                "message": "Runtime kimliği geçersiz.",
            }

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_start_runtime_websocket_if_needed", lambda: False)
    monkeypatch.setattr(runtime, "_start_task_relay_if_ready", lambda: None)

    response = runtime.register_runtime(
        {
            "deviceId": VALID_DEVICE_ID,
            "deviceSecret": VALID_DEVICE_SECRET,
            "capabilities": [],
        }
    )

    snapshot = state_store.snapshot()
    assert response == {
        "ok": False,
        "error": {
            "code": "RUNTIME_REGISTER_INVALID_IDENTITY",
            "message": "Runtime kimliği geçersiz.",
        },
    }
    assert snapshot["runtime"]["lifecycleState"] == "offline"
    assert snapshot["runtime"]["lastErrorCode"] == "runtime_register_invalid_identity"
    assert snapshot["runtime"]["deviceId"] == ""
    assert snapshot["runtime"]["deviceSecret"] == ""
    assert snapshot["pairing"]["lastSessionId"] == ""


def test_concurrent_runtime_registration_is_single_flight(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "account": {"accessToken": "user-token"},
            "pairing": {
                "lastSessionId": "session-1",
                "desktopDeviceId": VALID_DEVICE_ID,
                "lastSessionStatus": "claimed",
            },
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "ready": False,
            },
        }
    )

    class FakeBackend:
        def __init__(self) -> None:
            self.register_calls = 0

        def runtime_register_identity_error(self) -> None:
            return None

        def register_runtime(self, _payload: dict[str, object]) -> BackendResult:
            self.register_calls += 1
            time.sleep(0.03)
            return BackendResult(
                ok=True,
                request_id="req-register",
                status_code=200,
                data={
                    "runtime": {"deviceId": VALID_DEVICE_ID, "connectionId": VALID_CONNECTION_ID},
                    "tokens": {"accessToken": "runtime-token"},
                },
            )

    runtime.backend = FakeBackend()  # type: ignore[assignment]

    def connect_once() -> tuple[bool, None]:
        state_store.update_state({"runtime": {"ready": True}})
        return True, None

    monkeypatch.setattr(runtime, "_connect_runtime_transport", connect_once)
    monkeypatch.setattr(runtime, "_prime_runtime_task_delivery", lambda: None)
    results: list[dict[str, object]] = []
    threads = [threading.Thread(target=lambda: results.append(runtime.ensure_runtime_registered())) for _ in range(2)]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=1)

    assert runtime.backend.register_calls == 1  # type: ignore[attr-defined]
    assert len(results) == 2
    assert all(result["ok"] is True for result in results)
    assert any(result.get("reused") is True for result in results)


def test_reconnect_registration_reuses_concurrent_explicit_token(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "runtimeToken": "runtime-token-old",
                "ready": False,
            }
        }
    )
    runtime = bridge.RuntimeBridge()

    class SignalingRegistrationLock:
        def __init__(self) -> None:
            self._lock = threading.RLock()
            self.reconnect_waiting = threading.Event()

        def __enter__(self) -> SignalingRegistrationLock:
            if threading.current_thread().name == "test-runtime-reconnect":
                self.reconnect_waiting.set()
            self._lock.acquire()
            return self

        def __exit__(self, _exc_type: object, _exc: object, _tb: object) -> None:
            self._lock.release()

    registration_lock = SignalingRegistrationLock()
    runtime._runtime_registration_lock = registration_lock  # type: ignore[assignment]

    class FakeBackend:
        configured = True
        loopback = False

        def __init__(self) -> None:
            self.register_calls = 0

        def register_runtime(self, _payload: dict[str, object]) -> BackendResult:
            self.register_calls += 1
            return BackendResult(
                ok=True,
                request_id="req_register_explicit",
                status_code=200,
                data={
                    "runtime": {"deviceId": VALID_DEVICE_ID, "connectionId": VALID_CONNECTION_ID},
                    "tokens": {"accessToken": "runtime-token-new"},
                },
                x_request_id="req-server-register-explicit",
            )

    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_connect_runtime_transport", lambda: (False, None))
    monkeypatch.setattr(runtime, "_prime_runtime_task_delivery", lambda: None)
    reconnect_results: list[bool] = []

    with registration_lock:
        reconnect_thread = threading.Thread(
            target=lambda: reconnect_results.append(runtime._refresh_runtime_registration_for_reconnect()),
            name="test-runtime-reconnect",
        )
        reconnect_thread.start()
        assert registration_lock.reconnect_waiting.wait(timeout=1)
        explicit_response = runtime.register_runtime({})

    reconnect_thread.join(timeout=2)
    snapshot = state_store.snapshot()

    assert explicit_response["ok"] is True
    assert reconnect_results == [True]
    assert reconnect_thread.is_alive() is False
    assert runtime.backend.register_calls == 1  # type: ignore[attr-defined]
    assert snapshot["runtime"]["runtimeToken"] == "runtime-token-new"
    assert snapshot["runtime"]["lifecycleState"] == "runtime_connecting"


def test_ensure_runtime_registered_primes_assigned_task_fetch(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
            }
        }
    )
    fetch_calls: list[int] = []

    class FakeBackend:
        def register_runtime(self, _payload: dict[str, object]) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_register",
                status_code=200,
                data={
                    "runtime": {"deviceId": VALID_DEVICE_ID, "connectionId": VALID_CONNECTION_ID},
                    "tokens": {"accessToken": "runtime-token"},
                },
            )

        def heartbeat(self, _payload: dict[str, object]) -> BackendResult:
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_start_runtime_websocket_if_needed", lambda: False)
    monkeypatch.setattr(runtime, "_start_task_relay_if_ready", lambda: None)
    monkeypatch.setattr(
        runtime,
        "execute_assigned_runtime_tasks",
        lambda limit=1: fetch_calls.append(limit) or {"ok": True, "executions": [], "fetched": 0},
    )

    response = runtime.ensure_runtime_registered()

    assert response["ok"] is True
    assert fetch_calls == [2]


def test_runtime_register_retry_recovers_after_transient_failure(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "account": {"accessToken": "user-token"},
            "pairing": {
                "lastSessionId": "session-1",
                "desktopDeviceId": VALID_DEVICE_ID,
                "lastSessionStatus": "claimed",
                "expiresAt": "2030-05-22T15:30:00Z",
            },
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "ready": False,
            },
        }
    )

    class FakeBackend:
        def __init__(self) -> None:
            self.register_calls = 0

        def register_runtime(self, payload: dict[str, object]) -> BackendResult:
            self.register_calls += 1
            if self.register_calls == 1:
                return BackendResult(
                    ok=False,
                    request_id="req_register_fail",
                    status_code=503,
                    data={"error": "backend_unavailable"},
                    error="backend_unavailable",
                )
            assert payload["deviceId"] == VALID_DEVICE_ID
            return BackendResult(
                ok=True,
                request_id="req_register_ok",
                status_code=200,
                data={
                    "runtime": {"deviceId": VALID_DEVICE_ID, "connectionId": VALID_CONNECTION_ID},
                    "tokens": {"accessToken": "runtime-token"},
                },
                x_request_id="req-register-2",
            )

        def heartbeat(self, payload: dict[str, object]) -> BackendResult:
            assert payload["status"] == "online"
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_start_runtime_websocket_if_needed", lambda: False)
    monkeypatch.setattr(runtime, "_start_task_relay_if_ready", lambda: None)
    monkeypatch.setattr(runtime, "_prime_runtime_task_delivery", lambda: None)
    monkeypatch.setattr(runtime, "_runtime_register_retry_backoff_seconds", lambda: [0.0, 0.01])

    snapshot = runtime._runtime_register_retry_snapshot()
    target = runtime._runtime_register_retry_target_from_snapshot(snapshot)
    assert target is not None
    with runtime._runtime_register_retry_lock:
        runtime._runtime_register_retry_generation += 1
        generation = runtime._runtime_register_retry_generation
        runtime._runtime_register_retry_target = target
    runtime._runtime_register_retry_loop(target, generation)

    snapshot = state_store.snapshot()
    assert runtime.backend.register_calls >= 2  # type: ignore[attr-defined]
    assert snapshot["runtime"]["runtimeToken"] == "runtime-token"
    assert snapshot["runtime"]["ready"] is True


def test_runtime_register_retry_uses_paired_device_after_qr_session_expires(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "account": {"accessToken": "user-token"},
            "pairing": {
                "lastSessionId": "session-1",
                "desktopDeviceId": VALID_DEVICE_ID,
                "lastSessionStatus": "claimed",
                "expiresAt": "2020-01-01T00:00:00Z",
            },
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "ready": False,
            },
        }
    )

    snapshot = runtime._runtime_register_retry_snapshot()

    assert runtime._runtime_register_retry_eligible(snapshot) is True
    assert runtime._runtime_register_retry_target_from_snapshot(snapshot) == {
        "lastSessionId": "session-1",
        "desktopDeviceId": VALID_DEVICE_ID,
        "deviceId": VALID_DEVICE_ID,
    }


def test_pairing_claim_poll_recovers_claimed_session_and_registers_runtime(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "account": {"accessToken": "user-token"},
            "pairing": {
                "lastSessionId": "session-1",
                "pairingToken": "pair-token",
                "desktopDeviceId": VALID_DEVICE_ID,
                "lastSessionStatus": "pending",
                "expiresAt": "2020-01-01T00:00:00Z",
            },
            "runtime": {"ready": False},
        }
    )

    class FakeBackend:
        def __init__(self) -> None:
            self.pairing_get_calls = 0
            self.register_calls = 0

        def pairing_get_session(self, session_id: str) -> BackendResult:
            self.pairing_get_calls += 1
            assert session_id == "session-1"
            state_store.update_state(
                {
                    "pairing": {
                        "lastSessionStatus": "claimed",
                        "lastClaimedAt": "2026-06-21T09:12:49.053Z",
                    },
                    "runtime": {
                        "deviceId": VALID_DEVICE_ID,
                        "deviceSecret": VALID_DEVICE_SECRET,
                    },
                }
            )
            return BackendResult(
                ok=True,
                request_id="req_pair",
                status_code=200,
                data={
                    "sessionId": "session-1",
                    "desktopDeviceId": VALID_DEVICE_ID,
                    "status": "claimed",
                    "runtimeAuth": {
                        "deviceId": VALID_DEVICE_ID,
                        "deviceSecret": VALID_DEVICE_SECRET,
                    },
                },
            )

        def register_runtime(self, payload: dict[str, object]) -> BackendResult:
            self.register_calls += 1
            assert payload["deviceId"] == VALID_DEVICE_ID
            assert payload["deviceSecret"] == VALID_DEVICE_SECRET
            return BackendResult(
                ok=True,
                request_id="req_register_ok",
                status_code=200,
                data={
                    "runtime": {"deviceId": VALID_DEVICE_ID, "connectionId": VALID_CONNECTION_ID},
                    "tokens": {"accessToken": "runtime-token"},
                },
                x_request_id="req-register",
            )

        def heartbeat(self, payload: dict[str, object]) -> BackendResult:
            assert payload["status"] == "online"
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

        def runtime_websocket_url(self) -> str:
            return ""

    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_start_runtime_websocket_if_needed", lambda: False)
    monkeypatch.setattr(runtime, "_start_task_relay_if_ready", lambda: None)
    monkeypatch.setattr(runtime, "_prime_runtime_task_delivery", lambda: None)
    monkeypatch.setattr(runtime, "_pairing_claim_poll_backoff_seconds", lambda: [0.0, 0.01])

    runtime._start_pairing_claim_poll_if_needed()

    deadline = time.monotonic() + 0.5
    while time.monotonic() < deadline:
        snapshot = state_store.snapshot()
        if snapshot["runtime"].get("ready") is True:
            break
        time.sleep(0.02)

    snapshot = state_store.snapshot()
    assert runtime.backend.pairing_get_calls >= 1  # type: ignore[attr-defined]
    assert runtime.backend.register_calls == 1  # type: ignore[attr-defined]
    assert snapshot["pairing"]["lastSessionStatus"] == "claimed"
    assert snapshot["runtime"]["runtimeToken"] == "runtime-token"
    assert snapshot["runtime"]["ready"] is True


def test_pairing_claim_poll_marks_runtime_waiting_for_phone(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "pairing": {
                "lastSessionId": "session-pending",
                "pairingToken": "pair-token",
                "lastSessionStatus": "pending",
                "expiresAt": "2030-05-22T15:30:00Z",
            },
            "runtime": {
                "lifecycleState": "offline",
                "lastErrorCode": "runtime_unauthorized",
            },
        }
    )
    monkeypatch.setattr(runtime, "_pairing_claim_poll_loop", lambda *_args: None)

    runtime._start_pairing_claim_poll_if_needed()

    state = state_store.snapshot()
    assert state["runtime"]["lifecycleState"] == "waiting_claim"
    assert state["runtime"]["lastErrorCode"] == ""
    assert state["runtime"]["ready"] is False


def test_runtime_unauthorized_endpoint_starts_register_retry(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "account": {"accessToken": "user-token"},
            "pairing": {
                "lastSessionId": "session-1",
                "desktopDeviceId": VALID_DEVICE_ID,
                "lastSessionStatus": "claimed",
                "expiresAt": "2020-01-01T00:00:00Z",
            },
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "ready": False,
            },
        }
    )
    started = {"value": False}

    class FakeBackend:
        def runtime_session(self) -> BackendResult:
            state_store.update_state(
                {
                    "runtime": {
                        "runtimeToken": "",
                        "ready": False,
                        "lastErrorCode": "runtime_unauthorized",
                    }
                }
            )
            return BackendResult(
                ok=False,
                request_id="req_runtime_session",
                status_code=401,
                data={"error": "runtime_unauthorized"},
                error="runtime_unauthorized",
            )

    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(
        runtime,
        "_start_runtime_register_retry_if_needed",
        lambda: started.__setitem__("value", True),
    )

    response = runtime.runtime_session()

    assert response["ok"] is False
    assert started["value"] is True


def test_status_registers_paired_runtime_when_offline(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "account": {"accessToken": "user-token"},
            "pairing": {
                "lastSessionId": "session-1",
                "desktopDeviceId": VALID_DEVICE_ID,
                "lastSessionStatus": "claimed",
                "expiresAt": "2020-01-01T00:00:00Z",
            },
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "ready": False,
            },
        }
    )

    class FakeBackend:
        configured = True
        loopback = False

        def register_runtime(self, payload: dict[str, object]) -> BackendResult:
            assert payload["deviceId"] == VALID_DEVICE_ID
            assert payload["deviceSecret"] == VALID_DEVICE_SECRET
            return BackendResult(
                ok=True,
                request_id="req_register_ok",
                status_code=200,
                data={
                    "runtime": {"deviceId": VALID_DEVICE_ID, "connectionId": VALID_CONNECTION_ID},
                    "tokens": {"accessToken": "runtime-token"},
                },
            )

        def heartbeat(self, payload: dict[str, object]) -> BackendResult:
            assert payload["status"] == "online"
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

        def runtime_websocket_url(self) -> str:
            return ""

    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_start_task_relay_if_ready", lambda: None)
    monkeypatch.setattr(runtime, "_prime_runtime_task_delivery", lambda: None)

    payload = runtime.status()

    snapshot = state_store.snapshot()
    assert payload["runtimeLifecycleState"] == "ready"
    assert snapshot["runtime"]["runtimeToken"] == "runtime-token"
    assert snapshot["runtime"]["ready"] is True


def test_stale_bridge_background_workers_stop_after_state_store_switch(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    target = {"deviceId": VALID_DEVICE_ID, "deviceSecret": VALID_DEVICE_SECRET}
    runtime._runtime_register_retry_target = target
    runtime._runtime_register_retry_generation = 1

    monkeypatch.setattr(state_store, "STATE_PATH", tmp_path / "next-profile" / "elyan_state.json")

    assert runtime._runtime_register_retry_should_continue(target, generation=1) is False


def test_runtime_register_retry_invalidation_wakes_sleeping_thread(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "account": {"accessToken": "user-token"},
            "pairing": {
                "lastSessionId": "session-1",
                "desktopDeviceId": VALID_DEVICE_ID,
                "lastSessionStatus": "claimed",
                "expiresAt": "2030-05-22T15:30:00Z",
            },
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "ready": False,
            },
        }
    )

    class FakeBackend:
        register_calls = 0

        def runtime_register_identity_error(self) -> dict[str, str] | None:
            return None

        def register_runtime(self, _payload: dict[str, object]) -> BackendResult:
            self.register_calls += 1
            return BackendResult(
                ok=False,
                request_id="req_unexpected",
                status_code=503,
                error="backend_unavailable",
            )

    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_runtime_register_retry_backoff_seconds", lambda: [30.0])

    runtime._start_runtime_register_retry_if_needed()
    thread = runtime._runtime_register_retry_thread
    assert thread is not None

    runtime._invalidate_runtime_register_retry()
    thread.join(timeout=0.5)

    assert not thread.is_alive()
    assert runtime.backend.register_calls == 0  # type: ignore[attr-defined]


def test_relay_interval_shrinks_while_active_remote_tasks_exist(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    state_store.sync_task_inbox(
        [
            {
                "id": "task-1",
                "title": "Yeni görev",
                "status": "queued",
                "targetDeviceId": VALID_DEVICE_ID,
                "queuePosition": 1,
                "summary": "",
                "error": "",
                "approvalRequest": {},
                "createdAt": "2030-01-01T00:00:00Z",
                "startedAt": "",
                "completedAt": "",
                "canceledAt": "",
                "updatedAt": "2030-01-01T00:00:00Z",
                "artifactCount": 0,
                "origin": "mobile",
            }
        ]
    )

    monkeypatch.setenv("ELYAN_RUNTIME_RELAY_ACTIVE_INTERVAL_SECONDS", "2.2")
    monkeypatch.setenv("ELYAN_RUNTIME_RELAY_RECONNECTING_INTERVAL_SECONDS", "3.4")
    monkeypatch.setenv("ELYAN_RUNTIME_RELAY_IDLE_INTERVAL_SECONDS", "5.0")

    assert runtime._relay_mode() == "active"
    assert runtime._relay_interval_seconds() == 2.2
    assert runtime._relay_task_fetch_limit() == 2

    state_store.sync_task_inbox([])
    state_store.update_state({"runtime": {"lifecycleState": "reconnecting"}})

    assert runtime._relay_mode() == "reconnecting"
    assert runtime._relay_interval_seconds() == 3.4
    assert runtime._relay_task_fetch_limit() == 3

    state_store.update_state({"runtime": {"lifecycleState": "offline"}})

    assert runtime._relay_mode() == "idle"
    assert runtime._relay_interval_seconds() == 5.0
    assert runtime._relay_task_fetch_limit() == 1


def test_terminal_task_status_triggers_detail_resync(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    resynced: list[str] = []

    monkeypatch.setattr(runtime, "_send_runtime_socket_message", lambda payload: False)
    monkeypatch.setattr(
        runtime.backend,
        "runtime_task_status",
        lambda task_id, payload: bridge.BackendResult(
            ok=True,
            request_id="req_status",
            status_code=200,
            data={"ok": True, "taskId": task_id},
        ),
    )
    monkeypatch.setattr(runtime, "_resync_terminal_remote_task", lambda task_id: resynced.append(task_id))

    result = runtime._report_runtime_task_status(
        "task-1",
        {"status": "completed", "summary": "Hazır."},
    )

    assert result is not None
    assert result.ok is True
    assert resynced == ["task-1"]


def test_deterministic_router_handles_system_info_without_cloud(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    monkeypatch.setattr(
        bridge,
        "run_capability",
        lambda tool, args, _state: {
            "ok": True,
            "tool": tool,
            "output": "Saat: 10:15",
            "error": None,
        },
    )
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_time",
            "taskId": "task_time",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "saat kaç"},
        }
    )

    result = response["result"]
    assert response["ok"] is True
    assert result["assistantMessage"] == "Saat: 10:15"
    assert result["provider"] == "local_tool"
    assert result["toolEvents"][0]["tool"] == "sys_info"


def test_shell_route_returns_plan_preview_instead_of_direct_execution(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_shell",
            "taskId": "task_shell",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "terminal whoami çalıştır"},
        }
    )

    result = response["result"]
    assert response["ok"] is True
    assert result["chatOk"] is True
    assert result["needsConfirmation"] is True
    assert result["executionMode"] == "plan_preview"
    assert result["pendingPlanId"]
    assert "whoami" in result["assistantMessage"]


def test_direct_capability_uses_registry_policy(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_direct_shell",
            "taskId": "task_direct_shell",
            "capability": "shell_run",
            "payload": {"command": "uptime"},
        }
    )

    assert response["ok"] is False
    assert response["error"]["code"] == "PERMISSION_REQUIRED"


def test_browser_route_requires_explicit_permission(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_browser_permission",
            "taskId": "task_browser_permission",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "google da github ara"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["permissionNeeded"] is True
    assert response["result"]["assistantMessage"] == "Tam yetki kapalı. Tarayıcı ve medya işlemleri için önce Ayarlar > Gizlilik bölümünden tam yetkiyi aç."
    assert response["result"]["permissionKey"] == "allow_browser_control"
    assert response["result"]["canGrantPersistently"] is True
    assert response["result"]["systemPermissionKey"] == "accessibility"


def test_screen_route_requires_explicit_permission(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_screen_permission",
            "taskId": "task_screen_permission",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "ekranı analiz et"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["permissionNeeded"] is True
    assert response["result"]["assistantMessage"] == "Tam yetki kapalı. Ekran analizi için önce Ayarlar > Gizlilik bölümünden tam yetkiyi aç."
    assert response["result"]["permissionKey"] == "allow_screen_analysis"


def test_browser_route_surfaces_os_permission_requirement_when_runtime_permission_is_enabled(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    snapshot_path = _write_native_desktop_snapshot(tmp_path)
    monkeypatch.setenv("ELYAN_DESKTOP_NATIVE_STATE_PATH", str(snapshot_path))
    state_store.update_state(
        {
            "account": {"dangerousAreaEnabled": True},
            "permissions": {"allow_browser_control": True},
        }
    )
    runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(
        bridge,
        "_execute_capability_with_preprocessing",
        lambda capability, _args, _state, source="": (
            {
                "ok": False,
                "tool": capability,
                "output": "Spotify otomasyonu icin erisilebilirlik izni gerekiyor.",
                "error": {
                    "code": "PERMISSION_REQUIRED",
                    "message": "Spotify otomasyonu icin erisilebilirlik izni gerekiyor.",
                },
            },
            [],
        ),
    )

    response = runtime.handle(
        {
            "id": "req_os_permission",
            "taskId": "task_os_permission",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "youtube'dan müslüm gürses çal"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["permissionNeeded"] is True
    assert response["result"]["permissionErrorCode"] == "OS_PERMISSION_REQUIRED"
    assert response["result"]["osPermissionStatus"] == "required"
    assert response["result"]["assistantMessage"] == "macOS erişilebilirlik izni kapalı. Ayarlar > Gizlilik bölümünden erişilebilirliği açıp tekrar dene."


def test_play_media_direct_capability_requires_browser_permission(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_play_media",
            "taskId": "task_play_media",
            "capability": "play_media",
            "payload": {"query": "lofi", "provider": "youtube"},
        }
    )

    assert response["ok"] is False
    assert response["error"]["code"] == "PERMISSION_REQUIRED"


def test_personal_action_direct_capability_requires_confirmation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "account": {"dangerousAreaEnabled": True},
            "permissions": {"allow_personal_actions": True},
        }
    )

    response = runtime.handle(
        {
            "id": "req_calendar_direct",
            "taskId": "task_calendar_direct",
            "capability": "add_calendar_event",
            "payload": {"title": "Demo", "start_iso": "2026-05-21T10:00"},
        }
    )

    assert response["ok"] is False
    assert response["error"]["code"] == "PERMISSION_REQUIRED"


def test_semantic_whatsapp_route_returns_pending_plan_with_permission(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "account": {"dangerousAreaEnabled": True},
            "permissions": {"allow_personal_actions": True},
        }
    )
    monkeypatch.setattr(
        bridge,
        "_semantic_route",
        lambda *_args, **_kwargs: {
            "intent": "send_whatsapp_message",
            "capability": "send_whatsapp_message",
            "args": {
                "message": "Merhaba",
                "phone_number": "+905551112233",
            },
            "confidence": 0.86,
            "provider": "semantic_planner",
            "privacyClass": "local_private",
        },
    )
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_whatsapp_plan",
            "taskId": "task_whatsapp_plan",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "Ayşe'ye WhatsApp'tan merhaba yaz"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["chatOk"] is True
    assert response["result"]["needsConfirmation"] is True
    assert response["result"]["executionMode"] == "plan_preview"
    assert response["result"]["pendingPlanId"]


def test_speech_capture_requires_direct_ui_gesture(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_speech_capture",
            "taskId": "task_speech_capture",
            "capability": "speech.capture",
            "payload": {"action": "start", "_uiGesture": False},
        }
    )

    assert response["ok"] is False
    assert response["error"]["code"] == "PERMISSION_REQUIRED"


def test_speech_bridge_requests_return_structured_result(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    monkeypatch.setattr(
        bridge,
        "run_capability",
        lambda tool, args, _state: {
            "ok": True,
            "tool": tool,
            "output": "Transkript hazır",
            "result": {
                "kind": "speech_to_text",
                "text": "Merhaba Elyan",
                "language": "tr",
                "durationMs": 1200,
                "segments": [{"start": 0.0, "end": 1.2, "text": "Merhaba Elyan"}],
                "audioPath": args.get("audioPath", ""),
            },
            "artifacts": [],
            "error": None,
        }
        if tool == "speech_to_text"
        else {
            "ok": False,
            "tool": tool,
            "output": "unexpected",
            "error": {"code": "UNEXPECTED", "message": "unexpected"},
        },
    )

    response = runtime.handle(
        {
            "id": "req_speech_transcribe",
            "taskId": "task_speech_transcribe",
            "capability": "speech.transcribe",
            "payload": {"audioPath": "/tmp/sample.wav", "languageHint": "tr"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["kind"] == "speech_to_text"
    assert response["result"]["text"] == "Merhaba Elyan"
    assert response["events"][0]["tool"] == "speech_to_text"


def test_direct_document_write_requires_confirmation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_direct_docx",
            "taskId": "task_direct_docx",
            "capability": "document_write",
            "payload": {"prompt": "test", "outputPath": "notes/test.docx"},
        }
    )

    assert response["ok"] is False
    assert response["error"]["code"] == "PERMISSION_REQUIRED"


def test_pdf_report_request_routes_to_research_then_pdf_canvas() -> None:
    routed = bridge.route_text_to_tool(
        "Masaüstümde Atatürk hakkında kaynaklı, düzenli, yaklaşık 4 sayfalık bir araştırma raporu hazırla. "
        "Başlık, kısa giriş, ana bölümler, sonuç ve kaynakça olsun. "
        "Raporu PDF olarak oluştur ve ataturk_arastirma_raporu.pdf adıyla kaydet."
    )

    assert routed is not None
    assert routed.tool_name == "canvas_write"
    assert routed.intent == "pdf_report"
    assert routed.requires_confirmation is True
    assert routed.plan_preview is not None
    steps = routed.plan_preview["steps"]
    assert [step["capability"] for step in steps] == ["web_research", "canvas_write"]
    assert steps[1]["args"]["outputFormat"] == "pdf"
    assert steps[1]["args"]["width"] == 595
    assert steps[1]["args"]["height"] == 842
    assert steps[1]["args"]["outputPath"].endswith("ataturk_arastirma_raporu.pdf")
    assert Path(steps[1]["args"]["outputPath"]).parent == Path.home() / "Desktop"


def test_pdf_report_honors_location_mentioned_at_end_of_request() -> None:
    routed = bridge.route_text_to_tool(
        "Atatürk hakkında kaynaklı 4 sayfalık araştırma raporu hazırla ve PDF olarak Masaüstüne kaydet"
    )

    assert routed is not None
    assert routed.intent == "pdf_report"
    assert Path(routed.args["outputPath"]).parent == Path.home() / "Desktop"


def test_budget_pipeline_routes_to_spreadsheet_chart_and_pdf() -> None:
    routed = bridge.route_text_to_tool(
        "Masaüstümde 2026 ilk 6 ay için örnek kişisel bütçe analizi hazırla. "
        "Gelir, kira, market, ulaşım, fatura, eğitim ve birikim kategorileri olsun. "
        "Önce verileri tablo halinde Excel dosyasına kaydet, sonra aynı veriden kategori bazlı grafik üret, "
        "son olarak bulguları kısa bir PDF raporda özetle. "
        "Dosya adları butce_analizi.xlsx, butce_grafigi.png ve butce_raporu.pdf olsun."
    )

    assert routed is not None
    assert routed.intent == "data_artifact_pipeline"
    assert routed.requires_confirmation is True
    assert routed.plan_preview is not None
    steps = routed.plan_preview["steps"]
    assert [step["capability"] for step in steps] == ["spreadsheet_write", "chart_generate", "canvas_write"]
    assert steps[0]["args"]["outputPath"].endswith("butce_analizi.xlsx")
    assert steps[1]["args"]["path"].endswith("butce_analizi.xlsx")
    assert steps[1]["args"]["outputPath"].endswith("butce_grafigi.png")
    assert steps[2]["args"]["sourcePath"].endswith("butce_grafigi.png")
    assert steps[2]["args"]["outputPath"].endswith("butce_raporu.pdf")
    assert Path(steps[0]["args"]["outputPath"]).parent == Path.home() / "Desktop"
    assert steps[0]["args"]["columns"] == ["Ay", "Kategori", "Tür", "Tutar"]
    assert len(steps[0]["args"]["rows"]) == 42


def test_budget_spreadsheet_can_feed_chart_output(tmp_path: Path) -> None:
    from actions.chart_generate import chart_generate
    from actions.spreadsheet_write import spreadsheet_write

    sheet_path = tmp_path / "butce_analizi.xlsx"
    chart_path = tmp_path / "butce_grafigi.png"
    routed = task_router._data_artifact_pipeline_route(
        "2026 ilk 6 ay için örnek bütçe analizi hazırla; Excel, grafik ve PDF üret"
    )
    assert routed is not None
    writer_args = routed.steps[0]["args"]
    sheet = spreadsheet_write(
        prompt=str(writer_args["prompt"]),
        output_path=str(sheet_path),
        title="butce_analizi",
        columns=writer_args["columns"],
        rows=writer_args["rows"],
        overwrite=False,
    )
    chart = chart_generate(
        str(sheet_path),
        chartType="bar",
        xColumn="Kategori",
        yColumn="Tutar",
        title="Bütçe Analizi",
        outputPath=str(chart_path),
        _selectedPaths=[str(sheet_path)],
    )

    assert sheet_path.exists()
    assert chart_path.exists()
    assert sheet["result"]["rowCount"] == 42
    assert sheet["result"]["columns"] == ["Ay", "Kategori", "Tür", "Tutar"]
    assert chart["result"]["outputPath"] == str(chart_path)
    assert chart["result"]["aggregation"] == "sum"
    assert chart["result"]["dataPointCount"] == 7
    assert chart["artifacts"][0]["path"] == str(chart_path)


def test_spreadsheet_writer_does_not_invent_rows_from_prompt(tmp_path: Path) -> None:
    from actions.spreadsheet_write import spreadsheet_write

    result = spreadsheet_write(
        prompt="Kişisel bütçe: gelir, kira ve market",
        output_path=str(tmp_path / "plain.xlsx"),
    )

    assert result["result"]["columns"] == []
    assert result["result"]["rowCount"] == 0


def test_canvas_write_expands_pdf_report_prompt_to_paginated_report(tmp_path: Path) -> None:
    from actions.canvas_write import canvas_write

    output = tmp_path / "ataturk_arastirma_raporu.pdf"
    result = canvas_write(
        prompt="Atatürk hakkında kaynakçalı, düzenli, yaklaşık 4 sayfalık araştırma raporu hazırla.",
        title="Atatürk Araştırma Raporu",
        output_path=str(output),
        output_format="pdf",
        width=595,
        height=842,
        source_context=(
            '[step_1] {"sources":[{"title":"Atatürk Araştırma Merkezi",'
            '"url":"https://example.test/ataturk","summary":"Kurumsal kaynak özeti"}]}'
        ),
        overwrite=False,
    )

    import fitz  # type: ignore[reportMissingImports]

    doc = fitz.open(output)
    try:
        text = "\n".join(page.get_text() for page in doc)
        assert doc.page_count == 4
        assert "Milli Mücadele" in text
        assert "Kaynakça" in text
        assert "Araştırma" in text
        assert "https://example.test/ataturk" in text
        assert "■" not in text
        assert "yaklaşık 4 sayfalık araştırma raporu hazırla" not in text
    finally:
        doc.close()
    assert result["result"]["outputFormat"] == "pdf"


def test_direct_image_generate_requires_confirmation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_direct_image",
            "taskId": "task_direct_image",
            "capability": "image_generate",
            "payload": {"prompt": "Sade bir çalışma yüzeyi", "outputPath": "images/clean.png"},
        }
    )

    assert response["ok"] is False
    assert response["error"]["code"] == "PERMISSION_REQUIRED"


def test_router_handles_turkish_web_search_apostrophe() -> None:
    routed = bridge.route_text_to_tool("Google’da Python öğren ara")

    assert routed is not None
    assert routed.tool_name == "browser_control"
    assert routed.args == {"action": "search", "query": "Python öğren"}


def test_router_handles_plain_google_search_spacing() -> None:
    routed = bridge.route_text_to_tool("google da github ara")

    assert routed is not None
    assert routed.tool_name == "browser_control"
    assert routed.args == {"action": "search", "query": "github"}


def test_router_requires_specific_topic_for_web_research() -> None:
    assert bridge.route_text_to_tool("kaynak ver") is None

    routed = bridge.route_text_to_tool("Atatürk hakkında araştırma yap")

    assert routed is not None
    assert routed.tool_name == "web_research"
    assert routed.args == {"query": "Atatürk"}


def test_router_handles_youtube_spacing_and_folder_alias() -> None:
    youtube = bridge.route_text_to_tool("youtube da lofi aç")
    finder = bridge.route_text_to_tool("dosya gezgini aç")

    assert youtube is not None
    assert youtube.tool_name == "browser_control"
    assert youtube.args == {"action": "play_youtube", "query": "lofi"}

    assert finder is not None
    assert finder.tool_name == "open_app"
    assert finder.args == {"app_name": "Finder"}


def test_router_handles_app_close_commands() -> None:
    routed = bridge.route_text_to_tool("Chrome'i kapat")

    assert routed is not None
    assert routed.tool_name == "close_app"
    assert routed.args == {"app_name": "Chrome"}


def test_router_strips_turkish_case_particles_for_app_open_close() -> None:
    open_route = bridge.route_text_to_tool("chrome u aç")
    close_route = bridge.route_text_to_tool("chrome u kapat")
    safari_route = bridge.route_text_to_tool("Safariyi aç")

    assert open_route is not None
    assert open_route.tool_name == "open_app"
    assert open_route.args == {"app_name": "chrome"}

    assert close_route is not None
    assert close_route.tool_name == "close_app"
    assert close_route.args == {"app_name": "chrome"}

    assert safari_route is not None
    assert safari_route.tool_name == "open_app"
    assert safari_route.args == {"app_name": "Safari"}


def test_router_handles_short_screenshot_command() -> None:
    routed = bridge.route_text_to_tool("ss al")

    assert routed is not None
    assert routed.tool_name == "desktop_operator.observe_screen"
    assert routed.args == {
        "query": "ss al",
        "target": "active_window",
        "preserveScreenshot": True,
    }


def test_router_handles_youtube_ablative_suffix() -> None:
    routed = bridge.route_text_to_tool("youtube dan müslüm gürses çal")

    assert routed is not None
    assert routed.tool_name == "browser_control"
    assert routed.args == {"action": "play_youtube", "query": "müslüm gürses"}


def test_router_handles_audio_transcription_path(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    import runtime.task_router as task_router

    audio_path = tmp_path / "meeting.wav"
    audio_path.write_bytes(b"RIFF")
    monkeypatch.setattr(task_router, "_workspace_root", lambda: tmp_path)

    routed = bridge.route_text_to_tool("meeting.wav ses kaydını yazıya çevir")

    assert routed is not None
    assert routed.tool_name == "speech_to_text"
    assert routed.args["path" if "path" in routed.args else "audioPath"] == str(audio_path)


def test_router_uses_selected_audio_artifact_for_transcription() -> None:
    routed = bridge.route_text_to_tool(
        "bunu yazıya çevir",
        selected_artifacts=[
            {
                "id": "artifact_audio",
                "name": "meeting.wav",
                "path": "/tmp/meeting.wav",
                "mimeType": "audio/wav",
                "sizeBytes": 12,
                "kind": "audio",
            }
        ],
    )

    assert routed is not None
    assert routed.tool_name == "speech_to_text"
    assert routed.args["audioPath"] == "/tmp/meeting.wav"
    assert routed.args["_selectedPaths"] == ["/tmp/meeting.wav"]


def test_router_handles_read_aloud_command() -> None:
    routed = bridge.route_text_to_tool('"Merhaba Elyan" sesli oku')

    assert routed is not None
    assert routed.tool_name == "text_to_speech"
    assert routed.args["text"] == "Merhaba Elyan"


def test_router_handles_generic_close_command() -> None:
    routed = bridge.route_text_to_tool("uygulamayı kapat")

    assert routed is not None
    assert routed.tool_name == "close_app"
    assert routed.args == {"app_name": ""}


def test_linux_open_app_route_launches_native_file_manager(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """open_app artık çapraz platform: Linux'ta Finder → nautilus/dolphin...
    aday ikilisi PATH'te aranıp detached başlatılır; bulunamazsa güvenli
    CAPABILITY_UNAVAILABLE mesajı döner (asla 'yalnız macOS' reddi değil)."""
    _isolate_state(monkeypatch, tmp_path)
    import actions.open_app as open_app_module

    monkeypatch.setattr(open_app_module.sys, "platform", "linux")
    launched: list[list[str]] = []
    monkeypatch.setattr(open_app_module.shutil, "which", lambda name: "/usr/bin/nautilus" if name == "nautilus" else None)
    monkeypatch.setattr(open_app_module, "_spawn_detached", lambda cmd: launched.append(cmd))
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_open_app_linux",
            "taskId": "task_open_app_linux",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "finder ac"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["chatOk"] is True
    assert launched == [["nautilus"]]


def test_linux_open_app_without_binary_fails_safe(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    import actions.open_app as open_app_module

    monkeypatch.setattr(open_app_module.sys, "platform", "linux")
    monkeypatch.setattr(open_app_module.shutil, "which", lambda name: None)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_open_app_linux_missing",
            "taskId": "task_open_app_linux_missing",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "finder ac"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["chatOk"] is False
    assert "bulunamadi" in response["result"]["assistantMessage"]


def test_router_builds_calendar_plan_preview() -> None:
    routed = bridge.route_text_to_tool("takvime cuma 14:00 ürün toplantısı ekle")

    assert routed is not None
    assert routed.tool_name == "add_calendar_event"
    assert routed.requires_confirmation is True
    assert routed.plan_preview is not None
    assert "ürün toplantısı" in routed.args["title"].lower()


def test_router_builds_reminder_plan_preview() -> None:
    routed = bridge.route_text_to_tool("yarın sabah 9da annemi ara diye hatırlatıcı ekle")

    assert routed is not None
    assert routed.tool_name == "add_reminder"
    assert routed.requires_confirmation is True
    assert routed.plan_preview is not None
    assert "annemi ara" in routed.args["title"].lower()


def test_router_builds_multi_step_browser_plan() -> None:
    routed = bridge.route_text_to_tool("safari aç ve openai sitesini aç")

    assert routed is not None
    assert routed.is_multi_step is True
    assert routed.requires_confirmation is True
    assert len(routed.steps) == 2


def test_router_builds_ocr_route_for_selected_image() -> None:
    routed = bridge.route_text_to_tool(
        "bu görseldeki yazıyı oku",
        selected_artifacts=[
            {
                "id": "artifact_image",
                "name": "scan.png",
                "path": "/tmp/scan.png",
                "mimeType": "image/png",
                "sizeBytes": 42,
                "kind": "image",
            }
        ],
    )

    assert routed is not None
    assert routed.tool_name == "ocr_read"
    assert routed.args["path"] == "/tmp/scan.png"
    assert routed.args["mode"] == "read"
    assert routed.args["_selectedPaths"] == ["/tmp/scan.png"]


def test_router_builds_document_route_for_selected_pdf() -> None:
    routed = bridge.route_text_to_tool(
        "bunu özetle",
        selected_artifacts=[
            {
                "id": "artifact_pdf",
                "name": "notes.pdf",
                "path": "/tmp/notes.pdf",
                "mimeType": "application/pdf",
                "sizeBytes": 128,
                "kind": "document",
            }
        ],
    )

    assert routed is not None
    assert routed.tool_name == "document_read"
    assert routed.args["path"] == "/tmp/notes.pdf"
    assert routed.args["mode"] == "summary"
    assert routed.args["_selectedPaths"] == ["/tmp/notes.pdf"]


def test_router_builds_image_read_route_for_selected_image() -> None:
    routed = bridge.route_text_to_tool(
        "bu görseli incele",
        selected_artifacts=[
            {
                "id": "artifact_image",
                "name": "poster.png",
                "path": "/tmp/poster.png",
                "mimeType": "image/png",
                "sizeBytes": 64,
                "kind": "image",
            }
        ],
    )

    assert routed is not None
    assert routed.tool_name == "image_read"
    assert routed.args["path"] == "/tmp/poster.png"
    assert routed.args["mode"] == "summary"
    assert routed.args["_selectedPaths"] == ["/tmp/poster.png"]


def test_selected_image_attributes_stay_single_read_only_task() -> None:
    artifact = {
        "id": "artifact_image",
        "name": "poster.png",
        "path": "/tmp/poster.png",
        "mimeType": "image/png",
        "kind": "image",
    }

    routed = bridge.route_text_to_tool(
        "Bu görseli incele; boyut, biçim ve temel renk bilgisini çıkar",
        selected_artifacts=[artifact],
    )

    assert routed is not None
    assert routed.tool_name == "image_read"
    assert routed.is_multi_step is False
    assert routed.requires_confirmation is False
    assert routed.args["mode"] == "palette"


def test_selected_image_color_extraction_is_not_misrouted_to_ocr() -> None:
    routed = bridge.route_text_to_tool(
        "temel renk bilgisini çıkar",
        selected_artifacts=[{
            "id": "artifact_image",
            "name": "poster.png",
            "path": "/tmp/poster.png",
            "mimeType": "image/png",
            "kind": "image",
        }],
    )

    assert routed is not None
    assert routed.tool_name == "image_read"


def test_router_builds_selected_pdf_to_presentation_pipeline() -> None:
    routed = bridge.route_text_to_tool(
        "bu PDF dosyasını 6 slaytlık PowerPoint sunumuna dönüştür ve /tmp/report-deck.pptx yoluna kaydet",
        selected_artifacts=[
            {
                "id": "report",
                "name": "report.pdf",
                "path": "/tmp/report.pdf",
                "mimeType": "application/pdf",
                "kind": "document",
            }
        ],
    )

    assert routed is not None
    assert routed.intent == "document_transform"
    assert routed.requires_confirmation is True
    assert [step["capability"] for step in routed.steps] == ["document_read", "presentation_write"]
    assert routed.steps[0]["args"]["_selectedPaths"] == ["/tmp/report.pdf"]
    assert routed.steps[1]["args"]["outputPath"] == "/private/tmp/report-deck.pptx"


def test_router_builds_data_analyze_route_for_selected_csv() -> None:
    routed = bridge.route_text_to_tool(
        "bu csvyi analiz et",
        selected_artifacts=[
            {
                "id": "artifact_csv",
                "name": "sales.csv",
                "path": "/tmp/sales.csv",
                "mimeType": "text/csv",
                "sizeBytes": 96,
                "kind": "document",
            }
        ],
    )

    assert routed is not None
    assert routed.tool_name == "data_analyze"
    assert routed.args["path"] == "/tmp/sales.csv"
    assert routed.args["mode"] == "summary"
    assert routed.args["_selectedPaths"] == ["/tmp/sales.csv"]


def test_router_builds_data_analyze_route_for_explicit_xlsx_path() -> None:
    routed = bridge.route_text_to_tool(
        "/tmp/elyan-release/butce.xlsx dosyasını analiz et ve özet istatistikleri çıkar"
    )

    assert routed is not None
    assert routed.tool_name == "data_analyze"
    assert routed.args["path"].endswith("/tmp/elyan-release/butce.xlsx")
    assert routed.args["mode"] == "profile"


def test_router_builds_chart_generate_route_for_selected_csv() -> None:
    routed = bridge.route_text_to_tool(
        "bundan histogram grafik çıkar",
        selected_artifacts=[
            {
                "id": "artifact_csv",
                "name": "sales.csv",
                "path": "/tmp/sales.csv",
                "mimeType": "text/csv",
                "sizeBytes": 96,
                "kind": "document",
            }
        ],
    )

    assert routed is not None
    assert routed.tool_name == "chart_generate"
    assert routed.args["path"] == "/tmp/sales.csv"
    assert routed.args["chartType"] == "histogram"
    assert routed.args["_selectedPaths"] == ["/tmp/sales.csv"]


def test_router_builds_math_route() -> None:
    routed = bridge.route_text_to_tool("x^2 - 5x + 6’yı çarpanlara ayır")

    assert routed is not None
    assert routed.tool_name == "math_solve"
    assert routed.args["mode"] == "factor"
    assert "x^2 - 5x + 6" in routed.args["expression"]


def test_router_builds_latex_math_route() -> None:
    routed = bridge.route_text_to_tool(r"\frac{x+1}{2} ifadesini sadeleştir")

    assert routed is not None
    assert routed.tool_name == "math_solve"
    assert routed.args["mode"] == "simplify"
    assert routed.args["_latexInput"] == r"\frac{x+1}{2}"


def test_router_builds_spreadsheet_write_plan() -> None:
    routed = bridge.route_text_to_tool("xlsx tablo oluştur")

    assert routed is not None
    assert routed.tool_name == "spreadsheet_write"
    assert routed.requires_confirmation is True
    assert routed.plan_preview is not None
    assert routed.args["outputPath"].endswith(".xlsx")


def test_public_research_spreadsheet_uses_structured_sources(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr("runtime.task_router._resolve_location_path", lambda _text: str(tmp_path))

    routed = bridge.route_text_to_tool(
        "Masaüstümde 2026 Türkiye yangınlarını raporla excel tablosu oluştur"
    )

    assert routed is not None
    assert routed.intent == "research_spreadsheet"
    assert routed.is_multi_step is True
    assert [step["capability"] for step in routed.steps] == ["web_research", "spreadsheet_write"]
    assert routed.steps[0]["args"]["query"] == "2026 Türkiye yangınlarını"
    writer_args = routed.steps[1]["args"]
    assert writer_args["columns"] == ["Başlık", "URL", "Özet"]
    assert writer_args["rows"] == "{{steps.step_1.result.sources}}"
    assert Path(writer_args["outputPath"]).parent == tmp_path


def test_npm_bootstrap_and_tool_contract_expose_structured_data_stack() -> None:
    root = Path(__file__).resolve().parents[1]
    core_requirements = (root / "requirements-core.txt").read_text(encoding="utf-8").splitlines()
    launcher = (root / "bin" / "elyan.js").read_text(encoding="utf-8")
    shell_installer = (root / "scripts" / "install.sh").read_text(encoding="utf-8")

    assert "openpyxl==3.1.5" in core_requirements
    assert "openpyxl" in launcher
    for optional_package in ("pandas", "matplotlib", "reportlab"):
        assert not any(item.startswith(optional_package) for item in core_requirements)
    assert "PACKAGE_VERSION" in launcher
    assert "coreReady()" in launcher
    assert "shouldRestartDaemon" in launcher
    assert "minor < 14" in launcher
    assert "shouldReplaceVenv" in launcher
    assert "detached: true" not in launcher
    assert "install_extras.py" in launcher
    repair_block = launcher[launcher.index("if (shouldRepair)") :]
    assert repair_block.index("shouldRestartDaemon && !run") < repair_block.index("bootstrap({ recreateVenv")
    assert "Uyumsuz Python sanal ortamı yenileniyor" in shell_installer


def test_private_spreadsheet_report_does_not_trigger_web_research() -> None:
    routed = bridge.route_text_to_tool("2026 bütçemi raporla excel tablosu oluştur")

    assert routed is not None
    assert routed.intent == "spreadsheet_write"



def test_spreadsheet_write_neutralizes_formula_cells_and_preserves_numbers(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    import actions.spreadsheet_write as spreadsheet_module
    from openpyxl import load_workbook

    monkeypatch.setattr(spreadsheet_module, "_workspace_root", lambda: tmp_path)
    output = tmp_path / "safe.xlsx"
    spreadsheet_module.spreadsheet_write(
        prompt="Güvenli veri",
        output_path=str(output),
        columns=["Başlık", "Sayı"],
        rows=[["=HYPERLINK(\"https://example.invalid\",\"x\")", -42]],
    )

    values = list(load_workbook(output, read_only=True).active.iter_rows(values_only=True))
    assert values[1][0].startswith("'=")
    assert values[1][1] == -42


def test_spreadsheet_write_infers_and_aligns_dictionary_columns(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import actions.spreadsheet_write as spreadsheet_module
    from openpyxl import load_workbook

    monkeypatch.setattr(spreadsheet_module, "_workspace_root", lambda: tmp_path)
    output = tmp_path / "aligned.xlsx"
    result = spreadsheet_module.spreadsheet_write(
        output_path=str(output),
        rows=[
            {"name": "first", "score": 10},
            {"score": 20, "name": "second"},
        ],
    )

    workbook = load_workbook(output, read_only=False)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    assert values == [
        ("name", "score"),
        ("first", 10),
        ("second", 20),
    ]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:B3"
    assert sheet.tables
    assert sheet["A1"].font.bold is True
    assert sheet.column_dimensions["A"].width >= 10
    assert result["result"]["columns"] == ["name", "score"]
    assert result["result"]["rowCount"] == 2


def test_router_builds_presentation_write_plan() -> None:
    routed = bridge.route_text_to_tool("pptx sunumu oluştur")

    assert routed is not None
    assert routed.tool_name == "presentation_write"
    assert routed.requires_confirmation is True
    assert routed.plan_preview is not None
    assert routed.args["outputPath"].endswith(".pptx")


def test_router_builds_canvas_write_plan() -> None:
    routed = bridge.route_text_to_tool("canvas ile tablo ve görsel düzeni oluştur")

    assert routed is not None
    assert routed.tool_name == "canvas_write"
    assert routed.requires_confirmation is True
    assert routed.plan_preview is not None
    assert routed.args["outputPath"].endswith(".pdf")


def test_conversation_send_returns_pending_plan_for_side_effecting_schedule(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "account": {"dangerousAreaEnabled": True},
            "permissions": {"allow_personal_actions": True},
        }
    )
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_plan",
            "taskId": "task_plan",
            "capability": "conversation.send",
            "payload": {
                "conversationId": "",
                "text": "takvime cuma 14:00 ürün toplantısı ekle",
            },
        }
    )

    result = response["result"]
    assert response["ok"] is True
    assert result["chatOk"] is True
    assert result["needsConfirmation"] is True
    assert result["executionMode"] == "plan_preview"
    assert result["pendingPlanId"]


def test_conversation_send_passes_structured_result_to_ui_state(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(
        bridge,
        "route_text_to_tool",
        lambda _text, **_kwargs: bridge.RoutedTask(
            "math_solve",
            {"expression": "2 + 2", "mode": "evaluate"},
            "math_solve",
            intent="math_solve",
            confidence=0.96,
        ),
    )
    monkeypatch.setattr(
        bridge,
        "run_capability",
        lambda _capability, _args, _state: {
            "ok": True,
            "tool": "math_solve",
            "output": "Sayısal sonuç: 4",
            "result": {
                "kind": "math_solve",
                "expression": "2 + 2",
                "mode": "evaluate",
                "result": "4",
            },
            "artifacts": [],
            "error": None,
        },
    )

    response = runtime.handle(
        {
            "id": "req_structured",
            "taskId": "task_structured",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "2 + 2 hesapla"},
        }
    )

    result = response["result"]
    assert result["structuredResult"]["kind"] == "math_solve"
    active = state_store.get_active_conversation()
    assert active is not None
    assert active["messages"][-1]["structuredResult"]["mode"] == "evaluate"


def test_conversation_send_runs_latex_parse_before_math_solve(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    calls: list[tuple[str, dict]] = []

    monkeypatch.setattr(
        bridge,
        "route_text_to_tool",
        lambda _text, **_kwargs: bridge.RoutedTask(
            "math_solve",
            {
                "expression": r"\frac{x+1}{2}",
                "mode": "simplify",
                "_latexInput": r"\frac{x+1}{2}",
            },
            "math_solve",
            intent="math_solve",
            confidence=0.94,
        ),
    )

    def fake_run_capability(capability: str, args: dict, _state: dict) -> dict:
        calls.append((capability, dict(args)))
        if capability == "latex_parse":
            return {
                "ok": True,
                "tool": "latex_parse",
                "output": "LaTeX normalize edildi",
                "result": {
                    "kind": "latex_parse",
                    "normalizedExpression": "(x + 1)/2",
                    "latex": r"\frac{x + 1}{2}",
                },
                "artifacts": [],
                "error": None,
            }
        if capability == "math_solve":
            return {
                "ok": True,
                "tool": "math_solve",
                "output": "Sadeleştirilmiş ifade: x/2 + 1/2",
                "result": {
                    "kind": "math_solve",
                    "expression": "(x + 1)/2",
                    "mode": "simplify",
                    "result": "x/2 + 1/2",
                },
                "artifacts": [],
                "error": None,
            }
        return {
            "ok": False,
            "tool": capability,
            "output": "unexpected",
            "error": {"code": "UNEXPECTED", "message": "unexpected"},
        }

    monkeypatch.setattr(bridge, "run_capability", fake_run_capability)

    response = runtime.handle(
        {
            "id": "req_latex_chain",
            "taskId": "task_latex_chain",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": r"\frac{x+1}{2} ifadesini sadeleştir"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["structuredResult"]["kind"] == "math_solve"
    assert response["result"]["structuredResult"]["latexParse"]["normalizedExpression"] == "(x + 1)/2"
    assert [item[0] for item in calls] == ["latex_parse", "math_solve"]
    assert calls[1][1]["expression"] == "(x + 1)/2"


def test_conversation_send_routes_selected_artifact_and_clears_composer_state(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    captured: dict[str, object] = {}
    image_path = tmp_path / "poster.png"
    image_path.write_bytes(b"test-image")

    def fake_run(capability: str, args: dict[str, object], _state: dict[str, object]) -> dict[str, object]:
        captured["capability"] = capability
        captured["args"] = dict(args)
        return {
            "ok": True,
            "tool": capability,
            "output": "image_read:poster.png",
            "result": {"kind": "image_read", "sourcePath": str(image_path)},
            "artifacts": [],
            "error": None,
        }

    monkeypatch.setattr(bridge, "run_capability", fake_run)

    response = runtime.handle(
        {
            "id": "req_selected_artifact",
            "taskId": "task_selected_artifact",
            "capability": "conversation.send",
            "payload": {
                "conversationId": "",
                "text": "bu görseli incele",
                "selectedArtifacts": [
                    {
                        "id": "artifact_image",
                        "name": "poster.png",
                        "path": str(image_path),
                        "mimeType": "image/png",
                        "sizeBytes": 64,
                        "kind": "image",
                    }
                ],
            },
        }
    )

    assert response["ok"] is True
    assert response["result"]["chatOk"] is True
    assert captured["capability"] == "image_read"
    captured_args = captured["args"]
    assert isinstance(captured_args, dict)
    assert captured_args["path"] == str(image_path)
    assert captured_args["mode"] == "summary"
    assert captured_args["_selectedPaths"] == [str(image_path)]
    assert captured_args["_confirmed"] is False
    assert response["result"]["state"]["composer"]["selectedArtifacts"] == []


def test_runtime_status_includes_speech_truth_surface(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_runtime_status",
            "taskId": "task_runtime_status",
            "capability": "runtime.status",
            "payload": {},
        }
    )

    assert response["ok"] is True
    assert "speechStatus" in response["result"]
    assert "dependencyStatus" in response["result"]
    assert "faster_whisper" in response["result"]["dependencyStatus"]
    assert "artifactSelectionStatus" in response["result"]


def test_confirm_plan_executes_local_steps(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "account": {"dangerousAreaEnabled": True},
            "permissions": {"allow_personal_actions": True},
        }
    )
    runtime = bridge.RuntimeBridge()
    sent = runtime.handle(
        {
            "id": "req_plan_confirm",
            "taskId": "task_plan_confirm",
            "capability": "conversation.send",
            "payload": {
                "conversationId": "",
                "text": "takvime cuma 14:00 ürün toplantısı ekle",
            },
        }
    )
    result = sent["result"]
    plan_id = result["pendingPlanId"]
    conversation_id = result["conversationId"]

    monkeypatch.setattr(
        bridge,
        "run_capability",
        # P3 kanıt sözleşmesi: sağlayıcı adımı provider kimliği döndürür.
        lambda tool, args, _state: {
            "ok": True,
            "tool": tool,
            "output": f"{tool}:{args.get('title', args.get('app_name', 'ok'))}",
            "result": {"kind": tool, "eventId": "evt_test_1"},
            "error": None,
        },
    )

    confirmed = runtime.handle(
        {
            "id": "req_confirm",
            "taskId": "task_confirm",
            "capability": "conversation.confirm_plan",
            "payload": {
                "conversationId": conversation_id,
                "pendingPlanId": plan_id,
                "approved": True,
            },
        }
    )

    payload = confirmed["result"]
    assert confirmed["ok"] is True
    assert payload["chatOk"] is True
    assert payload["executionMode"] == "confirmed_plan"
    assert "add_calendar_event" in payload["assistantMessage"]


def test_confirm_plan_injects_confirmation_for_file_write(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    plan = state_store.save_pending_plan(
        {
            "conversationId": "",
            "query": "xlsx tablo oluştur",
            "intent": "spreadsheet_write",
            "capability": "spreadsheet_write",
            "confidence": 0.83,
            "privacyClass": "local_private",
            "steps": [
                {
                    "capability": "spreadsheet_write",
                    "args": {"prompt": "xlsx tablo oluştur", "outputPath": "tables/test.xlsx", "overwrite": False},
                }
            ],
        }
    )
    conversation = state_store.create_conversation("")
    captured: dict[str, object] = {}

    def fake_run_capability(capability: str, args: dict[str, object], _state: dict[str, object]) -> dict[str, object]:
        captured["capability"] = capability
        captured["confirmed"] = args.get("_confirmed")
        output_path = (tmp_path / "tables" / "test.xlsx").resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(b"verified")
        return {
            "ok": True,
            "tool": capability,
            "output": "spreadsheet_write:test.xlsx",
            "result": {
                "kind": "spreadsheet_write",
                "outputPath": str(output_path),
                "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "created": True,
                "sourceContext": "prompt",
            },
            "artifacts": [
                {
                    "kind": "file",
                    "name": "test.xlsx",
                    "path": str(output_path),
                }
            ],
            "error": None,
        }

    monkeypatch.setattr(bridge, "run_capability", fake_run_capability)

    confirmed = runtime.handle(
        {
            "id": "req_confirm_write",
            "taskId": "task_confirm_write",
            "capability": "conversation.confirm_plan",
            "payload": {
                "conversationId": conversation["id"],
                "pendingPlanId": plan["id"],
                "approved": True,
            },
        }
    )

    assert confirmed["ok"] is True
    assert confirmed["result"]["chatOk"] is True
    assert captured["capability"] == "spreadsheet_write"
    assert captured["confirmed"] is True
    assert confirmed["result"]["structuredResult"]["kind"] == "spreadsheet_write"
    assert confirmed["result"]["artifacts"][0]["name"] == "test.xlsx"


def test_confirm_plan_fails_closed_when_capability_not_ready(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    plan = state_store.save_pending_plan(
        {
            "conversationId": "",
            "query": "sunum hazırla",
            "intent": "presentation_write",
            "capability": "presentation_write",
            "confidence": 0.83,
            "privacyClass": "local_private",
            "steps": [
                {
                    "capability": "presentation_write",
                    "args": {"prompt": "sunum hazırla", "outputPath": "decks/test.pptx", "overwrite": False},
                }
            ],
        }
    )
    conversation = state_store.create_conversation("")
    called = {"run": False}

    monkeypatch.setattr(
        bridge,
        "capability_readiness",
        lambda capability, **_kwargs: {
            "available": False,
            "ready": False,
            "errorCode": "DEPENDENCY_UNAVAILABLE",
            "missingDependencies": ["python_pptx"],
        }
        if capability == "presentation_write"
        else {
            "available": True,
            "ready": True,
            "errorCode": "",
            "missingDependencies": [],
        },
    )

    def fake_run_capability(_capability: str, _args: dict[str, object], _state: dict[str, object]) -> dict[str, object]:
        called["run"] = True
        return {
            "ok": True,
            "tool": "presentation_write",
            "output": "should not run",
            "result": {"kind": "presentation_write"},
            "artifacts": [],
            "error": None,
        }

    monkeypatch.setattr(bridge, "run_capability", fake_run_capability)

    confirmed = runtime.handle(
        {
            "id": "req_confirm_unready",
            "taskId": "task_confirm_unready",
            "capability": "conversation.confirm_plan",
            "payload": {
                "conversationId": conversation["id"],
                "pendingPlanId": plan["id"],
                "approved": True,
            },
        }
    )

    assert confirmed["ok"] is True
    assert confirmed["result"]["chatOk"] is False
    assert confirmed["result"]["error"]["code"] == "DEPENDENCY_UNAVAILABLE"
    assert called["run"] is False


def test_follow_up_closes_last_opened_app(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    monkeypatch.setattr(
        bridge,
        "run_capability",
        lambda tool, args, _state: {
            "ok": True,
            "tool": tool,
            "output": f"{tool}:{args.get('app_name', '')}",
            "result": {
                "appName": args.get("app_name", ""),
                "foregroundConfirmed": tool == "open_app",
                "closedConfirmed": tool == "close_app",
            },
            "error": None,
        },
    )

    first = runtime.handle(
        {
            "id": "req_open",
            "taskId": "task_open",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "Safari aç"},
        }
    )
    conversation_id = first["result"]["conversationId"]

    second = runtime.handle(
        {
            "id": "req_close_followup",
            "taskId": "task_close_followup",
            "capability": "conversation.send",
            "payload": {"conversationId": conversation_id, "text": "onu kapat"},
        }
    )

    assert second["result"]["chatOk"] is True
    assert second["result"]["executionMode"] == "local_tool"
    assert second["result"]["assistantMessage"] == "close_app:Safari"


def test_focus_app_route_uses_open_app_capability(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    monkeypatch.setattr(
        bridge,
        "run_capability",
        lambda tool, args, _state: {
            "ok": True,
            "tool": tool,
            "output": f"{tool}:{args.get('app_name', '')}",
            "result": {
                "appName": args.get("app_name", ""),
                "foregroundConfirmed": tool == "open_app",
            },
            "error": None,
        },
    )

    response = runtime.handle(
        {
            "id": "req_focus",
            "taskId": "task_focus",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "spotify'ı öne getir"},
        }
    )

    assert response["result"]["chatOk"] is True
    assert response["result"]["executionMode"] == "local_tool"
    assert response["result"]["assistantMessage"] == "open_app:spotify"


def test_focus_command_without_explicit_target_returns_clarification(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_focus_clarify",
            "taskId": "task_focus_clarify",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "öne getir"},
        }
    )

    assert response["result"]["chatOk"] is True
    assert response["result"]["clarificationNeeded"] is True
    assert response["result"]["executionMode"] == "clarification"
    assert response["result"]["clarificationQuestion"] == "Hangi uygulamayı öne getireyim?"


def test_restart_app_route_creates_confirmation_plan(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_restart",
            "taskId": "task_restart",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "chrome'u yeniden aç"},
        }
    )

    assert response["result"]["chatOk"] is True
    assert response["result"]["needsConfirmation"] is True
    assert response["result"]["executionMode"] == "plan_preview"
    plan_preview = response["result"]["planPreview"]
    assert isinstance(plan_preview, dict)
    steps = plan_preview.get("steps", [])
    assert isinstance(steps, list)
    assert [step.get("capability") for step in steps] == ["close_app", "open_app"]


def test_pending_plan_revision_updates_schedule(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "account": {"dangerousAreaEnabled": True},
            "permissions": {"allow_personal_actions": True},
        }
    )
    runtime = bridge.RuntimeBridge()
    sent = runtime.handle(
        {
            "id": "req_plan_revise",
            "taskId": "task_plan_revise",
            "capability": "conversation.send",
            "payload": {
                "conversationId": "",
                "text": "takvime cuma 14:00 ürün toplantısı ekle",
            },
        }
    )

    revised = runtime.handle(
        {
            "id": "req_revise",
            "taskId": "task_revise",
            "capability": "conversation.revise_plan",
            "payload": {
                "conversationId": sent["result"]["conversationId"],
                "pendingPlanId": sent["result"]["pendingPlanId"],
                "revisionText": "15:00 yap",
            },
        }
    )

    assert revised["ok"] is True
    assert revised["result"]["chatOk"] is True
    assert revised["result"]["executionMode"] == "plan_revised"
    assert "Cuma 15:00" in revised["result"]["assistantMessage"]


def test_pending_plan_cancel_via_follow_up(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "account": {"dangerousAreaEnabled": True},
            "permissions": {"allow_personal_actions": True},
        }
    )
    runtime = bridge.RuntimeBridge()
    sent = runtime.handle(
        {
            "id": "req_plan_cancel",
            "taskId": "task_plan_cancel",
            "capability": "conversation.send",
            "payload": {
                "conversationId": "",
                "text": "takvime cuma 14:00 ürün toplantısı ekle",
            },
        }
    )

    cancelled = runtime.handle(
        {
            "id": "req_cancel_followup",
            "taskId": "task_cancel_followup",
            "capability": "conversation.send",
            "payload": {
                "conversationId": sent["result"]["conversationId"],
                "text": "iptal et",
            },
        }
    )

    assert cancelled["ok"] is True
    assert cancelled["result"]["executionMode"] == "plan_cancelled"
    assert cancelled["result"]["assistantMessage"] == "İşlem iptal edildi."


def test_low_confidence_semantic_route_returns_clarification(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(bridge, "route_text_to_tool", lambda _text, **_kwargs: None)
    monkeypatch.setattr(
        bridge,
        "_semantic_route",
        lambda _state, _conversation, _text, **_kwargs: {
            "intent": "open_app",
            "capability": "open_app",
            "args": {"app_name": "Safari"},
            "confidence": 0.2,
            "requiresConfirmation": False,
            "isMultiStep": False,
            "privacyClass": "local_private",
            "planPreview": None,
            "provider": "ollama",
        },
    )

    response = runtime.handle(
        {
            "id": "req_clarify",
            "taskId": "task_clarify",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "onu aç"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["clarificationNeeded"] is True
    assert response["result"]["executionMode"] == "clarification"


def test_semantic_route_clarifies_when_document_target_is_missing(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(bridge, "route_text_to_tool", lambda _text, **_kwargs: None)
    monkeypatch.setattr(
        bridge,
        "_semantic_route",
        lambda _state, _conversation, _text, **_kwargs: {
            "intent": "document_read",
            "capability": "document_read",
            "args": {"mode": "summary"},
            "confidence": 0.92,
            "requiresConfirmation": False,
            "isMultiStep": False,
            "privacyClass": "local_private",
            "planPreview": None,
            "provider": "ollama",
        },
    )

    response = runtime.handle(
        {
            "id": "req_semantic_missing_doc",
            "taskId": "task_semantic_missing_doc",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "bu belgeyi özetle"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["clarificationNeeded"] is True
    assert "belge seç" in response["result"]["clarificationQuestion"].lower()


def test_semantic_route_clarifies_when_browser_url_is_missing(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(bridge, "route_text_to_tool", lambda _text, **_kwargs: None)
    monkeypatch.setattr(
        bridge,
        "_semantic_route",
        lambda _state, _conversation, _text, **_kwargs: {
            "intent": "browser_control",
            "capability": "browser_control",
            "args": {"action": "open_url"},
            "confidence": 0.93,
            "requiresConfirmation": False,
            "isMultiStep": False,
            "privacyClass": "public_text",
            "planPreview": None,
            "provider": "ollama",
        },
    )

    response = runtime.handle(
        {
            "id": "req_semantic_missing_url",
            "taskId": "task_semantic_missing_url",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "siteyi aç"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["clarificationNeeded"] is True
    assert response["result"]["clarificationQuestion"] == "Hangi adresi açayım?"


def test_provider_lock_semantic_candidates_do_not_escape() -> None:
    state = state_store._ensure_defaults(
        {
            "providers": {
                "routingPolicy": "provider_lock",
                "active": "openai",
                "openai": {
                    "enabled": True,
                    "apiKey": "key",
                    "baseUrl": "https://api.openai.com/v1",
                    "defaultModel": "gpt-4.1-mini",
                },
                "ollama": {
                    "enabled": True,
                    "defaultModel": "llama3.2:3b",
                },
            }
        }
    )

    assert bridge._semantic_candidate_providers(state, privacy_class="public_text") == ["openai"]


def test_public_snapshot_redacts_provider_base_urls() -> None:
    state = state_store._ensure_defaults(
        {
            "providers": {
                "openai": {
                    "enabled": True,
                    "apiKey": "key",
                    "baseUrl": "https://api.openai.com/v1",
                    "defaultModel": "gpt-4.1-mini",
                },
                "ollama": {
                    "enabled": True,
                    "baseUrl": "http://127.0.0.1:11434",
                    "defaultModel": "llama3.1:8b",
                },
            }
        }
    )

    public = state_store.public_snapshot(state)

    assert public["providers"]["openai"]["baseUrl"] == ""
    assert public["providers"]["ollama"]["baseUrl"] == ""


def test_provider_catalog_payload_hides_base_urls(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(
        bridge.RuntimeBridge,
        "local_models_status",
        lambda self: {
            "status": {
                "available": True,
                "reachable": True,
                "configured": True,
                "baseUrl": "http://127.0.0.1:11434",
                "defaultModel": "llama3.1:8b",
                "latencyMs": 12,
                "lastCheckedAt": "2026-06-03T12:00:00Z",
                "errorCode": "",
                "jobs": [],
            },
            "models": {"ok": True, "available": True, "models": [{"name": "llama3.1:8b"}]},
            "jobs": [],
            "selectedRuntime": "ollama",
            "selectedRuntimeStatus": {
                "providerId": "ollama",
                "available": True,
                "reachable": True,
                "configured": True,
                "baseUrl": "http://127.0.0.1:11434",
                "defaultModel": "llama3.1:8b",
                "latencyMs": 12,
                "lastCheckedAt": "2026-06-03T12:00:00Z",
                "errorCode": "",
                "jobs": [],
            },
            "runtimes": {
                "ollama": {
                    "providerId": "ollama",
                    "available": True,
                    "reachable": True,
                    "configured": True,
                    "baseUrl": "http://127.0.0.1:11434",
                    "defaultModel": "llama3.1:8b",
                    "latencyMs": 12,
                    "lastCheckedAt": "2026-06-03T12:00:00Z",
                    "errorCode": "",
                    "jobs": [],
                }
            },
            "defaultLocalModel": "llama3.1:8b",
        },
    )

    response = runtime.handle(
        {
            "id": "req_provider_catalog",
            "taskId": "task_provider_catalog",
            "capability": "providers.catalog",
            "payload": {},
        }
    )

    assert response["ok"] is True
    catalog = response["result"]["result"]
    first_provider = next(item for item in catalog["providers"] if item["id"] == "openai")
    assert "baseUrl" not in first_provider
    assert first_provider["endpointConfigured"] is True


def test_server_brain_is_first_public_chat_candidate_when_ready(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    state = state_store._ensure_defaults(
        {
            "account": {"accessToken": "user-token"},
            "controlPlane": {
                "brainProfile": {
                    "chat": {
                        "isChatUsable": True,
                        "servingProvider": "ollama",
                        "localProviderHint": "ollama",
                        "connection": {
                            "serverBrainReady": True,
                        },
                    },
                    "bridge": {
                        "serverBrainReady": True,
                    },
                }
            },
            "providers": {
                "routingPolicy": "local_first",
                "fallbackToCloud": True,
                "ollama": {"enabled": True, "defaultModel": "llama3.1:8b"},
                "openai": {
                    "enabled": True,
                    "apiKey": "key",
                    "baseUrl": "https://api.openai.com/v1",
                    "defaultModel": "gpt-4.1-mini",
                },
            },
        }
    )

    assert bridge._chat_provider_candidates(state, privacy_class="public_text")[0] == "server_brain"

    monkeypatch.setattr(
        bridge,
        "_chat_with_ollama",
        lambda *_args, **_kwargs: {
            "ok": True,
            "content": "Merhaba",
            "provider": "ollama",
            "model": "llama3.1:8b",
        },
    )

    result = bridge._invoke_provider_chat(state, "server_brain", [{"role": "user", "text": "selam"}], "selam")

    assert result["ok"] is True
    assert result["provider"] == "server_brain"


def test_server_brain_is_first_public_chat_candidate_even_when_snapshot_is_stale(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    state = state_store._ensure_defaults(
        {
            "account": {"accessToken": "user-token"},
            "providers": {
                "routingPolicy": "local_first",
                "fallbackToCloud": True,
                "ollama": {"enabled": True, "defaultModel": "llama3.1:8b"},
                "openai": {
                    "enabled": True,
                    "apiKey": "key",
                    "baseUrl": "https://api.openai.com/v1",
                    "defaultModel": "gpt-4.1-mini",
                },
            },
        }
    )

    assert bridge._chat_provider_candidates(state, privacy_class="public_text")[0] == "server_brain"


def test_server_brain_uses_backend_chat_messages_when_no_hidden_hint(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state({"account": {"accessToken": "user-token"}})

    class FakeBackend:
        configured = True
        loopback = False

        def __init__(self) -> None:
            self.calls: list[str] = []

        def brain_profile(self) -> BackendResult:
            self.calls.append("brain_profile")
            return BackendResult(
                ok=True,
                request_id="req_brain_profile",
                status_code=200,
                data={
                    "chat": {
                        "isChatUsable": True,
                        "serverBrainName": "Elyan",
                    },
                    "bridge": {
                        "serverBrainReady": True,
                    },
                },
            )

        def brain_retrieval_search(self, _payload: dict[str, object]) -> BackendResult:
            self.calls.append("brain_retrieval_search")
            return BackendResult(
                ok=True,
                request_id="req_brain_search",
                status_code=200,
                data={"results": []},
            )

        def chat_messages(self, payload: dict[str, object]) -> BackendResult:
            self.calls.append("chat_messages")
            assert payload["source"] == "desktop"
            assert payload["content"] == "selam"
            # P1: brain'e yapılandırılmış cowork bağlamı metadata olarak gider.
            metadata = payload["metadata"]
            assert isinstance(metadata, dict)
            cowork = metadata["coworkContext"]
            assert isinstance(cowork, dict)
            assert cowork["contract"] == "elyan.cowork.v1"
            return BackendResult(
                ok=True,
                request_id="req_chat_messages",
                status_code=200,
                data={
                    "assistantMessage": {"role": "assistant", "content": "Backend yanıtı"},
                    "brain": {"serverBrainReady": True, "provider": "server_brain"},
                    "dispatched": False,
                    "reused": False,
                },
            )

    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {
            "controlPlane": {
                "health": {
                    "ok": True,
                    "agent": {"chatReady": True, "serverBrainReady": True},
                }
            }
        }
    )
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    response = runtime.handle(
        {
            "id": "req_server_brain",
            "taskId": "task_server_brain",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "selam"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["chatOk"] is True
    assert response["result"]["assistantMessage"] == "Backend yanıtı"
    assert response["result"]["provider"] == "server_brain"
    assert "chat_messages" in runtime.backend.calls  # type: ignore[attr-defined]


def test_server_brain_reconciles_local_conversation_to_canonical_backend_session(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "account": {"accessToken": "user-token"},
            "controlPlane": {
                "health": {
                    "ok": True,
                    "agent": {"chatReady": True, "serverBrainReady": True},
                }
            },
        }
    )
    now = "2026-06-21T09:00:00.000Z"
    session = {
        "id": VALID_CHAT_SESSION_ID,
        "title": "Selam",
        "source": "desktop",
        "status": "active",
        "createdAt": now,
        "updatedAt": now,
        "lastMessageAt": now,
        "messageCount": 2,
    }
    messages = [
        {
            "id": "message-user",
            "role": "user",
            "content": "selam",
            "status": "completed",
            "createdAt": now,
        },
        {
            "id": "message-assistant",
            "role": "assistant",
            "content": "Backend yanıtı",
            "blocks": [
                {
                    "type": "summary",
                    "stableBlockId": "summary-result",
                    "visibility": "user_visible",
                    "title": "Sonuç",
                    "content": "Backend yanıtı",
                }
            ],
            "status": "completed",
            "createdAt": now,
        },
    ]

    class FakeBackend:
        configured = True
        loopback = False

        def __init__(self) -> None:
            self.message_payloads: list[dict[str, object]] = []

        def chat_messages(self, payload: dict[str, object]) -> BackendResult:
            self.message_payloads.append(dict(payload))
            if len(self.message_payloads) == 1:
                assert "sessionId" not in payload
            else:
                assert payload["sessionId"] == VALID_CHAT_SESSION_ID
            assert payload["source"] == "desktop"
            transport = payload["metadata"]["desktopTransport"]  # type: ignore[index]
            assert transport["rawPrivateDataUploaded"] is False
            assert transport["derivedContextOnly"] is True
            return BackendResult(
                ok=True,
                request_id="req_chat_messages",
                status_code=200,
                data={
                    "session": session,
                    "userMessage": messages[0],
                    "assistantMessage": messages[1],
                    "delivery": {"route": "server_brain", "presentation": "chat"},
                    "brain": {"serverBrainReady": True, "provider": "server_brain"},
                    "dispatched": False,
                    "reused": False,
                },
            )

        def chat_sessions(self, **_kwargs: object) -> BackendResult:
            assert _kwargs["limit"] == 30
            return BackendResult(
                ok=True,
                request_id="req_chat_sessions",
                status_code=200,
                data={"sessions": [session]},
            )

        def chat_session_detail(self, session_id: str) -> BackendResult:
            assert session_id == VALID_CHAT_SESSION_ID
            return BackendResult(
                ok=True,
                request_id="req_chat_session_detail",
                status_code=200,
                data={"session": session, "messages": messages},
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(
        runtime,
        "_shared_brain_context_for_conversation",
        lambda **_kwargs: ("", {}, None),
    )

    first = runtime.handle(
        {
            "id": "req_server_canonical_first",
            "taskId": "task_server_canonical_first",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "selam"},
        }
    )

    assert first["ok"] is True
    assert first["result"]["conversationId"] == VALID_CHAT_SESSION_ID
    assert first["result"]["executionMode"] == "server_brain"
    state = first["result"]["state"]
    assert state["conversation"]["activeId"] == VALID_CHAT_SESSION_ID
    assert [item["id"] for item in state["conversation"]["items"]] == [VALID_CHAT_SESSION_ID]
    assistant = state["conversation"]["items"][0]["messages"][1]
    assert assistant["text"] == "Backend yanıtı"
    assert assistant["blocks"][0]["type"] == "summary"

    second = runtime.handle(
        {
            "id": "req_server_canonical_second",
            "taskId": "task_server_canonical_second",
            "capability": "conversation.send",
            "payload": {"conversationId": VALID_CHAT_SESSION_ID, "text": "devam"},
        }
    )

    assert second["ok"] is True
    assert second["result"]["conversationId"] == VALID_CHAT_SESSION_ID
    assert len(runtime.backend.message_payloads) == 2  # type: ignore[attr-defined]


def test_server_brain_normalizes_content_blob_text_blocks_without_raw_dict(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "account": {"accessToken": "user-token"},
            "controlPlane": {
                "health": {
                    "ok": True,
                    "agent": {"chatReady": True, "serverBrainReady": True},
                }
            },
        }
    )
    now = "2026-06-21T10:30:00.000Z"
    session = {
        "id": VALID_CHAT_SESSION_ID,
        "title": "Selam",
        "source": "desktop",
        "status": "active",
        "createdAt": now,
        "updatedAt": now,
        "lastMessageAt": now,
        "messageCount": 2,
    }
    messages = [
        {"id": "message-user", "role": "user", "content": "selam", "status": "completed", "createdAt": now},
        {
            "id": "message-assistant",
            "role": "assistant",
            "content": "",
            "contentBlob": {
                "blocks": [
                    {
                        "type": "text",
                        "markdown": "Merhaba, Elyan hazır.",
                        "format": "markdown",
                        "version": 1,
                    }
                ]
            },
            "status": "completed",
            "createdAt": now,
        },
    ]

    class FakeBackend:
        configured = True
        loopback = False

        def chat_messages(self, _payload: dict[str, object]) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_chat_messages",
                status_code=200,
                data={
                    "session": session,
                    "userMessage": messages[0],
                    "assistantMessage": messages[1],
                    "delivery": {"route": "server_brain", "presentation": "chat"},
                    "brain": {"serverBrainReady": True, "provider": "server_brain"},
                },
            )

        def chat_sessions(self, **_kwargs: object) -> BackendResult:
            return BackendResult(ok=True, request_id="req_chat_sessions", status_code=200, data={"sessions": [session]})

        def chat_session_detail(self, _session_id: str) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_chat_session_detail",
                status_code=200,
                data={"session": session, "messages": messages},
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_shared_brain_context_for_conversation", lambda **_kwargs: ("", {}, None))

    response = runtime.handle(
        {
            "id": "req_server_content_blob",
            "taskId": "task_server_content_blob",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "selam"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["assistantMessage"] == "Merhaba, Elyan hazır."
    assistant = response["result"]["state"]["conversation"]["items"][0]["messages"][1]
    assert assistant["text"] == "Merhaba, Elyan hazır."
    assert assistant["content"] == "Merhaba, Elyan hazır."
    assert assistant["blocks"] == [{"type": "text", "markdown": "Merhaba, Elyan hazır.", "format": "markdown", "version": 1}]
    assert "contentBlob" in assistant


def test_server_brain_waits_for_final_text_after_running_trace(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    monkeypatch.setattr(bridge.time, "sleep", lambda _seconds: None)
    state_store.update_state(
        {
            "account": {"accessToken": "user-token"},
            "controlPlane": {
                "health": {
                    "ok": True,
                    "agent": {"chatReady": True, "serverBrainReady": True},
                }
            },
        }
    )
    now = "2026-06-21T10:45:00.000Z"
    session = {
        "id": VALID_CHAT_SESSION_ID,
        "title": "Merhaba",
        "source": "desktop",
        "status": "active",
        "createdAt": now,
        "updatedAt": now,
        "lastMessageAt": now,
        "messageCount": 2,
    }
    running_assistant = {
        "id": "message-assistant-running",
        "sessionId": VALID_CHAT_SESSION_ID,
        "role": "assistant",
        "status": "queued",
        "content": "",
        "blocks": [{"type": "task_trace", "status": "running", "title": "Görev yürütülüyor"}],
    }
    final_assistant = {
        "id": "message-assistant-final",
        "sessionId": VALID_CHAT_SESSION_ID,
        "role": "assistant",
        "status": "completed",
        "content": "Merhaba, nasıl yardımcı olabilirim?",
        "blocks": [
            {
                "type": "text",
                "markdown": "Merhaba, nasıl yardımcı olabilirim?",
                "format": "markdown",
                "version": 1,
            }
        ],
    }

    class FakeBackend:
        configured = True
        loopback = False

        def __init__(self) -> None:
            self.detail_calls = 0

        def chat_messages(self, _payload: dict[str, object]) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_chat_messages",
                status_code=200,
                data={
                    "session": session,
                    "userMessage": {"id": "message-user", "role": "user", "content": "merhaba"},
                    "assistantMessage": running_assistant,
                    "delivery": {"route": "server_brain", "presentation": "chat"},
                    "brain": {"serverBrainReady": True, "provider": "server_brain"},
                },
            )

        def chat_sessions(self, **_kwargs: object) -> BackendResult:
            return BackendResult(ok=True, request_id="req_chat_sessions", status_code=200, data={"sessions": [session]})

        def chat_session_detail(self, _session_id: str) -> BackendResult:
            self.detail_calls += 1
            messages = [
                {"id": "message-user", "role": "user", "content": "merhaba"},
                final_assistant if self.detail_calls >= 2 else running_assistant,
            ]
            return BackendResult(
                ok=True,
                request_id="req_chat_session_detail",
                status_code=200,
                data={"session": session, "messages": messages},
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(runtime, "_shared_brain_context_for_conversation", lambda **_kwargs: ("", {}, None))

    response = runtime.handle(
        {
            "id": "req_server_finalized",
            "taskId": "task_server_finalized",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "merhaba"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["assistantMessage"] == "Merhaba, nasıl yardımcı olabilirim?"
    assistant = response["result"]["state"]["conversation"]["items"][0]["messages"][1]
    assert assistant["text"] == "Merhaba, nasıl yardımcı olabilirim?"
    assert assistant["blocks"][0]["type"] == "text"
    assert runtime.backend.detail_calls >= 2  # type: ignore[attr-defined]


def test_server_brain_can_live_probe_when_snapshot_is_stale(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state({"account": {"accessToken": "user-token"}})

    class FakeBackend:
        configured = True
        loopback = False

        def __init__(self) -> None:
            self.calls: list[str] = []

        def auth_me(self) -> BackendResult:
            self.calls.append("auth_me")
            return BackendResult(
                ok=True,
                request_id="req_brain_auth",
                status_code=200,
                data={"user": {"email": "user@example.com"}},
            )

        def mobile_bootstrap(self) -> BackendResult:
            self.calls.append("mobile_bootstrap")
            return BackendResult(
                ok=True,
                request_id="req_brain_bootstrap",
                status_code=200,
                data={"devices": []},
            )

        def health(self) -> BackendResult:
            self.calls.append("health")
            return BackendResult(
                ok=True,
                request_id="req_brain_health",
                status_code=200,
                data={"ok": True, "agent": {"chatReady": True, "serverBrainReady": True}, "database": {"status": "up"}},
            )

        def brain_profile(self) -> BackendResult:
            self.calls.append("brain_profile")
            return BackendResult(
                ok=True,
                request_id="req_brain_profile",
                status_code=200,
                data={
                    "chat": {
                        "isChatUsable": True,
                        "serverBrainName": "Elyan",
                        "connection": {"serverBrainReady": True},
                    },
                    "bridge": {"serverBrainReady": True},
                },
            )

        def brain_retrieval_search(self, _payload: dict[str, object]) -> BackendResult:
            self.calls.append("brain_retrieval_search")
            return BackendResult(
                ok=True,
                request_id="req_brain_search",
                status_code=200,
                data={"results": []},
            )

        def chat_messages(self, payload: dict[str, object]) -> BackendResult:
            self.calls.append("chat_messages")
            assert payload["source"] == "desktop"
            assert payload["content"] == "selam"
            # P1: brain'e yapılandırılmış cowork bağlamı metadata olarak gider.
            metadata = payload["metadata"]
            assert isinstance(metadata, dict)
            cowork = metadata["coworkContext"]
            assert isinstance(cowork, dict)
            assert cowork["contract"] == "elyan.cowork.v1"
            return BackendResult(
                ok=True,
                request_id="req_chat_messages",
                status_code=200,
                data={
                    "assistantMessage": {"role": "assistant", "content": "Backend yanıtı"},
                    "brain": {"serverBrainReady": True, "provider": "server_brain"},
                    "dispatched": False,
                    "reused": False,
                },
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    response = runtime.handle(
        {
            "id": "req_server_brain_live",
            "taskId": "task_server_brain_live",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "selam"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["chatOk"] is True
    assert response["result"]["assistantMessage"] == "Backend yanıtı"
    assert response["result"]["provider"] == "server_brain"
    assert "auth_me" in runtime.backend.calls  # type: ignore[attr-defined]
    assert "brain_profile" in runtime.backend.calls  # type: ignore[attr-defined]
    assert "chat_messages" in runtime.backend.calls  # type: ignore[attr-defined]


def test_local_private_without_local_provider_returns_safe_local_runtime_error(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.save_state(
        {
            "providers": {
                "routingPolicy": "local_first",
                "fallbackToCloud": True,
                "ollama": {"enabled": True, "defaultModel": ""},
                "openai": {
                    "enabled": True,
                    "apiKey": "key",
                    "baseUrl": "https://api.openai.com/v1",
                    "defaultModel": "gpt-4.1-mini",
                },
            }
        }
    )
    monkeypatch.setattr(bridge, "route_text_to_tool", lambda _text, **_kwargs: None)
    monkeypatch.setattr(bridge, "_semantic_route", lambda _state, _conversation, _text, **_kwargs: None)

    response = runtime.handle(
        {
            "id": "req_permission",
            "taskId": "task_permission",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "bu belgeyi yorumla"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["chatOk"] is False
    assert response["result"]["error"]["code"] == "LOCAL_MODEL_NOT_CONFIGURED"
    assert "yerel runtime" in response["result"]["assistantMessage"].lower()


def test_public_text_chat_falls_back_to_cloud_when_local_runtime_fails(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.save_state(
        {
            "providers": {
                "routingPolicy": "local_first",
                "fallbackToCloud": True,
                "ollama": {"enabled": True, "defaultModel": "llama3.2:3b"},
                "openai": {
                    "enabled": True,
                    "apiKey": "key",
                    "baseUrl": "https://api.openai.com/v1",
                    "defaultModel": "gpt-4.1-mini",
                },
            }
        }
    )
    monkeypatch.setattr(bridge, "route_text_to_tool", lambda _text, **_kwargs: None)
    monkeypatch.setattr(bridge, "_semantic_route", lambda _state, _conversation, _text, **_kwargs: None)

    def fake_invoke(_state: dict[str, object], provider: str, _conversation: list[dict[str, str]], text: str) -> dict[str, object]:
        if provider == "ollama":
            return {"ok": False, "error": "provider_unreachable"}
        return {"ok": True, "content": f"Cloud yanit: {text}", "provider": provider}

    monkeypatch.setattr(bridge, "_invoke_provider_chat", fake_invoke)

    response = runtime.handle(
        {
            "id": "req_cloud_fallback",
            "taskId": "task_cloud_fallback",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "merhaba nasilsin"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["chatOk"] is True
    assert response["result"]["provider"] == "openai"


def test_local_private_chat_fails_closed_when_selected_runtime_unreachable(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.save_state(
        {
            "providers": {
                "routingPolicy": "local_first",
                "fallbackToCloud": True,
                "ollama": {"enabled": True, "defaultModel": "llama3.2:3b"},
                "openai": {
                    "enabled": True,
                    "apiKey": "key",
                    "baseUrl": "https://api.openai.com/v1",
                    "defaultModel": "gpt-4.1-mini",
                },
            }
        }
    )
    monkeypatch.setattr(bridge, "route_text_to_tool", lambda _text, **_kwargs: None)
    monkeypatch.setattr(bridge, "_semantic_route", lambda _state, _conversation, _text, **_kwargs: None)
    monkeypatch.setattr(
        bridge,
        "_local_runtime_status_from_state",
        lambda _state, provider_id="": (
            provider_id or "ollama",
            {
                "providerId": provider_id or "ollama",
                "available": False,
                "reachable": False,
                "configured": True,
                "baseUrl": "http://127.0.0.1:11434",
                "defaultModel": "llama3.2:3b",
                "latencyMs": 0,
                "lastCheckedAt": "2026-06-03T12:00:00Z",
                "errorCode": "provider_unreachable",
                "jobs": [],
            },
        ),
    )

    response = runtime.handle(
        {
            "id": "req_local_private_unreachable",
            "taskId": "task_local_private_unreachable",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "bu belgeyi yorumla"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["chatOk"] is False
    assert response["result"]["error"]["code"] == "LOCAL_MODEL_UNREACHABLE"


def test_semantic_route_filters_workspace_context_for_cloud_provider(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured: dict[str, str] = {}
    state = state_store._ensure_defaults(
        {
            "providers": {
                "routingPolicy": "cloud_fallback",
                "active": "openai",
                "openai": {
                    "enabled": True,
                    "apiKey": "key",
                    "baseUrl": "https://api.openai.com/v1",
                    "defaultModel": "gpt-4.1-mini",
                },
                "ollama": {
                    "enabled": True,
                    "defaultModel": "",
                },
            }
        }
    )

    monkeypatch.setattr(
        bridge,
        "run_capability",
        lambda capability, _args, _state: {
            "ok": True,
            "result": {
                "kind": "retrieve_context",
                "sources": ["workspace", "conversations"],
                "strategy": "lexical",
                "matches": [
                    {"source": "workspace", "title": "secret.txt", "snippet": "workspace only snippet"},
                    {"source": "conversations", "title": "conv", "snippet": "conversation snippet"},
                ],
            },
            "artifacts": [],
            "tool": capability,
        }
        if capability == "retrieve_context"
        else {"ok": False, "error": {"code": "unexpected", "message": "unexpected"}},
    )

    def fake_invoke(_state: dict, provider: str, conversation: list[dict[str, str]], _text: str) -> dict[str, str]:
        captured["provider"] = provider
        captured["prompt"] = _text
        captured["conversation"] = conversation
        return {
            "ok": True,
            "content": '{"intent":"chat","capability":"","args":{},"confidence":0.7,"requiresConfirmation":false,"isMultiStep":false,"privacyClass":"public_text"}',
            "provider": provider,
        }

    monkeypatch.setattr(bridge, "_invoke_provider_chat", fake_invoke)

    result = bridge._semantic_route(
        state,
        [{"role": "user", "text": "önceki özel konuşma içeriği"}],
        "önceki proje konuşmasını özetle",
        conversation_id="",
    )

    assert result is not None
    assert captured["provider"] == "openai"
    assert captured["conversation"] == []
    assert "conversation snippet" in captured["prompt"]
    assert "workspace only snippet" not in captured["prompt"]
    assert "önceki özel konuşma içeriği" not in captured["prompt"]
    planning_envelope = json.loads(captured["prompt"])
    assert planning_envelope["type"] == "elyan.plan.request"
    assert planning_envelope["request"]["text"] == "önceki proje konuşmasını özetle"


def test_semantic_route_never_uses_cloud_for_local_private_goal(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    state = state_store._ensure_defaults(
        {
            "providers": {
                "routingPolicy": "cloud_fallback",
                "active": "openai",
                "openai": {
                    "enabled": True,
                    "apiKey": "key",
                    "baseUrl": "https://api.openai.com/v1",
                    "defaultModel": "gpt-4.1-mini",
                },
                "ollama": {"enabled": False, "defaultModel": ""},
            }
        }
    )
    invoked: list[str] = []
    monkeypatch.setattr(
        bridge,
        "_invoke_provider_chat",
        lambda _state, provider, _conversation, _text: invoked.append(provider),
    )

    result = bridge._semantic_route(
        state,
        [],
        "summarize this",
        goal_context={"goalContract": {"privacy": "local_private"}},
    )

    assert result is None
    assert invoked == []


def test_operator_planner_uses_local_first_then_cloud_with_sanitized_observation(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured: list[tuple[str, str]] = []
    state = state_store._ensure_defaults(
        {
            "providers": {
                "routingPolicy": "local_first",
                "fallbackToCloud": True,
                "active": "openai",
                "openai": {
                    "enabled": True,
                    "apiKey": "key",
                    "baseUrl": "https://api.openai.com/v1",
                    "defaultModel": "gpt-4.1-mini",
                },
                "ollama": {
                    "enabled": True,
                    "defaultModel": "llama3.2:3b",
                },
            }
        }
    )

    def fake_invoke(_state: dict, provider: str, conversation: list[dict[str, str]], _text: str) -> dict[str, str]:
        captured.append((provider, conversation[0]["text"]))
        if provider == "ollama":
            return {"ok": True, "content": '{"steps":[],"confidence":0.2,"clarificationQuestion":""}', "provider": provider}
        return {
            "ok": True,
            "content": '{"steps":[{"action":"click","targetText":"Continue","elementType":"button"}],"confidence":0.86,"clarificationQuestion":""}',
            "provider": provider,
        }

    monkeypatch.setattr(bridge, "_invoke_provider_chat", fake_invoke)

    result = bridge.plan_visual_operator_steps(
        "devam et",
        {
            "activeApp": "Google Chrome",
            "activeWindow": "Docs",
            "resolutionMode": "browser_first",
            "screenshotPath": "/tmp/private.png",
            "elements": [{"type": "button", "text": "Continue", "source": "browser_dom", "bbox": {"x": 1, "y": 1, "w": 10, "h": 10}}],
        },
        state=state,
    )

    assert result["steps"][0]["action"] == "click"
    assert result["provider"] == "openai"
    assert captured[0][0] == "ollama"
    assert "private.png" not in captured[-1][1]


def test_document_read_requires_explicit_or_workspace_target(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    import runtime.task_router as task_router

    monkeypatch.setattr(task_router, "_workspace_root", lambda: tmp_path)

    assert bridge.route_text_to_tool("masaüstündeki son pdfyi özetle") is None


def test_conversation_send_clarifies_when_artifact_target_missing(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_missing_artifact",
            "taskId": "task_missing_artifact",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "bu dosyayı oku"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["clarificationNeeded"] is True
    assert response["result"]["executionMode"] == "clarification"
    assert "belge seç" in response["result"]["clarificationQuestion"].lower()


def test_conversation_send_clarifies_when_data_target_missing(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_missing_data_artifact",
            "taskId": "task_missing_data_artifact",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "bu tabloyu analiz et"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["clarificationNeeded"] is True
    assert "csv/json" in response["result"]["clarificationQuestion"].lower()


def test_runtime_status_exposes_dependency_truth(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "runtime": {
                "currentTaskId": "task_123",
                "capabilities": [
                    "web_research",
                    "email_draft",
                    "math_solve",
                    "shell_run",
                ],
            }
        }
    )
    runtime = bridge.RuntimeBridge()

    status = runtime.status()

    assert status["ok"] is True
    assert "dependencyStatus" in status
    assert "sympy" in status["dependencyStatus"]
    assert "pandas" in status["dependencyStatus"]
    assert "matplotlib" in status["dependencyStatus"]
    assert "latex2sympy2_extended" in status["dependencyStatus"]
    assert "python_docx" in status["dependencyStatus"]
    assert "openpyxl" in status["dependencyStatus"]
    assert "langgraph" in status["dependencyStatus"]
    assert "litellm" in status["dependencyStatus"]
    assert "qiskit" in status["dependencyStatus"]
    assert "qiskit_aer" in status["dependencyStatus"]
    assert "dimod" in status["dependencyStatus"]
    assert "ocean_sdk" in status["dependencyStatus"]
    assert "retrievalStatus" in status
    assert status["retrievalStatus"]["model"] == "all-MiniLM-L6-v2"
    assert "taskIntelligenceStatus" in status
    assert status["taskIntelligenceStatus"]["available"] is True
    assert status["taskIntelligenceStatus"]["responseStyle"]["tone"] == "professional"
    assert "runtimeCapabilities" in status
    assert "web_research" in status["runtimeCapabilities"]
    assert status["runtimeCapabilityCount"] == 4
    assert "runtimeCapabilityMetadataSummary" in status
    assert status["runtimeCapabilityMetadataSummary"]["total"] == 4
    assert set(status["runtimeCapabilityGroups"]) == {
        "research_docs",
        "communication_approval",
        "math_quantum",
        "local_execution",
    }
    assert status["runtimeCapabilityGroups"]["research_docs"]["count"] == 1
    assert status["runtimeCapabilityGroups"]["communication_approval"]["count"] == 1
    assert status["runtimeCapabilityGroups"]["math_quantum"]["count"] == 1
    assert status["runtimeCapabilityGroups"]["local_execution"]["count"] == 1
    assert status["runtimeTransport"]["mode"] in {"websocket", "heartbeat", "unavailable"}
    assert "lastErrorCode" in status["runtimeTransport"]
    assert status["runtimeCurrentTaskId"] == "task_123"
    assert "executorStatus" in status
    assert status["executorStatus"]["available"] is True
    assert status["executorStatus"]["activeExecutionCount"] == 0
    assert status["executorStatus"]["capabilityMetadataSummary"]["total"] == 4


def test_runtime_state_patch_clears_stale_pairing_realtime_ready(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "runtime": {
                "ready": True,
                "lifecycleState": "ready",
                "websocketConnected": True,
            },
            "pairing": {"realtimeReady": True},
        }
    )
    runtime = bridge.RuntimeBridge()

    runtime._runtime_state_patch(
        lifecycle_state="reconnecting",
        ready=False,
        websocket_connected=False,
        error_code="ws_error",
    )

    state = state_store.snapshot()
    assert state["runtime"]["lifecycleState"] == "reconnecting"
    assert state["runtime"]["ready"] is False
    assert state["runtime"]["websocketConnected"] is False
    assert state["pairing"]["realtimeReady"] is False


def test_runtime_backend_snapshot_clears_stale_ready_when_backend_not_ready(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "runtime": {
                "runtimeToken": "runtime-token",
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "ready": True,
                "lifecycleState": "ready",
                "websocketConnected": True,
            },
            "pairing": {"realtimeReady": True},
        }
    )
    runtime = bridge.RuntimeBridge()

    monkeypatch.setattr(
        runtime.backend,
        "auth_me",
        lambda: bridge.BackendResult(
            ok=True,
            request_id="req_auth_me",
            status_code=200,
            data={"ok": True},
        ),
    )
    monkeypatch.setattr(
        runtime.backend,
        "mobile_bootstrap",
        lambda: bridge.BackendResult(
            ok=True,
            request_id="req_mobile_bootstrap",
            status_code=200,
            data={"recentTasks": [], "summary": {"pendingApprovals": 0, "activeTasks": 0}},
        ),
    )
    monkeypatch.setattr(
        runtime.backend,
        "runtime_session",
        lambda: bridge.BackendResult(
            ok=True,
            request_id="req_runtime_session",
            status_code=200,
            data={
                "readiness": {
                    "targetStatus": "registering",
                    "canReceiveTasks": False,
                },
                "connection": {"status": "offline"},
            },
        ),
    )

    runtime._runtime_backend_snapshot()

    state = state_store.snapshot()
    assert state["runtime"]["lifecycleState"] == "reconnecting"
    assert state["runtime"]["ready"] is False
    assert state["pairing"]["realtimeReady"] is False


def test_runtime_status_exposes_retrieval_truth_after_indexing(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    import actions.retrieve_context as retrieve_context
    import runtime.capability_registry as registry

    runtime = bridge.RuntimeBridge()
    document = tmp_path / "notes.md"
    document.write_text("Elyan retrieval status document for workspace planning.", encoding="utf-8")
    monkeypatch.setattr(retrieve_context, "_workspace_root", lambda: tmp_path)
    monkeypatch.setattr(retrieve_context, "_embedding_rank", lambda _query, _documents: None)

    result = registry.run_capability(
        "retrieve_context",
        {"query": "workspace planning", "sources": ["workspace"], "limit": 2},
        state_store.snapshot(),
    )
    status = runtime.status()

    assert result["ok"] is True
    assert status["retrievalStatus"]["cacheReady"] is True
    assert status["retrievalStatus"]["indexedWorkspaceChunks"] >= 1
    assert status["retrievalStatus"]["strategy"] == "lexical"
    assert status["retrievalStatus"]["lastIndexedAt"]


def test_runtime_payloads_and_status_include_local_file_index_capability_state(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    snapshot_path = _write_native_desktop_snapshot(tmp_path)
    monkeypatch.setenv("ELYAN_DESKTOP_NATIVE_STATE_PATH", str(snapshot_path))
    state_store.update_state(
        {
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "capabilityStates": {
                    "browser.control": {"available": True, "ready": True},
                },
            }
        }
    )
    native_index_state = {
        "available": True,
        "ready": False,
        "version": "1.0.0",
        "stats": {
            "rootCount": 1,
            "indexedFileCount": 0,
            "lastScanAt": "2026-06-03T10:00:00Z",
        },
        "errorCode": "no_approved_roots",
    }
    monkeypatch.setattr(
        bridge.native_file_indexer,
        "current_capability_state",
        lambda _state=None: native_index_state,
    )
    local_models_status = {
        "status": {
            "available": True,
            "reachable": True,
            "configured": True,
            "baseUrl": "http://127.0.0.1:11434",
            "defaultModel": "llama3.1:8b",
            "latencyMs": 42,
            "lastCheckedAt": "2026-06-03T10:00:00Z",
            "errorCode": "",
            "binary": "/usr/local/bin/ollama",
            "jobs": [],
        },
        "models": {
            "ok": True,
            "available": True,
            "models": [
                {"name": "llama3.1:8b"},
                {"name": "llama3.2:3b"},
            ],
        },
        "jobs": [],
        "selectedRuntime": "ollama",
        "selectedRuntimeStatus": {
            "providerId": "ollama",
            "available": True,
            "reachable": True,
            "configured": True,
            "baseUrl": "http://127.0.0.1:11434",
            "defaultModel": "llama3.1:8b",
            "latencyMs": 42,
            "lastCheckedAt": "2026-06-03T10:00:00Z",
            "errorCode": "",
            "jobs": [],
        },
        "runtimes": {
            "ollama": {
                "providerId": "ollama",
                "available": True,
                "reachable": True,
                "configured": True,
                "baseUrl": "http://127.0.0.1:11434",
                "defaultModel": "llama3.1:8b",
                "latencyMs": 42,
                "lastCheckedAt": "2026-06-03T10:00:00Z",
                "errorCode": "",
                "jobs": [],
            }
        },
        "defaultLocalModel": "llama3.1:8b",
    }
    monkeypatch.setattr(bridge.RuntimeBridge, "local_models_status", lambda self: local_models_status)

    runtime = bridge.RuntimeBridge()
    register_payload = runtime._runtime_register_payload()
    heartbeat_payload = runtime._runtime_heartbeat_payload("online")
    status = runtime.status()

    assert register_payload is not None
    assert register_payload["capabilityStates"]["local_files.index"] == native_index_state
    assert register_payload["capabilityStates"]["local_models.api"]["available"] is True
    assert register_payload["capabilityStates"]["local_models.api"]["ready"] is True
    assert register_payload["capabilityStates"]["desktop_os.status"]["available"] is True
    assert register_payload["capabilityStates"]["desktop_os.permissions"]["available"] is True
    assert register_payload["capabilityStates"]["desktop_os.active_window"]["available"] is True
    assert heartbeat_payload["capabilityStates"]["local_files.index"] == native_index_state
    assert heartbeat_payload["capabilityStates"]["local_models.api"]["stats"]["defaultLocalModel"] == "llama3.1:8b"
    assert heartbeat_payload["capabilityStates"]["local_models.api"]["stats"]["selectedRuntime"] == "ollama"
    assert heartbeat_payload["capabilityStates"]["local_models.api"]["stats"]["selectedRuntimeReady"] is True
    assert heartbeat_payload["capabilityStates"]["local_models.api"]["stats"]["selectedRuntimeReachable"] is True
    assert heartbeat_payload["capabilityStates"]["desktop_os.status"]["available"] is True
    assert heartbeat_payload["capabilityStates"]["desktop_os.permissions"]["available"] is True
    assert heartbeat_payload["capabilityStates"]["desktop_os.active_window"]["available"] is True
    assert status["runtimeCapabilityStates"]["browser.control"] == {
        "available": True,
        "ready": True,
    }
    assert status["runtimeCapabilityStates"]["local_files.index"] == native_index_state
    assert status["runtimeCapabilityStates"]["local_models.api"]["available"] is True
    assert status["controlPlane"]["localModels"]["defaultLocalModel"] == "llama3.1:8b"
    assert status["localModels"]["status"]["available"] is True
    assert status["localModels"]["selectedRuntimeStatus"]["reachable"] is True
    assert status["executorStatus"]["modelRouterReadiness"]["localProviders"]["ollama"]["reachable"] is True
    assert status["executorStatus"]["modelRouterReadiness"]["localProviders"]["ollama"]["latencyMs"] == 42
    assert status["retrievalStatus"]["localFileIndex"] == native_index_state


def test_runtime_payloads_publish_computed_capability_readiness(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
            }
        }
    )
    monkeypatch.setattr(
        bridge,
        "dependency_status_snapshot",
        lambda: {
            "python_pptx": {"available": False, "label": "python-pptx"},
            "sympy": {"available": True, "label": "SymPy"},
        },
    )
    monkeypatch.setattr(
        bridge.native_file_indexer,
        "current_capability_state",
        lambda _state=None: {
            "available": False,
            "ready": False,
            "errorCode": "native_sidecar_unavailable",
        },
    )
    monkeypatch.setattr(
        bridge.RuntimeBridge,
        "local_models_status",
        lambda self: {
            "status": {"available": False, "baseUrl": "", "defaultModel": "", "binary": "", "jobs": []},
            "models": {"ok": False, "available": False, "models": []},
            "jobs": [],
            "defaultLocalModel": "",
        },
    )

    runtime = bridge.RuntimeBridge()
    register_payload = runtime._runtime_register_payload()
    status = runtime.status()

    assert register_payload is not None
    assert register_payload["capabilityStates"]["presentation_write"]["available"] is False
    assert register_payload["capabilityStates"]["presentation_write"]["ready"] is False
    assert register_payload["capabilityStates"]["presentation_write"]["errorCode"] == "DEPENDENCY_UNAVAILABLE"
    assert register_payload["capabilityStates"]["presentation_write"]["missingDependencies"] == ["python_pptx"]
    assert status["runtimeCapabilityStates"]["presentation_write"]["missingDependencies"] == ["python_pptx"]
    assert status["runtimeCapabilityStates"]["math_solve"]["available"] is True
    assert status["runtimeCapabilityStates"]["math_solve"]["ready"] is True


def test_runtime_payloads_refresh_stale_dependency_unavailable_state(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "capabilityStates": {
                    "math_solve": {
                        "available": False,
                        "ready": False,
                        "errorCode": "DEPENDENCY_UNAVAILABLE",
                        "missingDependencies": ["sympy"],
                    },
                },
            }
        }
    )
    monkeypatch.setattr(
        bridge,
        "dependency_status_snapshot",
        lambda: {
            "sympy": {"available": True, "label": "SymPy"},
        },
    )

    runtime = bridge.RuntimeBridge()
    register_payload = runtime._runtime_register_payload()

    assert register_payload is not None
    assert register_payload["runtimeVersion"] == bridge._package_version()
    assert register_payload["capabilityStates"]["math_solve"]["available"] is True
    assert register_payload["capabilityStates"]["math_solve"]["ready"] is True
    assert register_payload["capabilityStates"]["math_solve"]["missingDependencies"] == []


def test_runtime_register_retry_does_not_start_without_backend_register_method(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    assert runtime._runtime_register_retry_thread is None

    state_store.update_state(
        {
            "runtime": {
                "deviceId": VALID_DEVICE_ID,
                "deviceSecret": VALID_DEVICE_SECRET,
                "lifecycleState": "offline",
                "lastErrorCode": "runtime_unauthorized",
            }
        }
    )

    runtime.backend = object()  # type: ignore[assignment]
    runtime._start_runtime_register_retry_if_needed()

    assert runtime._runtime_register_retry_thread is None


def test_mcp_server_upsert_invalid_config_returns_safe_error(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    response = runtime.handle(
        {
            "id": "req_mcp_invalid",
            "taskId": "task_mcp_invalid",
            "capability": "mcp.server.upsert",
            "payload": {"name": "broken", "command": "", "transport": "stdio"},
        }
    )

    assert response["ok"] is False
    assert response["error"]["code"] == "MCP_SERVER_INVALID"


def test_mcp_server_upsert_refresh_and_remove_roundtrip(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    server_path = tmp_path / "roundtrip_mcp_server.py"
    server_path.write_text(
        """
from mcp.server.fastmcp import FastMCP
from mcp.types import ToolAnnotations

mcp = FastMCP("Roundtrip Server")

@mcp.tool(annotations=ToolAnnotations(readOnlyHint=True))
def ping() -> dict[str, str]:
    return {"pong": "ok"}

if __name__ == "__main__":
    mcp.run(transport="stdio")
        """.strip(),
        encoding="utf-8",
    )

    runtime = bridge.RuntimeBridge()

    upsert = runtime.handle(
        {
            "id": "req_mcp_upsert",
            "taskId": "task_mcp_upsert",
            "capability": "mcp.server.upsert",
            "payload": {
                "name": "roundtrip",
                "transport": "stdio",
                "command": "python3",
                "args": [str(server_path)],
                "cwd": str(tmp_path),
                "enabled": True,
            },
        }
    )

    assert upsert["ok"] is True
    server_id = upsert["result"]["server"]["id"]
    assert upsert["result"]["mcpStatus"]["toolCount"] == 1

    tools = runtime.handle(
        {
            "id": "req_mcp_tools",
            "taskId": "task_mcp_tools",
            "capability": "mcp.tools.list",
            "payload": {"refresh": False},
        }
    )
    assert tools["ok"] is True
    assert tools["result"]["tools"][0]["name"] == "ping"

    removed = runtime.handle(
        {
            "id": "req_mcp_remove",
            "taskId": "task_mcp_remove",
            "capability": "mcp.server.remove",
            "payload": {"serverId": server_id},
        }
    )

    assert removed["ok"] is True
    assert removed["result"]["removed"] is True


def test_conversation_send_surfaces_retrieval_metadata(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    monkeypatch.setattr(bridge, "route_text_to_tool", lambda _text, **_kwargs: None)
    monkeypatch.setattr(
        bridge,
        "_semantic_route",
        lambda _state, _conversation, _text, **_kwargs: {
            "intent": "math_solve",
            "capability": "math_solve",
            "args": {"expression": "2+2", "mode": "evaluate"},
            "confidence": 0.92,
            "requiresConfirmation": False,
            "isMultiStep": False,
            "privacyClass": "public_text",
            "planPreview": None,
            "provider": "ollama",
            "retrieval": {
                "kind": "retrieve_context",
                "sources": ["workspace", "conversations"],
                "strategy": "embedding",
                "model": "all-MiniLM-L6-v2",
                "indexedAt": "2026-05-21T10:00:00Z",
                "matches": [
                    {"source": "workspace", "title": "notes.md", "snippet": "planning notes", "chunkId": "a1"},
                    {"source": "conversations", "title": "conv", "snippet": "last turn", "chunkId": "b2"},
                ],
            },
        },
    )

    response = runtime.handle(
        {
            "id": "req_retrieval_metadata",
            "taskId": "task_retrieval_metadata",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "bunu hesapla"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["retrievalUsed"] is True
    assert response["result"]["retrievalStrategy"] == "embedding"
    assert response["result"]["retrievalSources"] == ["conversations", "workspace"]
    assert response["result"]["agentStatus"]["verificationUsed"] is True
    assert response["result"]["agentStatus"]["displayStage"] == "Kontrol ediyor"


def test_conversation_send_runs_shared_retrieval_on_every_outbound_message(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state({"account": {"accessToken": "user-token"}})

    class FakeBackend:
        configured = True
        loopback = False

        def __init__(self) -> None:
            self.profile_calls = 0
            self.search_calls = 0
            self.search_queries: list[str] = []

        def brain_profile(self) -> BackendResult:
            self.profile_calls += 1
            return BackendResult(
                ok=True,
                request_id="req_brain_profile",
                status_code=200,
                data={"chat": {"activeSharedModel": "shared-gpt", "localProviderHint": ""}},
            )

        def brain_retrieval_search(self, payload: dict[str, object]) -> BackendResult:
            self.search_calls += 1
            query = str(payload["query"])
            self.search_queries.append(query)
            return BackendResult(
                ok=True,
                request_id="req_brain_search",
                status_code=200,
                data={
                    "results": [
                        {
                            "source": "shared_docs",
                            "title": "Shared notes",
                            "snippet": "desktop için ortak bilgi",
                        }
                    ]
                },
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    captured_system_messages: list[str] = []

    def fake_route(
        _state: dict[str, object],
        conversation: list[dict[str, str]],
        text: str,
        **_kwargs: object,
    ) -> dict[str, object]:
        system_messages = [item["text"] for item in conversation if item.get("role") == "system"]
        captured_system_messages.extend(system_messages)
        return {"ok": True, "content": f"Yanıt: {text}", "provider": "ollama"}

    monkeypatch.setattr(bridge, "_route_chat", fake_route)

    first = runtime.handle(
        {
            "id": "req_shared_first",
            "taskId": "task_shared_first",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "ilk mesaj"},
        }
    )
    conversation_id = first["result"]["conversationId"]
    second = runtime.handle(
        {
            "id": "req_shared_second",
            "taskId": "task_shared_second",
            "capability": "conversation.send",
            "payload": {"conversationId": conversation_id, "text": "ikinci mesaj"},
        }
    )

    assert first["ok"] is True
    assert first["result"]["sharedRetrievalUsed"] is True
    assert first["result"]["sharedRetrievalCount"] == 1
    assert first["result"]["sharedRetrievalSources"] == ["shared_docs"]
    assert first["result"]["sharedModelSnapshot"]["activeSharedModel"] == "shared-gpt"
    assert any("Relevant shared knowledge" in message for message in captured_system_messages)
    assert second["ok"] is True
    assert second["result"]["sharedRetrievalUsed"] is True
    assert runtime.backend.profile_calls == 2  # type: ignore[attr-defined]
    assert runtime.backend.search_calls == 2  # type: ignore[attr-defined]
    assert runtime.backend.search_queries == ["ilk mesaj", "ikinci mesaj"]  # type: ignore[attr-defined]


def test_conversation_send_keeps_local_chat_working_when_shared_retrieval_fails(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state({"account": {"accessToken": "user-token"}})

    class FakeBackend:
        configured = True
        loopback = False

        def brain_profile(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_brain_profile",
                status_code=200,
                data={"chat": {"activeSharedModel": "shared-gpt", "localProviderHint": ""}},
            )

        def brain_retrieval_search(self, _payload: dict[str, object]) -> BackendResult:
            return BackendResult(
                ok=False,
                request_id="req_brain_search",
                status_code=503,
                data={"error": "unavailable"},
                error="unavailable",
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]
    monkeypatch.setattr(
        bridge,
        "_route_chat",
        lambda _state, _conversation, text, **_kwargs: {
            "ok": True,
            "content": f"Yanıt: {text}",
            "provider": "ollama",
        },
    )

    response = runtime.handle(
        {
            "id": "req_shared_failure",
            "taskId": "task_shared_failure",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "merhaba"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["chatOk"] is True
    assert response["result"]["assistantMessage"] == "Yanıt: merhaba"
    assert response["result"]["sharedRetrievalUsed"] is False
    assert response["result"]["sharedModelSnapshot"]["activeSharedModel"] == "shared-gpt"


def test_conversation_send_records_retrieval_metrics(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    document = tmp_path / "notes.md"
    document.write_text("Elyan workspace retrieval metric test.", encoding="utf-8")
    import actions.retrieve_context as retrieve_context

    monkeypatch.setattr(retrieve_context, "_workspace_root", lambda: tmp_path)
    monkeypatch.setattr(retrieve_context, "_embedding_rank", lambda _query, _documents: None)
    state_store.save_state(
        {
            "providers": {
                "routingPolicy": "local_first",
                "fallbackToCloud": False,
                "ollama": {
                    "enabled": True,
                    "defaultModel": "llama3.2:3b",
                },
                "local": {
                    "defaultModel": "llama3.2:3b",
                    "runtimeFamily": "ollama",
                },
            }
        }
    )

    monkeypatch.setattr(
        bridge.RuntimeBridge,
        "_shared_brain_context_for_conversation",
        lambda self, **_kwargs: (
            "",
            {
                "sharedRetrievalUsed": False,
                "sharedRetrievalCount": 0,
                "sharedRetrievalSources": [],
                "sharedModelSnapshot": {},
            },
            {},
        ),
    )
    monkeypatch.setattr(bridge, "route_text_to_tool", lambda _text, **_kwargs: None)
    monkeypatch.setattr(bridge, "_semantic_route", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(bridge, "_contextual_route", lambda *_args, **_kwargs: None)

    def fake_invoke(
        _state: dict[str, object],
        provider: str,
        conversation: list[dict[str, str]],
        text: str,
    ) -> dict[str, object]:
        system_messages = [item["text"] for item in conversation if item.get("role") == "system"]
        assert any("Relevant local context" in message for message in system_messages)
        return {
            "ok": True,
            "content": f"Yanıt: {text}",
            "provider": provider,
        }

    monkeypatch.setattr(bridge, "_invoke_provider_chat", fake_invoke)

    runtime = bridge.RuntimeBridge()
    response = runtime.handle(
        {
            "id": "req_metrics",
            "taskId": "task_metrics",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "workspace notlarini ozetle"},
        }
    )

    status = runtime.status()
    assert response["ok"] is True
    assert status["executorStatus"]["metrics"]["retrievalHits"] >= 1
    assert status["executorStatus"]["metrics"]["retrievalFallbacks"] >= 1
    assert status["executorStatus"]["metrics"]["sharedRetrievalHits"] == 0


def test_chat_provider_candidates_respect_routing_policy() -> None:
    state = state_store._ensure_defaults(
        {
            "providers": {
                "routingPolicy": "cloud_fallback",
                "active": "openai",
                "openai": {
                    "enabled": True,
                    "apiKey": "key",
                    "baseUrl": "https://api.openai.com/v1",
                    "defaultModel": "gpt-4.1-mini",
                },
                "ollama": {
                    "enabled": True,
                    "defaultModel": "llama3.2:3b",
                },
            }
        }
    )

    candidates = bridge._chat_provider_candidates(state, privacy_class="public_text")

    assert candidates[0] == "openai"
    assert "ollama" in candidates


def test_local_first_private_semantic_candidates_do_not_include_cloud() -> None:
    state = state_store._ensure_defaults(
        {
            "providers": {
                "routingPolicy": "local_first",
                "fallbackToCloud": True,
                "openai": {
                    "enabled": True,
                    "apiKey": "key",
                    "baseUrl": "https://api.openai.com/v1",
                    "defaultModel": "gpt-4.1-mini",
                },
                "ollama": {
                    "enabled": True,
                    "defaultModel": "llama3.2:3b",
                },
            }
        }
    )

    candidates = bridge._semantic_candidate_providers(state, privacy_class="local_private")

    assert candidates == ["ollama"]


def test_cloud_fallback_private_chat_candidates_do_not_include_cloud() -> None:
    state = state_store._ensure_defaults(
        {
            "providers": {
                "routingPolicy": "cloud_fallback",
                "active": "openai",
                "openai": {
                    "enabled": True,
                    "apiKey": "key",
                    "baseUrl": "https://api.openai.com/v1",
                    "defaultModel": "gpt-4.1-mini",
                },
                "ollama": {
                    "enabled": True,
                    "defaultModel": "llama3.2:3b",
                },
            }
        }
    )

    candidates = bridge._chat_provider_candidates(state, privacy_class="local_private")

    assert candidates == ["ollama"]


def test_ollama_model_probe_requires_the_configured_tag(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        bridge,
        "_OLLAMA_TAGS_CACHE",
        {"at": float("inf"), "names": ["llama3.2:1b", "qwen2.5:latest"]},
    )

    assert bridge._ollama_model_installed({}, "llama3.2:3b") is False
    assert bridge._ollama_model_installed({}, "llama3.2:1b") is True
    assert bridge._ollama_model_installed({}, "qwen2.5") is True


def test_state_store_normalizes_legacy_device_name(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)

    state_store.save_state(
        {
            "pairing": {
                "deviceName": "Elyan Desktop",
            }
        }
    )

    snapshot = state_store.load_state()
    assert snapshot["pairing"]["deviceName"] == "Elyan"


def test_rejected_plan_records_negative_intelligence(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    conversation = state_store.create_conversation("Plan")
    plan = state_store.save_pending_plan(
        {
            "conversationId": conversation["id"],
            "query": "chrome aç",
            "intent": "open_app",
            "capability": "open_app",
            "confidence": 0.8,
            "steps": [{"capability": "open_app", "args": {"app_name": "Chrome"}}],
            "planPreview": {"summary": "Chrome açılacak.", "steps": []},
        }
    )

    response = runtime.confirm_conversation_plan(conversation["id"], plan["id"], False)

    quality = state_store.capability_quality_snapshot("open_app")
    snapshot = state_store.load_state()["taskIntelligence"]
    assert response["ok"] is True
    assert response["executionMode"] == "plan_cancelled"
    assert quality["rejections"] == 1
    assert len(snapshot["rejectedPlanPatterns"]) == 1


def test_revised_plan_records_revision_intelligence(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    conversation = state_store.create_conversation("Plan")
    plan = state_store.save_pending_plan(
        {
            "conversationId": conversation["id"],
            "query": "sunumu hazırla",
            "intent": "presentation_write",
            "capability": "presentation_write",
            "confidence": 0.81,
            "steps": [{"capability": "presentation_write", "args": {"outputPath": "deck.pptx"}}],
            "planPreview": {"summary": "Sunum oluşturulacak.", "steps": []},
        }
    )
    monkeypatch.setattr(
        bridge,
        "revise_plan_payload",
        lambda _plan, _text: {
            "capability": "presentation_write",
            "steps": [
                {
                    "capability": "presentation_write",
                    "args": {"outputPath": "deck-final.pptx"},
                }
            ],
            "planPreview": {"summary": "Sunum deck-final.pptx olarak güncellendi.", "steps": []},
        },
    )

    response = runtime.revise_conversation_plan(
        conversation["id"],
        plan["id"],
        "çıktıyı deck-final.pptx yap",
    )

    quality = state_store.capability_quality_snapshot("presentation_write")
    snapshot = state_store.load_state()["taskIntelligence"]
    assert response["ok"] is True
    assert response["executionMode"] == "plan_revised"
    assert quality["revisions"] == 1
    assert len(snapshot["corrections"]) == 1


def test_confirmed_plan_records_success_intelligence(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    conversation = state_store.create_conversation("Plan")
    plan = state_store.save_pending_plan(
        {
            "conversationId": conversation["id"],
            "query": "özeti docx yap",
            "intent": "document_write",
            "capability": "document_write",
            "confidence": 0.83,
            "steps": [{"capability": "document_write", "args": {"outputPath": "notes.docx"}}],
            "planPreview": {"summary": "notes.docx oluşturulacak.", "steps": []},
        }
    )
    monkeypatch.setattr(
        runtime,
        "_execute_plan_steps",
        lambda _steps, **_context: (True, "notes.docx hazır.", [], "", {"kind": "document_write"}, []),
    )

    response = runtime.confirm_conversation_plan(conversation["id"], plan["id"], True)

    quality = state_store.capability_quality_snapshot("document_write")
    snapshot = state_store.load_state()["taskIntelligence"]
    assert response["ok"] is True
    assert response["chatOk"] is True
    assert quality["successes"] == 1
    assert len(snapshot["confirmedPlanPatterns"]) == 1


def test_confirmed_plan_stays_available_until_execution_finishes(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Onay round-trip'i planı yürütme başlamadan silmemeli; plan yalnız
    yürütme sonucu oluştuktan sonra temizlenmeli."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    conversation = state_store.create_conversation("Plan")
    plan = state_store.save_pending_plan(
        {
            "conversationId": conversation["id"],
            "query": "Finder aç",
            "intent": "open_app",
            "capability": "open_app",
            "confidence": 0.9,
            "steps": [{"capability": "open_app", "args": {"app_name": "Finder"}}],
        }
    )
    observed_during_execution: list[bool] = []

    def fake_execute(_steps: list[dict], **_context: object) -> tuple:
        observed_during_execution.append(state_store.get_pending_plan(plan["id"]) is not None)
        return True, "Finder açıldı.", [], "", {"kind": "open_app"}, []

    monkeypatch.setattr(runtime, "_execute_plan_steps", fake_execute)

    response = runtime.confirm_conversation_plan(conversation["id"], plan["id"], True)

    assert response["chatOk"] is True
    assert observed_during_execution == [True]
    assert state_store.get_pending_plan(plan["id"]) is None


def test_confirmed_plan_rejects_duplicate_execution_while_inflight(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Plan yürütme boyunca state'te kaldığı için ikinci doğrudan confirm
    aynı yan etkili adımları paralel başlatmamalı."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    conversation = state_store.create_conversation("Plan")
    plan = state_store.save_pending_plan(
        {
            "conversationId": conversation["id"],
            "query": "Finder aç",
            "intent": "open_app",
            "capability": "open_app",
            "confidence": 0.9,
            "steps": [{"capability": "open_app", "args": {"app_name": "Finder"}}],
        }
    )
    started = threading.Event()
    release = threading.Event()
    calls: list[int] = []
    first_result: list[dict] = []

    def fake_execute(_steps: list[dict], **_context: object) -> tuple:
        calls.append(len(calls) + 1)
        if len(calls) == 1:
            started.set()
            assert release.wait(2.0)
        return True, "Finder açıldı.", [], "", {"kind": "open_app"}, []

    monkeypatch.setattr(runtime, "_execute_plan_steps", fake_execute)
    worker = threading.Thread(
        target=lambda: first_result.append(
            runtime.confirm_conversation_plan(conversation["id"], plan["id"], True)
        )
    )
    worker.start()
    assert started.wait(1.0)
    second_runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(second_runtime, "_execute_plan_steps", fake_execute)
    try:
        duplicate = second_runtime.confirm_conversation_plan(conversation["id"], plan["id"], True)
    finally:
        release.set()
        worker.join(timeout=2.0)

    assert calls == [1]
    assert duplicate["executionMode"] == "plan_execution_in_progress"
    assert first_result and first_result[0]["chatOk"] is True
    assert state_store.get_pending_plan(plan["id"]) is None


def test_interrupted_confirmed_plan_fails_closed_without_replay(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    conversation = state_store.create_conversation("Plan")
    plan = state_store.save_pending_plan(
        {
            "conversationId": conversation["id"],
            "query": "mail gönder",
            "intent": "email_send",
            "capability": "email_send",
            "executionState": "failed",
            "executionErrorCode": "PLAN_EXECUTION_INTERRUPTED",
            "steps": [{"capability": "email_send", "args": {"to": ["user@example.com"]}}],
        }
    )
    monkeypatch.setattr(
        runtime,
        "_execute_plan_steps",
        lambda _steps: (_ for _ in ()).throw(AssertionError("kesintili plan replay edilmemeli")),
    )

    result = runtime.confirm_conversation_plan(conversation["id"], plan["id"], True)

    assert result["chatOk"] is False
    assert result["executionMode"] == "plan_execution_interrupted"
    assert result["error"]["code"] == "PLAN_EXECUTION_INTERRUPTED"
    assert state_store.get_pending_plan(plan["id"])["executionState"] == "failed"


def test_bad_history_forces_semantic_plan_preview(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.record_route_outcome(
        outcome="revised",
        query="finder aç",
        intent="open_app",
        capability="open_app",
    )
    state_store.record_route_outcome(
        outcome="rejected",
        query="finder aç",
        intent="open_app",
        capability="open_app",
    )
    monkeypatch.setattr(bridge, "route_text_to_tool", lambda _text, **_kwargs: None)
    monkeypatch.setattr(
        bridge,
        "_semantic_route",
        lambda _state, _conversation, _text, **_kwargs: {
            "intent": "open_app",
            "capability": "open_app",
            "args": {"app_name": "Finder"},
            "confidence": 0.96,
            "requiresConfirmation": False,
            "isMultiStep": False,
            "privacyClass": "local_private",
            "planPreview": None,
            "provider": "ollama",
        },
    )

    response = runtime.handle(
        {
            "id": "req_history_plan",
            "taskId": "task_history_plan",
            "capability": "conversation.send",
            "payload": {"conversationId": "", "text": "finder aç"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["needsConfirmation"] is True
    assert response["result"]["executionMode"] == "plan_preview"


def test_task_intelligence_outcomes_stay_bounded(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    for index in range(24):
        state_store.record_route_outcome(
            outcome="clarified",
            query=f"belge {index}",
            intent="document_read",
            capability="document_read",
            question="Hangi belgeyi açayım?",
        )
    snapshot = state_store.load_state()["taskIntelligence"]

    assert len(snapshot["recentClarifications"]) <= 16


def test_select_conversation_preserves_active_session_when_backend_detail_fails(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state({"account": {"accessToken": "user-token"}})

    class FakeBackend:
        configured = True
        loopback = False

        def chat_session_detail(self, _session_id: str) -> BackendResult:
            return BackendResult(
                ok=False,
                request_id="req_chat_detail_failed",
                status_code=503,
                data=None,
                error="service_unavailable",
            )

        def chat_sessions(self, **_kwargs: object) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_chat_sessions",
                status_code=200,
                data={
                    "sessions": [
                        {
                            "id": VALID_CHAT_SESSION_ID,
                            "title": "Geçmiş sohbet",
                            "updatedAt": "2030-05-22T15:30:00Z",
                        }
                    ]
                },
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    response = runtime.handle(
        {
            "capability": "conversation.select",
            "payload": {"conversationId": VALID_CHAT_SESSION_ID},
        }
    )

    assert response["ok"] is True
    assert response["result"]["activeConversationId"] == VALID_CHAT_SESSION_ID
    assert response["result"]["state"]["conversation"]["activeId"] == VALID_CHAT_SESSION_ID
    assert response["result"]["warning"]["code"] == "SERVICE_UNAVAILABLE"


def test_backend_device_deactivate_removes_local_mobile_truth_when_refresh_fails(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state(
        {
            "controlPlane": {
                "mobileBootstrap": {
                    "ok": True,
                    "data": {
                        "devices": [
                            {"id": "mobile-1", "type": "mobile", "label": "iPhone"},
                            {"id": "mobile-2", "type": "mobile", "label": "iPad"},
                        ]
                    },
                }
            },
            "pairing": {
                "connectedDevices": [
                    {"id": "mobile-1", "label": "iPhone"},
                    {"id": "mobile-2", "label": "iPad"},
                ]
            },
        }
    )

    class FakeBackend:
        configured = True
        loopback = False

        def device_deactivate(self, device_id: str) -> BackendResult:
            assert device_id == "mobile-1"
            return BackendResult(ok=True, request_id="req_device_delete", status_code=200, data={"id": device_id})

        def mobile_bootstrap(self) -> BackendResult:
            return BackendResult(
                ok=False,
                request_id="req_mobile_bootstrap_failed",
                status_code=503,
                data=None,
                error="service_unavailable",
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    response = runtime.handle(
        {
            "capability": "backend.device_deactivate",
            "payload": {"deviceId": "mobile-1"},
        }
    )

    assert response["ok"] is True
    state = state_store.snapshot()
    devices = state["controlPlane"]["mobileBootstrap"]["data"]["devices"]
    assert [device["id"] for device in devices] == ["mobile-2"]
    assert [device["id"] for device in state["pairing"]["connectedDevices"]] == ["mobile-2"]


def test_append_message_stores_common_block_message_contract(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    conversation = state_store.create_conversation("Block session")

    message = state_store.append_message(conversation["id"], "assistant", "Hazır cevap")

    assert message["sessionId"] == conversation["id"]
    assert message["content"] == "Hazır cevap"
    assert message["status"] == "completed"
    assert message["createdAt"]
    assert message["blocks"] == [{"type": "text", "markdown": "Hazır cevap", "version": 1}]


def test_runtime_full_access_session_is_volatile_and_shapes_effective_state(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    granted = runtime.handle(
        {
            "capability": "runtime.access.grant_session",
            "payload": {"source": "test", "ttlSeconds": 120},
        }
    )

    assert granted["ok"] is True
    access = granted["result"]["access"]
    assert access["fullAccessSession"]["enabled"] is True
    assert access["effectivePermissions"]["allow_shell"] is True
    assert access["effectivePermissions"]["allow_computer_control"] is True
    assert state_store.snapshot().get("runtime", {}).get("access") is None

    restarted = bridge.RuntimeBridge()
    status = restarted.handle({"capability": "runtime.access.status", "payload": {}})
    assert status["result"]["access"]["fullAccessSession"]["enabled"] is False


def test_runtime_full_access_allows_shell_without_persistent_permission(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    import actions.shell as shell

    class FakeCompleted:
        stdout = "ok\n"
        stderr = ""
        returncode = 0

    monkeypatch.setattr(shell.subprocess, "run", lambda *_args, **_kwargs: FakeCompleted())
    runtime = bridge.RuntimeBridge()
    runtime.runtime_access_grant_session({"source": "test"})

    response = runtime.handle(
        {
            "capability": "shell_run",
            "payload": {"command": "npm test", "mode": "full_access"},
        }
    )

    assert response["ok"] is True
    assert response["result"]["result"]["result"]["classifiedRisk"] == "mutating"


def test_runtime_task_terminal_payload_marks_private_artifacts_share_required(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    payload, artifacts, ok = runtime._runtime_task_terminal_payload(
        {
            "chatOk": True,
            "assistantMessage": "Yerel rapor hazır.",
            "provider": "local",
            "artifacts": [{"id": "art_1", "name": "rapor.pdf", "path": str(tmp_path / "rapor.pdf")}],
            "planPreview": {"privacyClass": "local_private"},
            "executionTrace": {"verificationState": {"status": "passed", "checkedSteps": 1}},
        }
    )

    assert ok is True
    assert payload["accessMode"] == "permission_gated"
    assert payload["privacyClass"] == "local_private"
    assert payload["verification"]["status"] == "passed"
    assert payload["safeSummary"] == "Yerel rapor hazır."
    assert artifacts[0]["shareable"] is False
    assert artifacts[0]["requiresUserShare"] is True


def test_synthesize_success_summary_from_execution_trace(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    summary = runtime._synthesize_success_summary(
        {
            "chatOk": True,
            "assistantMessage": "",
            "executionTrace": {
                "steps": [
                    {"label": "Chrome açıldı", "status": "completed"},
                    {"label": "Fatura indirildi", "status": "completed"},
                    {"label": "Sonraki adım", "status": "pending"},
                ]
            },
        }
    )
    assert summary
    assert "Chrome açıldı" in summary
    assert "Fatura indirildi" in summary
    # Tamamlanmamış adım özetе girmez.
    assert "Sonraki adım" not in summary


def test_synthesize_success_summary_prefers_tool_events(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    summary = runtime._synthesize_success_summary(
        {
            "toolEvents": [
                {"tool": "open_app", "ok": True, "output": "Safari açıldı."},
                {"tool": "broken", "ok": False, "output": "hata mesajı"},
            ]
        }
    )
    assert "Safari açıldı." in summary
    # Başarısız araç çıktısı özete girmez.
    assert "hata mesajı" not in summary


def test_terminal_payload_synthesizes_summary_when_assistant_message_empty(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    payload, _artifacts, ok = runtime._runtime_task_terminal_payload(
        {
            "chatOk": True,
            "assistantMessage": "",
            "provider": "local",
            "executionTrace": {"steps": [{"label": "Klasör oluşturuldu", "status": "completed"}]},
        }
    )
    assert ok is True
    # Boş asistan metni yerine kanıttan sentezlenen içerik gitmeli (statik
    # tek satır değil) — kullanıcının #1 şikâyetinin masaüstü tarafı.
    assert payload["summary"]
    assert "Klasör oluşturuldu" in payload["summary"]
    assert payload["safeSummary"]
    assert payload["result"]["assistantMessage"]


def test_user_facing_plan_summary_filters_routing_variants(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    for phrase in (
        "Görev desktopa yönlendirildi.",
        "Görev masaüstüne yönlendirildi.",
        "gorev desktopa yonlendirildi",
        "Kullanıcı dispatch butonu ile masaüstüne yönlendirdi.",
    ):
        assert bridge._user_facing_plan_summary(phrase) == ""
    # Gerçek asistan cevabı korunur.
    assert bridge._user_facing_plan_summary("Fatura hazır.") == "Fatura hazır."


def test_execution_timeout_extended_for_multi_step_plan(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    from runtime import remote_task_runner as rtr

    long_task = {
        "payload": {
            "planPreview": {
                "steps": [
                    {"capability": "document_write"},
                    {"capability": "document_write"},
                    {"capability": "document_write"},
                    {"capability": "document_write"},
                ]
            }
        }
    }
    assert (
        runtime.remote_task_runner._execution_timeout_for(long_task)
        == rtr.REMOTE_TASK_LONG_EXECUTION_TIMEOUT_SECONDS
    )
    short_task = {
        "payload": {
            "planPreview": {
                "steps": [
                    {"capability": "open_app"},
                    {"capability": "sys_info"},
                ]
            }
        }
    }
    assert (
        runtime.remote_task_runner._execution_timeout_for(short_task)
        == rtr.REMOTE_TASK_EXECUTION_TIMEOUT_SECONDS
    )


def test_pending_terminal_stashed_when_delivery_fails_and_drained_on_recovery(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(runtime, "_send_runtime_socket_message", lambda payload: False)
    monkeypatch.setattr(runtime, "_resync_terminal_remote_task", lambda task_id: None)
    fail = bridge.BackendResult(ok=False, request_id="r", status_code=503, data={})
    monkeypatch.setattr(runtime.backend, "runtime_task_status", lambda task_id, payload: fail)

    runtime._report_runtime_task_status("task-9", {"status": "completed", "summary": "Hazır."})
    # WS+HTTP ikisi de düştü → terminal kaybolmasın diye kuyruğa alınır.
    assert "task-9" in runtime._pending_terminal_reports

    ok = bridge.BackendResult(ok=True, request_id="r", status_code=200, data={"ok": True})
    monkeypatch.setattr(runtime.backend, "runtime_task_status", lambda task_id, payload: ok)
    runtime._drain_pending_terminal_reports()
    # Transport geri gelince yeniden gönderilir ve kuyruktan düşer.
    assert "task-9" not in runtime._pending_terminal_reports


def test_reassert_pending_task_status_resends_running_until_terminal(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    sent: list[dict] = []
    monkeypatch.setattr(runtime, "_resync_terminal_remote_task", lambda task_id: None)
    monkeypatch.setattr(
        runtime,
        "_send_runtime_socket_message",
        lambda payload: (sent.append(payload) or True),
    )

    runtime._report_runtime_task_status("task-5", {"status": "running", "taskRunId": "run-5"})
    assert "task-5" in runtime._last_status_reports
    sent.clear()

    # Reconnect re-assertion → aktif görevin running durumu yeniden bildirilir.
    runtime._reassert_pending_task_status()
    assert any(p.get("taskId") == "task-5" for p in sent)

    # Terminal gelince cache düşer; bir daha running re-assert edilmez.
    runtime._report_runtime_task_status("task-5", {"status": "completed", "summary": "Bitti."})
    assert "task-5" not in runtime._last_status_reports


def test_backend_truth_sync_preserves_local_only_conversations(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Yerel-öncelikli konuşmalar backend truth-sync'te silinmemeli.

    Sidebar tek gerçek kaynak olarak backend session listesini kullandığından,
    buluta hiç gitmeyen yerel konuşmalar (id 'conv_...') her senkronda
    kayboluyordu. Bu regresyon testi kaybı önlediğimizi garanti eder.
    """
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state({"account": {"refreshToken": "rt_local", "accessToken": ""}})

    # Bir yerel-öncelikli konuşma (buluta gitmemiş) + state'e yaz.
    local_id = "conv_1700000000000_abcdef01"
    state_store.update_state(
        {
            "conversation": {
                "activeId": local_id,
                "items": [
                    {
                        "id": local_id,
                        "title": "Safari'yi aç",
                        "updatedAt": "2026-07-09T10:00:00Z",
                        "messages": [
                            {"role": "user", "text": "safariyi aç"},
                            {"role": "assistant", "text": "Safari açıldı."},
                        ],
                    }
                ],
            }
        }
    )

    class FakeBackend:
        def chat_sessions(self, *, limit: int = 30) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_sessions",
                status_code=200,
                data={
                    "sessions": [
                        {
                            "id": VALID_CHAT_SESSION_ID,
                            "title": "Bulut sohbeti",
                            "updatedAt": "2026-07-09T11:00:00Z",
                        }
                    ]
                },
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    runtime._sync_conversation_truth_from_backend()

    ids = {
        str(item.get("id", ""))
        for item in state_store.snapshot().get("conversation", {}).get("items", [])
    }
    # Hem bulut session'ı hem de yerel-öncelikli konuşma korunmalı.
    assert VALID_CHAT_SESSION_ID in ids
    assert local_id in ids


def test_backend_truth_sync_clear_all_drops_local_conversations(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """clear_all=True açıkça tüm geçmişi temizler; yerel konuşmalar da gitmeli."""
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state({"account": {"refreshToken": "rt_local", "accessToken": ""}})
    local_id = "conv_1700000000000_abcdef02"
    state_store.update_state(
        {
            "conversation": {
                "activeId": local_id,
                "items": [
                    {
                        "id": local_id,
                        "title": "Yerel",
                        "updatedAt": "2026-07-09T10:00:00Z",
                        "messages": [{"role": "user", "text": "merhaba"}],
                    }
                ],
            }
        }
    )

    class FakeBackend:
        def chat_sessions(self, *, limit: int = 30) -> BackendResult:
            return BackendResult(
                ok=True, request_id="req", status_code=200, data={"sessions": []}
            )

    runtime = bridge.RuntimeBridge()
    runtime.backend = FakeBackend()  # type: ignore[assignment]

    runtime._sync_conversation_truth_from_backend(clear_all=True)

    ids = {
        str(item.get("id", ""))
        for item in state_store.snapshot().get("conversation", {}).get("items", [])
    }
    assert local_id not in ids


def test_recoverable_replan_falls_back_web_research_to_local_context(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Web araştırma ağ hatasıyla düşerse ReAct replan yerel bağlama düşmeli ve
    kalan yazıcı adımını (outputPath dahil) korumalı."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    revised = runtime._recoverable_replan(
        {
            "failedCapability": "web_research",
            "errorCode": "NETWORK_FAILED",
            "failedArgs": {"query": "elyan nedir"},
            "remainingSteps": [
                {"capability": "document_write", "args": {"outputPath": "r.docx", "title": "R"}}
            ],
        }
    )

    assert [s["capability"] for s in revised] == ["retrieve_context", "document_write"]
    assert revised[0]["args"]["query"] == "elyan nedir"
    assert revised[1]["args"]["outputPath"] == "r.docx"  # yazıcı hedefi korunur


def test_skill_expansion_preserves_parent_dependencies_and_template_references(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(
        bridge.skill_runtime,
        "prepare_skill_run",
        lambda _skill_id, _payload, _state: {
            "steps": [
                {
                    "id": "collect",
                    "capability": "retrieve_context",
                    "args": {"query": "topic"},
                },
                {
                    "id": "solve",
                    "capability": "math_solve",
                    "forEach": "{{steps.collect.result.items}}",
                    "args": {"expression": "{{item.expression}}"},
                },
            ]
        },
    )

    expanded = runtime._expand_skill_plan_steps(
        [
            {"id": "before", "capability": "sys_info", "args": {"query": "all"}},
            {
                "id": "skill",
                "capability": "run_skill",
                "dependsOn": ["before"],
                "args": {"skillId": "test.skill", "payload": {}},
            },
            {
                "id": "after",
                "capability": "document_write",
                "dependsOn": ["skill"],
                "args": {"sections": "{{steps.skill.result.items}}"},
            },
        ]
    )

    assert [step["id"] for step in expanded] == [
        "before",
        "skill__collect",
        "skill",
        "after",
    ]
    assert expanded[1]["dependsOn"] == ["before"]
    assert expanded[2]["dependsOn"] == ["skill__collect"]
    assert expanded[2]["forEach"] == "{{steps.skill__collect.result.items}}"
    assert expanded[3]["dependsOn"] == ["skill"]


def test_recoverable_replan_ignores_non_recoverable_failures(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()

    # Ağ dışı hata → replan yok (izin hatası gibi kurtarılamaz).
    assert runtime._recoverable_replan(
        {"failedCapability": "web_research", "errorCode": "ACCESS_DENIED", "failedArgs": {"query": "x"}}
    ) == []
    # Alakasız capability → replan yok.
    assert runtime._recoverable_replan(
        {"failedCapability": "document_write", "errorCode": "NETWORK_FAILED", "failedArgs": {}}
    ) == []


def test_capability_execution_telemetry_and_reliability_surfacing(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Gerçek yürütme sonucu capability telemetrisine yazılmalı ve sık başarısız
    olan araç planlayıcı bağlamına 'reliability' kaydı olarak düşmeli."""
    _isolate_state(monkeypatch, tmp_path)
    from runtime import structured_planner

    # 3 yürütme, 2 başarısız (%67 hata) — güvenilmez eşiği aşar.
    state_store.record_capability_execution("web_research", False)
    state_store.record_capability_execution("web_research", False)
    state_store.record_capability_execution("web_research", True)

    snap = state_store.capability_quality_snapshot("web_research")
    assert snap["executions"] == 3
    assert snap["executionFailures"] == 2

    records = structured_planner.intelligence_context(state_store.snapshot())
    reliability = [r for r in records if r.get("kind") == "reliability"]
    assert reliability and reliability[0]["capability"] == "web_research"
    assert reliability[0]["failureRate"] == 0.67


def test_reliable_capability_not_flagged(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Düşük hata oranlı araç reliability kaydı üretmemeli."""
    _isolate_state(monkeypatch, tmp_path)
    from runtime import structured_planner

    for _ in range(9):
        state_store.record_capability_execution("document_write", True)
    state_store.record_capability_execution("document_write", False)  # %10 hata

    records = structured_planner.intelligence_context(state_store.snapshot())
    assert not [r for r in records if r.get("kind") == "reliability"]


def test_execute_step_with_telemetry_records_outcome(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Bridge execute_step sarmalayıcısı adım sonucunu telemetriye yazmalı."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(
        bridge,
        "_execute_capability_with_preprocessing",
        lambda capability, args, state, source: ({"ok": True, "output": "ok"}, []),
    )

    result, _events = runtime._execute_step_with_telemetry("math_solve", {}, {}, "confirmed_plan")

    assert result["ok"] is True
    assert state_store.capability_quality_snapshot("math_solve")["executions"] == 1


def test_clipboard_capabilities_registered_and_run(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """clipboard_read/write kayıtlı, güvenli (read_only) ve uçtan uca çalışıyor."""
    _isolate_state(monkeypatch, tmp_path)
    from runtime import capability_registry as cr

    assert "clipboard_read" in cr.capability_names()
    assert "clipboard_write" in cr.capability_names()
    meta = cr.capability_metadata("clipboard_read")
    assert meta["permissionClass"] == "read_only"
    assert meta["sideEffect"] is False

    # pbcopy/pbpaste'i taklit et — CI/sandbox'ta gerçek pano olmayabilir.
    from actions import clipboard
    store = {"value": ""}
    monkeypatch.setattr(
        clipboard.subprocess, "run",
        lambda cmd, **kw: _FakeClip(cmd, kw, store),
    )

    state = state_store._ensure_defaults({})
    w = cr.run_capability("clipboard_write", {"text": "elyan"}, state)
    assert w["ok"] is True
    r = cr.run_capability("clipboard_read", {}, state)
    assert r["ok"] is True
    assert "elyan" in (r.get("output") or "")


class _FakeClip:
    """pbcopy/pbpaste taklidi — paylaşılan sözlükte metin tutar."""
    def __init__(self, cmd, kw, store):
        self.returncode = 0
        if cmd == ["pbcopy"]:
            store["value"] = kw.get("input", "")
            self.stdout = ""
        else:  # pbpaste
            self.stdout = store["value"]


def test_billing_truth_refresh_is_throttled_and_syncs_quota(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Dispatch/sohbet sonrası kota senkronu: auth_me çağrılır ama throttle
    aralığında ikinci kez çağrılmaz (task loop'unu yavaşlatmaz)."""
    _isolate_state(monkeypatch, tmp_path)
    state_store.update_state({"account": {"accessToken": "user-token"}})
    runtime = bridge.RuntimeBridge()

    calls = {"n": 0}

    class FakeBackend:
        def auth_me(self):
            calls["n"] += 1
            return bridge.BackendResult(ok=True, request_id="r", status_code=200, data={})

    runtime.backend = FakeBackend()  # type: ignore[assignment]

    # Thread'i senkron çalıştır ki test deterministik olsun.
    class _SyncThread:
        def __init__(self, target=None, **_kw):
            self._target = target
        def start(self):
            if self._target:
                self._target()
    monkeypatch.setattr(bridge.threading, "Thread", _SyncThread)

    runtime._refresh_billing_truth_async()
    assert calls["n"] == 1
    # Throttle aralığında ikinci çağrı no-op.
    runtime._refresh_billing_truth_async()
    assert calls["n"] == 1
    # force=True aralığı bypass eder.
    runtime._refresh_billing_truth_async(force=True)
    assert calls["n"] == 2


def test_billing_refresh_skipped_when_not_authenticated(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    calls = {"n": 0}

    class FakeBackend:
        def auth_me(self):
            calls["n"] += 1
            return bridge.BackendResult(ok=True, request_id="r", status_code=200, data={})

    runtime.backend = FakeBackend()  # type: ignore[assignment]
    runtime._refresh_billing_truth_async(force=True)  # oturum yok → çağrılmamalı
    assert calls["n"] == 0


def _dispatch_work_order(required_capabilities: list[str]) -> dict:
    return {
        "schema": "elyan.desktop_work_order.v1",
        "source": "mobile_chat_dispatch",
        "goal": {
            "kind": "browser_task",
            "summary": "Chrome kapatılacak",
            "language": "tr",
            "sourceTextHash": "b" * 24,
        },
        "entities": [],
        "constraints": [],
        "requiredCapabilities": required_capabilities,
        "localContextNeeded": [],
        "expectedOutputs": [{"kind": "chat_result", "format": "elyan_blocks.v2", "required": True}],
        "verificationRules": [
            {"id": "runtime_completed", "description": "Runtime tamamlandı.", "evidence": "runtime_status"},
        ],
        "execution": {"mode": "cowork_dispatch", "approvalPolicy": "capability_policy", "maxSteps": 8},
        "planPreview": {
            "summary": "Chrome kapatılacak",
            "privacyClass": "local_private",
            "steps": [
                {
                    "id": "step_1",
                    "capability": required_capabilities[0],
                    "description": "Adım yürütülecek.",
                    "args": {},
                }
            ],
        },
    }


def _dispatch_auto_approve_backend(task_capabilities: list[str]):
    class FakeBackend:
        configured = True
        loopback = False

        def __init__(self) -> None:
            self.status_updates: list[tuple[str, dict]] = []
            self.heartbeats: list[dict] = []

        def runtime_tasks_assigned(self) -> BackendResult:
            return BackendResult(
                ok=True,
                request_id="req_tasks",
                status_code=200,
                data={
                    "tasks": [
                        {
                            "id": "task-dispatch",
                            "title": "Chrome u kapat",
                            "status": "queued",
                            "userId": "user-tests",
                            "targetDeviceId": VALID_DEVICE_ID,
                            "payload": {
                                "prompt": "Chrome u kapat",
                                "desktopWorkOrder": _dispatch_work_order(task_capabilities),
                            },
                        }
                    ]
                },
            )

        def runtime_task_status(self, task_id: str, payload: dict) -> BackendResult:
            self.status_updates.append((task_id, payload))
            return BackendResult(ok=True, request_id="req_status", status_code=200, data={"ok": True})

        def runtime_task_artifacts(self, task_id: str, payload: dict) -> BackendResult:
            return BackendResult(ok=True, request_id="req_artifacts", status_code=200, data={"ok": True})

        def heartbeat(self, payload: dict) -> BackendResult:
            self.heartbeats.append(payload)
            return BackendResult(ok=True, request_id="req_heartbeat", status_code=200, data={"ok": True})

    return FakeBackend()


def test_dispatched_plan_within_work_order_capabilities_waits_for_backend_policy(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Mobil dispatch yalnız görev rotasını yetkilendirir. Kullanıcı eylemi
    masaüstünde otomatik onaylanmaz; güncel kullanıcı modu backend geçidinde
    değerlendirilir."""
    _isolate_state(monkeypatch, tmp_path)
    _arm_device_identity()
    runtime = bridge.RuntimeBridge()
    runtime.backend = _dispatch_auto_approve_backend(["close_app"])  # type: ignore[assignment]

    monkeypatch.setattr(runtime, "_execute_deterministic_remote_task", lambda *_a, **_k: None)
    monkeypatch.setattr(
        runtime,
        "send_conversation",
        lambda *_args, **_kwargs: {
            "ok": True,
            "chatOk": True,
            "assistantMessage": "Chrome kapatılacak.",
            "provider": "local_planner",
            "toolEvents": [],
            "conversationId": "conv_dispatch",
            "needsConfirmation": True,
            "pendingPlanId": "plan_dispatch",
            "planPreview": {
                "summary": "Chrome kapatılacak.",
                "steps": [{"capability": "close_app", "description": "Chrome kapatılacak."}],
            },
        },
    )
    confirmed: list[tuple[str, str, bool]] = []

    def fake_confirm(conversation_id: str, plan_id: str, approved: bool) -> dict:
        confirmed.append((conversation_id, plan_id, approved))
        return {
            "ok": True,
            "chatOk": True,
            "assistantMessage": "Chrome kapatıldı.",
            "provider": "local_planner",
            "toolEvents": [{"tool": "close_app", "ok": True}],
            "conversationId": conversation_id,
        }

    monkeypatch.setattr(runtime, "confirm_conversation_plan", fake_confirm)
    monkeypatch.setattr(runtime, "_pending_plan_permission_error", lambda _plan_id: None)

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"][0]["status"] == "waiting_approval"
    assert confirmed == []
    statuses = [payload["status"] for _, payload in runtime.backend.status_updates]  # type: ignore[attr-defined]
    assert statuses[-1] == "waiting_approval"
    approval_request = runtime.backend.status_updates[-1][1]["approvalRequest"]  # type: ignore[attr-defined]
    assert approval_request["permission"] == "side_effect"
    assert approval_request["idempotency"] == "non_idempotent"


def test_dispatched_plan_with_blocklisted_capability_still_waits_for_approval(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Yıkıcı/dışa dönük adımlar (ör. email_send) dispatch kapsamında bile
    açık onay ister."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    _arm_device_identity()
    runtime.backend = _dispatch_auto_approve_backend(["email_send"])  # type: ignore[assignment]

    monkeypatch.setattr(runtime, "_execute_deterministic_remote_task", lambda *_a, **_k: None)
    monkeypatch.setattr(
        runtime,
        "send_conversation",
        lambda *_args, **_kwargs: {
            "ok": True,
            "chatOk": True,
            "assistantMessage": "E-posta gönderilecek.",
            "provider": "local_planner",
            "toolEvents": [],
            "conversationId": "conv_email",
            "needsConfirmation": True,
            "pendingPlanId": "plan_email",
            "planPreview": {
                "summary": "E-posta gönderilecek.",
                "steps": [{"capability": "email_send", "description": "E-posta gönder."}],
            },
        },
    )
    monkeypatch.setattr(
        runtime,
        "confirm_conversation_plan",
        lambda *_a, **_k: (_ for _ in ()).throw(AssertionError("blocklist adımı otomatik onaylanmamalı")),
    )

    result = runtime.execute_assigned_runtime_tasks()

    assert result["ok"] is True
    assert result["executions"][0]["status"] == "waiting_approval"
    statuses = [payload["status"] for _, payload in runtime.backend.status_updates]  # type: ignore[attr-defined]
    assert statuses[-1] == "waiting_approval"


def test_sanitize_contradictory_steps_drops_browser_relaunch_after_close() -> None:
    """"Chrome'u kapat" planındaki sonraki browser/aç adımları atılır —
    kapatılan uygulama aynı görevde geri açılmaz."""
    steps = [
        {"capability": "close_app", "args": {"app_name": "Chrome u"}, "description": "kapat"},
        {"capability": "browser_control", "args": {"action": "search", "query": "x"}, "description": "ara"},
        {"capability": "open_app", "args": {"app_name": "Chrome"}, "description": "aç"},
    ]
    sanitized = bridge._sanitize_contradictory_plan_steps(steps)
    assert [s["capability"] for s in sanitized] == ["close_app"]


def test_sanitize_contradictory_steps_keeps_unrelated_steps() -> None:
    steps = [
        {"capability": "close_app", "args": {"app_name": "Spotify"}, "description": "kapat"},
        {"capability": "browser_control", "args": {"action": "search", "query": "hava"}, "description": "ara"},
        {"capability": "open_app", "args": {"app_name": "Notes"}, "description": "aç"},
    ]
    sanitized = bridge._sanitize_contradictory_plan_steps(steps)
    assert [s["capability"] for s in sanitized] == ["close_app", "browser_control", "open_app"]


def test_delegates_multicap_cowork_task_to_llm_planner(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Çok-yetenekli serbest-metin cowork görevi, server_brain hazırsa backend
    regex planı yerine kataloglu LLM planlayıcıya (None → send_conversation)
    yönlendirilir."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {"controlPlane": {"health": {"ok": True, "agent": {"chatReady": True, "serverBrainReady": True}}}}
    )
    # document_read + web_research → basit doğrudan komut değil → LLM'e devret.
    assert runtime._remote_task_should_delegate_to_llm({"document_read", "web_research"}) is True


def test_simple_close_app_stays_on_fast_regex_path(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Yalnız uygulama kapatma gibi basit doğrudan komut, server_brain hazır
    olsa bile hız için regex yolunda kalır (LLM'e devredilmez)."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {"controlPlane": {"health": {"ok": True, "agent": {"chatReady": True, "serverBrainReady": True}}}}
    )
    assert runtime._remote_task_should_delegate_to_llm({"close_app"}) is False
    # Zayıf " ve " tek başına çok-adım saymaz — hız yolu korunur.
    assert (
        runtime._remote_task_should_delegate_to_llm({"play_media"}, "müzik aç ve keyfini çıkar")
        is False
    )


def test_sequential_simple_command_delegates_to_llm_planner(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Basit doğrudan yetenek olsa bile prompt açıkça sıralı çok-adım
    bildiriyorsa (regex tek eylemi yakalar) kataloglu LLM planlayıcıya gider."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    state_store.update_state(
        {"controlPlane": {"health": {"ok": True, "agent": {"chatReady": True, "serverBrainReady": True}}}}
    )
    assert (
        runtime._remote_task_should_delegate_to_llm(
            {"open_app"}, "önce Chrome'u aç sonra Not Defteri'ni aç"
        )
        is True
    )
    assert runtime._prompt_has_sequential_intent("Safari'yi aç, ardından kapat") is True
    assert runtime._prompt_has_sequential_intent("Safari'yi aç") is False


def test_multicap_task_stays_on_regex_when_server_brain_unavailable(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """server_brain erişilemezse LLM'e devretme — regex planı çevrimdışı yedek
    olarak korunur."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    assert runtime._remote_task_should_delegate_to_llm({"document_read", "web_research"}) is False


def test_bridge_binds_live_progress_emitter(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Başsız daemon dahil her modda executor'a canlı ilerleme emitter'ı bağlanır."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    assert runtime.executor_core._progress_emitter is not None


def test_live_progress_routes_to_active_task_and_throttles(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Aktif mobil görev bağlamındayken adım geçişi backend'e canlı 'running'
    güncellemesi akıtır; aynı durumda kısa aralıkta tekrar akmaz (throttle)."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    pushed: list[tuple[str, str, list]] = []
    monkeypatch.setattr(
        runtime,
        "_report_runtime_task_status",
        lambda task_id, payload: pushed.append(
            (task_id, payload["status"], [(s["id"], s["status"]) for s in payload["executionTrace"]["steps"]])
        ),
    )
    # Aktif görev yokken hiçbir şey akmaz.
    runtime._emit_remote_task_progress("conv", {"steps": [{"id": "s1", "status": "running", "label": "x"}]})
    assert pushed == []

    token = runtime._begin_active_remote_task("task-1", "run-1")
    try:
        block = {
            "status": "running",
            "title": "Görev",
            "activeStepId": "s1",
            "steps": [
                {"id": "s1", "status": "running", "label": "Kapatılıyor", "capability": "close_app"},
                {"id": "s2", "status": "pending", "label": "Doğrula", "capability": "sys_info"},
            ],
        }
        runtime._emit_remote_task_progress("conv", block)
        runtime._emit_remote_task_progress("conv", block)  # aynı sinyal → throttle
        block2 = {
            **block,
            "activeStepId": "s2",
            "steps": [
                {"id": "s1", "status": "completed", "label": "Kapatıldı", "capability": "close_app"},
                {"id": "s2", "status": "running", "label": "Doğrula", "capability": "sys_info"},
            ],
        }
        runtime._emit_remote_task_progress("conv", block2)  # durum değişti → akar
    finally:
        runtime._end_active_remote_task(token, "task-1")

    assert [status for _, status, _ in pushed] == ["running", "running"]
    assert pushed[0][2] == [("s1", "running"), ("s2", "pending")]
    assert pushed[1][2] == [("s1", "completed"), ("s2", "running")]
    # Görev bittiğinde throttle durumu temizlenir.
    assert "task-1" not in runtime._remote_progress_last_signature


def test_ws_cancel_reaches_copied_task_scope_and_discards_late_worker_success(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from runtime import mcp_runtime, remote_task_runner as rtr

    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    entered = threading.Event()
    release = threading.Event()
    observed_task_ids: list[str] = []

    def deterministic(*_args: object) -> dict[str, object]:
        observed_task_ids.append(mcp_runtime.current_task_id())
        entered.set()
        assert release.wait(2.0)
        return {"ok": True, "output": "must not escape"}

    monkeypatch.setattr(runtime, "_execute_deterministic_remote_task", deterministic)

    def cancel_from_socket() -> None:
        assert entered.wait(1.0)
        runtime._handle_runtime_ws_message(
            json.dumps({"type": "task.cancel", "taskId": "task-scoped-cancel"})
        )
        release.set()

    token = runtime._begin_active_remote_task("task-scoped-cancel", "run-1")
    canceller = threading.Thread(target=cancel_from_socket, daemon=True)
    try:
        canceller.start()
        result = runtime.remote_task_runner._execute_local_with_timeout(
            {},
            "prompt",
            "title",
            task_id="task-scoped-cancel",
        )
        canceller.join(2.0)

        assert result is rtr._EXECUTION_CANCELLED
        assert observed_task_ids == ["task-scoped-cancel"]
        assert runtime._active_remote_task_cancellation_reason("task-scoped-cancel") == "task_cancelled"
        inbox_item = state_store.get_task_inbox_item("task-scoped-cancel")
        assert inbox_item is not None
        assert inbox_item["status"] == "canceled"
    finally:
        release.set()
        runtime._end_active_remote_task(token, "task-scoped-cancel")

    # Bounded tombstone keeps a duplicate/retried dispatch canceled too.
    assert runtime._active_remote_task_cancellation_reason("task-scoped-cancel") == "task_cancelled"


def test_ws_cancel_before_scope_prevents_remote_task_execution(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    executed: list[str] = []
    monkeypatch.setattr(
        runtime,
        "_execute_deterministic_remote_task",
        lambda *_args: executed.append("executed") or {"ok": True},
    )

    runtime._handle_runtime_ws_message(
        json.dumps({"type": "task.cancel", "taskId": "task-cancel-before-scope"})
    )
    result = runtime.remote_task_runner.execute_runtime_task(
        {"id": "task-cancel-before-scope", "payload": {"prompt": "çalışmamalı"}}
    )

    assert result["status"] == "canceled"
    assert result["error"]["code"] == "TASK_CANCELLED"
    assert executed == []
    inbox_item = state_store.get_task_inbox_item("task-cancel-before-scope")
    assert inbox_item is not None
    assert inbox_item["status"] == "canceled"


def test_cancel_first_terminal_fence_suppresses_artifacts_and_completion(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    outbound: list[str] = []
    monkeypatch.setattr(
        runtime,
        "_report_runtime_task_artifacts",
        lambda *_args, **_kwargs: outbound.append("artifacts"),
    )
    monkeypatch.setattr(
        runtime,
        "_report_runtime_task_status",
        lambda *_args, **_kwargs: outbound.append("status"),
    )

    runtime._handle_runtime_ws_message(
        json.dumps({"type": "task.cancel", "taskId": "task-terminal-race"})
    )
    result = runtime._report_runtime_task_terminal_result(
        "task-terminal-race",
        {
            "chatOk": True,
            "assistantMessage": "geç başarı",
            "artifacts": [{"kind": "file", "path": "/tmp/late.txt"}],
        },
        dispatched_via_websocket=True,
    )

    assert result["status"] == "canceled"
    assert result["error"]["code"] == "TASK_CANCELLED"
    assert outbound == []


def test_scope_limit_after_running_reports_safe_terminal_failure(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    status_payloads: list[dict[str, object]] = []
    monkeypatch.setattr(runtime, "_runtime_task_preflight_error", lambda *_args: None)
    monkeypatch.setattr(runtime, "_remote_task_running_plan_preview", lambda *_args: {})
    monkeypatch.setattr(runtime, "_set_runtime_task_heartbeat", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(
        runtime,
        "_report_runtime_task_status",
        lambda _task_id, payload: status_payloads.append(payload)
        or BackendResult(ok=True, request_id="req", status_code=200, data={"ok": True}),
    )
    monkeypatch.setattr(
        runtime,
        "_begin_active_remote_task",
        lambda *_args: (_ for _ in ()).throw(
            SafeCapabilityError("TASK_SCOPE_LIMIT", "Çok fazla görev aynı anda çalışıyor.")
        ),
    )

    _arm_device_identity()
    result = runtime.remote_task_runner.execute_runtime_task(
        _trusted_task(
            {"id": "task-scope-limit", "payload": {"prompt": "kapsam sınırı"}},
            capabilities=["retrieve_context"],
        )
    )

    assert result["status"] == "failed"
    assert status_payloads[-1]["status"] == "failed"
    assert status_payloads[-1]["error"] == "TASK_SCOPE_LIMIT"


def test_websocket_approval_resume_is_idempotent_and_terminal_fenced(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Aynı approval iki kez gelirse ikinci resume paralel başlamamalı; yerel
    terminal görev de yeniden running durumuna diriltilmemeli."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    calls: list[str] = []

    def fake_resume(task_id: str, approved: bool, answer: str = "") -> dict:
        calls.append(task_id)
        nested = None
        if len(calls) == 1:
            nested = runtime._resume_remote_task_after_approval(task_id, approved, answer)
        return {"taskId": task_id, "ok": True, "status": "resumed", "nested": nested}

    monkeypatch.setattr(runtime.remote_task_runner, "resume_after_approval", fake_resume)

    first = runtime._resume_remote_task_after_approval("task-approval", True)

    assert calls == ["task-approval"]
    assert first["nested"]["status"] == "skipped_duplicate"

    runtime._remember_terminal_assigned_task("task-approval")
    terminal_retry = runtime._resume_remote_task_after_approval("task-approval", True)

    assert terminal_retry["status"] == "skipped_recent_terminal"
    assert calls == ["task-approval"]

    state_store.upsert_task_inbox_item({"id": "task-failed", "status": "failed"})
    persisted_terminal_retry = runtime._resume_remote_task_after_approval("task-failed", True)

    assert persisted_terminal_retry["status"] == "skipped_local_terminal"
    assert calls == ["task-approval"]


def test_force_runtime_reconnect_requires_auth(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """Kimlik hazır değilken zorla yeniden bağlanma güvenle es geçer."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(runtime, "_runtime_auth_ready", lambda: False)

    result = runtime.force_runtime_reconnect()

    assert result == {"ok": False, "reason": "auth_not_ready"}


def test_force_runtime_reconnect_closes_socket_and_restarts_when_thread_dead(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """Ölü soketi kapatır, thread çıkmışsa yeniden başlatmayı tetikler."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    monkeypatch.setattr(runtime, "_runtime_auth_ready", lambda: True)

    closed: list[str] = []

    class _FakeApp:
        def close(self) -> None:
            closed.append("closed")

    runtime._runtime_ws_app = _FakeApp()
    runtime._runtime_ws_thread = None  # thread çıkmış (5 dk boşluğunun sebebi)
    runtime._runtime_ws_connected = True
    started: list[str] = []
    monkeypatch.setattr(
        runtime,
        "_start_runtime_websocket_if_needed",
        lambda: started.append("start") or True,
    )

    result = runtime.force_runtime_reconnect()

    assert result["ok"] is True
    assert result["restarted"] is True
    assert closed == ["closed"]
    assert started == ["start"]
    assert runtime._runtime_ws_connected is False


def test_execute_local_with_timeout_returns_sentinel_on_hang(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Takılan yürütme zaman aşımı sentinel'ı döner (süresiz beklemez)."""
    from runtime import remote_task_runner as rtr

    monkeypatch.setattr(rtr, "REMOTE_TASK_EXECUTION_TIMEOUT_SECONDS", 0.2)

    class _Host:
        def _execute_deterministic_remote_task(self, *_args: object) -> object:
            time.sleep(5)  # takıldı
            return {"ok": True}

        def send_conversation(self, *_args: object) -> object:  # pragma: no cover
            return {"ok": True}

        def _runtime_diag(self, *_args: object, **_kwargs: object) -> None:
            pass

    runner = rtr.RemoteTaskRunner(_Host())
    result = runner._execute_local_with_timeout({}, "prompt", "title", task_id="task-xyz")

    assert result is rtr._EXECUTION_TIMEOUT


def test_execute_local_timeout_cancels_active_mcp_before_returning(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from runtime import remote_task_runner as rtr

    monkeypatch.setattr(rtr, "REMOTE_TASK_EXECUTION_TIMEOUT_SECONDS", 0.1)
    released = threading.Event()
    cancellation_calls: list[tuple[str, str]] = []

    class _Host:
        def _execute_deterministic_remote_task(self, *_args: object) -> object:
            assert released.wait(2.0)
            return {"ok": True, "output": "late"}

        def send_conversation(self, *_args: object) -> object:  # pragma: no cover
            return {"ok": True}

        def _runtime_diag(self, *_args: object, **_kwargs: object) -> None:
            pass

        def _cancel_active_remote_task(self, task_id: str, *, reason: str) -> int:
            cancellation_calls.append((task_id, reason))
            released.set()
            return 1

        def _active_remote_task_cancellation_reason(self, _task_id: str) -> str:
            return "task_execution_timeout" if cancellation_calls else ""

    runner = rtr.RemoteTaskRunner(_Host())
    result = runner._execute_local_with_timeout({}, "prompt", "title", task_id="task-mcp-timeout")

    assert result is rtr._EXECUTION_CANCELLED
    assert cancellation_calls == [("task-mcp-timeout", "task_execution_timeout")]


def test_execute_local_with_timeout_delegates_none_to_conversation(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Deterministik yol None dönerse (LLM'e delege) send_conversation sonucu döner."""
    from runtime import remote_task_runner as rtr

    class _Host:
        def _execute_deterministic_remote_task(self, *_args: object) -> object:
            return None

        def send_conversation(self, *_args: object, **_kwargs: object) -> object:
            return {"ok": True, "via": "conversation"}

        def _runtime_diag(self, *_args: object, **_kwargs: object) -> None:
            pass

    runner = rtr.RemoteTaskRunner(_Host())
    result = runner._execute_local_with_timeout({}, "prompt", "title", task_id="task-xyz")

    assert result == {"ok": True, "via": "conversation"}


class _ClarifyHost:
    """Netleştirme akışı testleri için minimal sahte host."""

    def __init__(self) -> None:
        self.canceled: list[str] = []
        self.status_reports: list[dict] = []

    def _report_runtime_task_status(self, task_id, payload):
        self.status_reports.append({"taskId": task_id, "payload": payload})
        from types import SimpleNamespace
        return SimpleNamespace(ok=True, to_dict=lambda: {})

    def _set_runtime_task_heartbeat(self, *_args, **_kwargs):
        return None

    def _cancel_remote_pending_task(self, task_id):
        self.canceled.append(task_id)

    def _remote_task_trace_payload(self, *_args, **_kwargs):
        return {"type": "task_trace", "steps": []}


def _new_clarify_runner():
    from runtime import remote_task_runner as rtr
    return rtr.RemoteTaskRunner(_ClarifyHost()), rtr


def test_is_clarification_result_detection() -> None:
    from runtime import remote_task_runner as rtr
    R = rtr.RemoteTaskRunner
    assert R._is_clarification_result({"clarificationNeeded": True, "clarificationQuestion": "Hangi dosya?"}) is True
    assert R._is_clarification_result({"executionMode": "clarification", "clarificationQuestion": "?"}) is True
    assert R._is_clarification_result({"clarificationNeeded": True, "clarificationQuestion": ""}) is False
    assert R._is_clarification_result({"capability": "open_app"}) is False
    assert R._is_clarification_result("düz metin") is False


def test_approval_resolution_notes_extraction() -> None:
    from runtime import remote_task_runner as rtr
    R = rtr.RemoteTaskRunner
    assert R._approval_resolution_notes({"resolution": {"notes": "rapor.pdf"}}) == "rapor.pdf"
    assert R._approval_resolution_notes({"resolution": {"answer": "evet"}}) == "evet"
    assert R._approval_resolution_notes({"resolution": {}}) == ""
    assert R._approval_resolution_notes({}) == ""


def test_pause_for_clarification_persists_context(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runner, _rtr = _new_clarify_runner()
    local_result = {
        "clarificationNeeded": True,
        "clarificationQuestion": "Hangi dosyayı özetleyeyim?",
        "conversationId": "c1",
    }
    res = runner._pause_for_clarification(
        "t1", "run1", "Başlık", "dosyayı özetle", local_result, False,
        task={"id": "t1", "payload": {}}, work_order=None,
    )
    assert res["status"] == "waiting_approval"
    assert res["clarification"] is True
    link = state_store.get_remote_task_link("t1")
    assert link["clarificationPending"] is True
    assert link["clarificationRounds"] == 1
    assert link["clarificationPrompt"] == "dosyayı özetle"
    assert link["clarificationQuestion"] == "Hangi dosyayı özetleyeyim?"
    assert link["clarificationTask"] == {"id": "t1", "payload": {}}


def test_pause_for_clarification_loop_guard(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runner, rtr = _new_clarify_runner()
    state_store.save_remote_task_link("t1", "", "", task_run_id="run1", status="waiting_approval")
    state_store.update_remote_task_link("t1", {"clarificationRounds": rtr.MAX_CLARIFICATION_ROUNDS})
    captured: dict = {}
    monkeypatch.setattr(runner, "_fail_safe", lambda *a, **k: captured.update(k) or {"ok": False, "status": "failed"})

    res = runner._pause_for_clarification(
        "t1", "run1", "B", "p",
        {"clarificationNeeded": True, "clarificationQuestion": "q"}, False, task={"id": "t1"},
    )
    assert res["status"] == "failed"
    assert captured["error_code"] == "clarification_unresolved"


def test_resume_after_clarification_merges_answer_and_replans(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runner, _rtr = _new_clarify_runner()
    state_store.save_remote_task_link("t1", "", "", task_run_id="run1", status="waiting_approval")
    state_store.update_remote_task_link("t1", {
        "clarificationPending": True,
        "clarificationPrompt": "dosyayı özetle",
        "clarificationQuestion": "Hangi dosya?",
        "clarificationTask": {"id": "t1", "payload": {}},
        "clarificationRounds": 1,
    })
    captured: dict = {}
    monkeypatch.setattr(runner, "execute_runtime_task", lambda task, **kw: captured.update(task=task, **kw) or {"ok": True})

    res = runner.resume_after_approval("t1", True, "rapor.pdf")
    assert res == {"ok": True}
    assert "dosyayı özetle" in captured["prompt_override"]
    assert "rapor.pdf" in captured["prompt_override"]
    assert captured["task"] == {"id": "t1", "payload": {}}


def test_resume_after_clarification_without_answer_fails_safe(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runner, _rtr = _new_clarify_runner()
    state_store.save_remote_task_link("t1", "", "", task_run_id="run1", status="waiting_approval")
    state_store.update_remote_task_link("t1", {
        "clarificationPending": True,
        "clarificationPrompt": "dosyayı özetle",
        "clarificationQuestion": "Hangi dosya?",
        "clarificationTask": {"id": "t1"},
        "clarificationRounds": 1,
    })
    captured: dict = {}
    monkeypatch.setattr(runner, "_fail_safe", lambda *a, **k: captured.update(k) or {"ok": False, "status": "failed"})

    runner.resume_after_approval("t1", True, "   ")
    assert captured["error_code"] == "clarification_unanswered"


def test_resume_after_clarification_rejection_cancels(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runner, _rtr = _new_clarify_runner()
    state_store.save_remote_task_link("t1", "", "", task_run_id="run1", status="waiting_approval")
    state_store.update_remote_task_link("t1", {"clarificationPending": True, "clarificationQuestion": "q"})

    res = runner.resume_after_approval("t1", False, "")
    assert res["status"] == "canceled"
    assert runner.host.canceled == ["t1"]


def test_plan_touches_blocklist_flags_irreversible_only() -> None:
    from runtime import remote_task_runner as rtr
    R = rtr.RemoteTaskRunner
    # Blocklist (geri alınamaz/dışa dönük) → True
    assert R._plan_touches_blocklist(
        {"planPreview": {"steps": [{"capability": "email_send"}]}}
    ) is True
    assert R._plan_touches_blocklist(
        {"planPreview": {"steps": [{"capability": "open_app"}, {"capability": "shell_run"}]}}
    ) is True
    # Zararsız yan etki (dosya/belge yazma, uygulama, takvim ekleme) → False
    assert R._plan_touches_blocklist(
        {"planPreview": {"steps": [{"capability": "document_write"}, {"capability": "add_calendar_event"}]}}
    ) is False
    assert R._plan_touches_blocklist({"planPreview": {"steps": []}}) is False
    assert R._plan_touches_blocklist({}) is False


def test_plan_requires_backend_approval_for_writes_and_unknown_capabilities() -> None:
    from runtime import remote_task_runner as rtr
    R = rtr.RemoteTaskRunner

    assert R._plan_requires_backend_approval(
        {"planPreview": {"steps": [{"capability": "web_research"}]}}
    ) is False
    assert R._plan_requires_backend_approval(
        {"planPreview": {"steps": [{"capability": "document_write"}]}}
    ) is True
    assert R._plan_requires_backend_approval(
        {"planPreview": {"steps": [{"capability": "browser_control"}]}}
    ) is True
    assert R._plan_requires_backend_approval(
        {"planPreview": {"steps": [{"capability": "unknown.write"}]}}
    ) is True
    assert R._plan_requires_backend_approval(
        {"planPreview": {"steps": []}}
    ) is True


def test_recoverable_replan_web_research_falls_back_to_local(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    revised = runtime._recoverable_replan({
        "reason": "tool_failure",
        "failedCapability": "web_research",
        "errorCode": "NETWORK_FAILED",
        "failedArgs": {"query": "kuantum"},
        "remainingSteps": [{"capability": "document_write", "args": {"prompt": "rapor"}}],
    })
    assert revised[0]["capability"] == "retrieve_context"
    assert revised[1]["capability"] == "document_write"


def test_recoverable_replan_skips_failed_observer_and_continues(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    # Gözlemci (document_read) patladı ama arkadan yazıcı adımı var → gözlemciyi
    # atla, kalanla devam (kısmi başarı > tam iptal).
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    revised = runtime._recoverable_replan({
        "reason": "verification_failure",
        "failedCapability": "document_read",
        "errorCode": "READ_FAILED",
        "failedArgs": {"path": "/x.pdf"},
        "remainingSteps": [{"capability": "document_write", "args": {"prompt": "özet"}}],
    })
    assert [s["capability"] for s in revised] == ["document_write"]


def test_recoverable_replan_gives_up_when_no_pattern(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    # Gözlemci değil (open_app) ve kurtarma deseni yok → boş (executor iptale düşer).
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    revised = runtime._recoverable_replan({
        "failedCapability": "open_app",
        "errorCode": "APP_NOT_FOUND",
        "remainingSteps": [{"capability": "close_app", "args": {"app_name": "X"}}],
    })
    assert revised == []


def test_capability_display_name_friendly_and_fallback() -> None:
    from runtime.capability_registry import capability_display_name
    assert capability_display_name("document_write") == "Belge oluşturma"
    assert capability_display_name("send_whatsapp_message") == "WhatsApp mesajı"
    assert capability_display_name("desktop_operator.run") == "Ekran otomasyonu"
    # bilinmeyen slug → okunabilir yedek (nokta sonrası, alt çizgi→boşluk, baş harf)
    assert capability_display_name("future_new.some_action") == "Some action"
    assert capability_display_name("") == "Bu işlem"


def test_readiness_error_is_friendly_and_structured() -> None:
    from runtime import remote_task_runner as rtr
    runner = rtr.RemoteTaskRunner(object())
    err = runner._readiness_error([
        {"capability": "open_app", "ready": True},
        {"capability": "document_write", "ready": False, "errorCode": "DEPENDENCY_UNAVAILABLE",
         "degradationReason": "dependency_unavailable"},
    ])
    assert err is not None
    assert err["capability"] == "document_write"
    assert err["displayName"] == "Belge oluşturma"
    assert err["code"] == "DEPENDENCY_UNAVAILABLE"
    # ham slug mesajda geçmez; dostane ad geçer
    assert "document_write" not in err["message"]
    assert "Belge oluşturma" in err["message"]
    assert err["degradationReason"] == "dependency_unavailable"


def test_readiness_error_none_when_all_ready() -> None:
    from runtime import remote_task_runner as rtr
    runner = rtr.RemoteTaskRunner(object())
    assert runner._readiness_error([{"capability": "open_app", "ready": True}]) is None


def test_fail_safe_surfaces_capability_unavailable_block(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    from runtime import remote_task_runner as rtr

    class _Host:
        def _report_runtime_task_status(self, _task_id, payload):
            from types import SimpleNamespace
            self.last_payload = payload
            return SimpleNamespace(ok=True, to_dict=lambda: {})

        def _remote_task_trace_payload(self, *_a, **_k):
            return {"type": "task_trace", "steps": []}

    host = _Host()
    runner = rtr.RemoteTaskRunner(host)
    runner._fail_safe(
        "t1", "run1",
        message="Belge oluşturma şu an kullanılamıyor: gerekli bileşen hazır değil.",
        error_code="DEPENDENCY_UNAVAILABLE",
        unavailable={"capability": "document_write", "displayName": "Belge oluşturma", "degradationReason": "dependency_unavailable"},
    )
    result = host.last_payload["result"]
    assert result["capabilityUnavailable"]["displayName"] == "Belge oluşturma"
    assert result["blocks"][0]["type"] == "capability_unavailable"
    assert result["blocks"][0]["displayName"] == "Belge oluşturma"


def test_server_brain_circuit_breaker_opens_and_resets(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    bridge._reset_server_brain_breaker()
    state = {"controlPlane": {"health": {"agent": {"serverBrainReady": True}}}}
    assert bridge._server_brain_ready(state) is True
    # Ardışık hata eşiği → devre açılır, brain atlanır (delegasyon/planlama düşer)
    for _ in range(bridge._SERVER_BRAIN_FAILURE_THRESHOLD):
        bridge._record_server_brain_outcome(False)
    assert bridge._server_brain_circuit_open() is True
    assert bridge._server_brain_ready(state) is False
    # Sağlıklı yanıt sayacı+cooldown'u sıfırlar
    bridge._record_server_brain_outcome(True)
    assert bridge._server_brain_circuit_open() is False
    assert bridge._server_brain_ready(state) is True
    bridge._reset_server_brain_breaker()


def test_server_brain_circuit_cooldown_expires(monkeypatch: pytest.MonkeyPatch) -> None:
    bridge._reset_server_brain_breaker()
    clock = {"now": 1000.0}
    monkeypatch.setattr(bridge.time, "monotonic", lambda: clock["now"])
    for _ in range(bridge._SERVER_BRAIN_FAILURE_THRESHOLD):
        bridge._record_server_brain_outcome(False)
    assert bridge._server_brain_circuit_open() is True
    clock["now"] += bridge._SERVER_BRAIN_COOLDOWN_SECONDS + 1.0
    assert bridge._server_brain_circuit_open() is False  # cooldown doldu
    bridge._reset_server_brain_breaker()


def test_single_failure_does_not_open_circuit(monkeypatch: pytest.MonkeyPatch) -> None:
    bridge._reset_server_brain_breaker()
    bridge._record_server_brain_outcome(False)
    assert bridge._server_brain_circuit_open() is False  # eşiğin altında
    bridge._reset_server_brain_breaker()


def test_tool_catalog_cached_per_platform() -> None:
    from runtime import structured_planner as sp
    a = sp.tool_catalog(platform="darwin")
    b = sp.tool_catalog(platform="darwin")
    assert a is b  # aynı önbelleklenmiş nesne (her çağrıda yeniden kurulmaz)
    c = sp.tool_catalog(platform="win32")
    assert c is not a


def test_recoverable_replan_corrects_app_content_open_app(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    # Eski istemci/backend planından gelen "Chrome dan kedi resmi" gibi uydurma
    # open_app hedefi APP_NOT_FOUND ile düşünce plan yerinde düzeltilir:
    # uygulamayı aç + içerik için tarayıcı adımı.
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    revised = runtime._recoverable_replan({
        "reason": "tool_failure",
        "failedCapability": "open_app",
        "errorCode": "APP_NOT_FOUND",
        "failedArgs": {"app_name": "Chrome dan kedi resmi"},
        "remainingSteps": [],
    })
    assert [s["capability"] for s in revised] == ["open_app", "browser_control"]
    assert revised[0]["args"]["app_name"] == "Google Chrome"


def test_replan_observation_includes_app_suggestions_for_app_not_found(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    from runtime import structured_planner

    observation = structured_planner.build_replan_observation({
        "reason": "tool_failure",
        "failedCapability": "open_app",
        "errorCode": "APP_NOT_FOUND",
        "failedArgs": {"app_name": "Chorme"},
        "appSuggestions": ["Google Chrome", "Chromium"],
    })
    assert observation["failedStep"]["suggestions"] == ["Google Chrome", "Chromium"]


def test_sweep_interrupted_tasks_fails_stuck_running_items(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    # Daemon yeniden başladığında önceki süreçten 'running' kalmış görevler
    # dürüstçe failed'a çekilir; waiting_approval/terminal görevlere dokunulmaz.
    _isolate_state(monkeypatch, tmp_path)
    state_store.upsert_task_inbox_item({"id": "11111111-1111-4111-8111-111111111111", "title": "Tarayıcı kapat", "status": "running"})
    state_store.upsert_task_inbox_item({"id": "22222222-2222-4222-8222-222222222222", "title": "Onay bekleyen", "status": "waiting_approval"})
    state_store.upsert_task_inbox_item({"id": "33333333-3333-4333-8333-333333333333", "title": "Biten", "status": "completed"})

    runtime = bridge.RuntimeBridge()
    reported: list[tuple[str, str]] = []
    monkeypatch.setattr(
        runtime,
        "_report_runtime_task_status",
        lambda task_id, payload: reported.append((task_id, str(payload.get("status", "")))) or None,
    )
    swept = runtime.remote_task_runner.sweep_interrupted_tasks()

    assert swept == 1
    assert reported and reported[0][0] == "11111111-1111-4111-8111-111111111111"
    assert reported[0][1] == "failed"
    stuck = state_store.get_task_inbox_item("11111111-1111-4111-8111-111111111111")
    assert stuck is not None and stuck["status"] == "failed"
    waiting = state_store.get_task_inbox_item("22222222-2222-4222-8222-222222222222")
    assert waiting is not None and waiting["status"] == "waiting_approval"


def test_status_payload_carries_task_title_for_local_inbox(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    payload = runtime.remote_task_runner._status_payload(
        "44444444-4444-4444-8444-444444444444",
        "run-1",
        status="running",
        message="Yürütülüyor.",
        task={"id": "44444444-4444-4444-8444-444444444444", "payload": {"prompt": "Chrome dan kedi resmi aç"}},
    )
    assert payload["title"] == "Chrome dan kedi resmi aç"
    runtime._sync_task_inbox_status("44444444-4444-4444-8444-444444444444", payload)
    item = state_store.get_task_inbox_item("44444444-4444-4444-8444-444444444444")
    assert item is not None and item["title"] == "Chrome dan kedi resmi aç"


def test_recoverable_replan_hands_browser_goal_from_failed_operator(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    # Görsel operatör izin/doğrulama nedeniyle düşerse tarayıcı-şekilli hedefi
    # tarayıcı ajanı devralır (canlı arıza regresyonu).
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    revised = runtime._recoverable_replan({
        "reason": "verification_failure",
        "failedCapability": "desktop_operator.run",
        "errorCode": "VERIFICATION_FAILED",
        "failedArgs": {"goal": "youtube kanalıma girip video linklerini topla"},
        "remainingSteps": [],
    })
    assert [s["capability"] for s in revised] == ["browser_agent.run"]
    assert revised[0]["args"]["goal"] == "youtube kanalıma girip video linklerini topla"

    # Tarayıcı-şekilli OLMAYAN hedef devredilmez (native iş tarayıcıda yapılamaz).
    revised_native = runtime._recoverable_replan({
        "reason": "verification_failure",
        "failedCapability": "desktop_operator.run",
        "errorCode": "PERMISSION_REQUIRED",
        "failedArgs": {"goal": "Ayarlarda karanlık modu aç"},
        "remainingSteps": [],
    })
    assert revised_native == []


def test_terminal_payload_strips_leaked_internal_envelope(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """Canlı arıza: iç planlama/replan zarfı asistan mesajına sızıp kullanıcıya
    ham JSON gösteriliyordu. Kalkan: zarf temizlenir, temiz yedeğe düşülür."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    envelope = (
        '{"type": "elyan.plan.replan", "contract": "elyan.plan.v2", '
        '"workOrder": {"capabilityScope": ["analyze_screen", "desktop_os.processes"]}}'
    )
    payload, _artifacts, ok = runtime._runtime_task_terminal_payload(
        {
            "chatOk": True,
            "assistantMessage": "analyze_screen\n" + envelope,
            "provider": "local",
            "executionTrace": {"steps": [{"label": "Ekran okundu", "status": "completed"}]},
        }
    )
    assert ok is True
    # Ham zarf ASLA kullanıcıya gitmez.
    assert "elyan.plan" not in payload["summary"]
    assert "capabilityScope" not in payload["summary"]
    assert "elyan.plan" not in payload["safeSummary"]
    # Temiz yedek (yürütme kanıtından sentez) gösterilir.
    assert payload["summary"].strip() != ""


def test_permission_required_is_not_replanned(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """İzin hatası LLM replan ile çözülemez; replan atlanır (zarf sızıntısı da
    olmaz), adımın okunaklı mesajı yüzeye çıkar."""
    _isolate_state(monkeypatch, tmp_path)
    runtime = bridge.RuntimeBridge()
    revised = runtime._recoverable_replan(
        {
            "failedCapability": "analyze_screen",
            "errorCode": "PERMISSION_REQUIRED",
            "message": "Ekran kaydı izni gerekiyor.",
            "remainingSteps": [],
        }
    )
    assert revised == []
