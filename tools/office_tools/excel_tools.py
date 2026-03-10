"""Excel tools with robust read/write and dataframe analysis flows."""

from __future__ import annotations

import asyncio
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Union

from config.settings import DESKTOP
from security.validator import validate_path
from utils.logger import get_logger

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except Exception:  # pragma: no cover - environment dependent
    Workbook = None
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    Table = None
    TableStyleInfo = None

try:
    import pandas as pd  # Optional dependency
except Exception:  # pragma: no cover - environment dependent
    pd = None

logger = get_logger("office.excel")

MAX_FILE_SIZE = 20 * 1024 * 1024


def _normalize_headers(raw_headers: Any) -> List[str]:
    if isinstance(raw_headers, (list, tuple)):
        return [str(h).strip() for h in raw_headers if str(h).strip()]
    return []


def _openpyxl_ready() -> tuple[bool, Optional[str]]:
    if Workbook is None or load_workbook is None:
        return False, "openpyxl kurulu degil. Excel islemleri icin 'pip install openpyxl' gerekli."
    return True, None


def _safe_sheet_name(name: str) -> str:
    cleaned = str(name or "Sheet").strip().replace("/", "-").replace("\\", "-")
    cleaned = cleaned.replace("*", "").replace("?", "").replace("[", "(").replace("]", ")")
    return cleaned[:31] or "Sheet"


def _normalize_rows(raw_data: Any, raw_headers: Any) -> tuple[List[str], List[Dict[str, Any]]]:
    cols: List[str] = _normalize_headers(raw_headers)
    rows: List[Dict[str, Any]] = []

    if raw_data is None:
        return cols, rows

    if isinstance(raw_data, dict):
        if not cols:
            cols = [str(k) for k in raw_data.keys()]
        rows.append({k: raw_data.get(k, "") for k in cols})
        return cols, rows

    if isinstance(raw_data, (list, tuple)):
        data_list = list(raw_data)
        if not data_list:
            return cols, rows

        first = data_list[0]
        if isinstance(first, dict):
            if not cols:
                ordered_keys: List[str] = []
                for item in data_list:
                    if not isinstance(item, dict):
                        continue
                    for key in item.keys():
                        s_key = str(key)
                        if s_key not in ordered_keys:
                            ordered_keys.append(s_key)
                cols = ordered_keys
            for item in data_list:
                if isinstance(item, dict):
                    rows.append({k: item.get(k, "") for k in cols})
                else:
                    rows.append({cols[0]: str(item) if cols else str(item)})
            return cols, rows

        if isinstance(first, (list, tuple)):
            matrix = [list(r) for r in data_list]
            width = max(len(r) for r in matrix) if matrix else 0
            if not cols:
                cols = [f"Sutun{i + 1}" for i in range(width)]
            for row_values in matrix:
                padded = list(row_values) + [""] * max(0, len(cols) - len(row_values))
                rows.append({col: padded[idx] if idx < len(padded) else "" for idx, col in enumerate(cols)})
            return cols, rows

        if not cols:
            cols = ["Veri"]
        for item in data_list:
            rows.append({cols[0]: item})
        return cols, rows

    if not cols:
        cols = ["Veri"]
    rows.append({cols[0]: raw_data})
    return cols, rows


def _sheet_payloads(
    data: Any,
    headers: Any,
    sheet_name: str,
    multi_sheet: bool,
) -> Dict[str, tuple[List[str], List[Dict[str, Any]]]]:
    if multi_sheet:
        if not isinstance(data, dict):
            return {}
        result: Dict[str, tuple[List[str], List[Dict[str, Any]]]] = {}
        header_map = headers if isinstance(headers, dict) else {}
        for s_name, s_data in data.items():
            s_headers = header_map.get(s_name) if isinstance(header_map, dict) else headers
            cols, rows = _normalize_rows(s_data, s_headers)
            if cols and rows:
                result[_safe_sheet_name(str(s_name))] = (cols, rows)
        return result

    cols, rows = _normalize_rows(data, headers)
    if not cols or not rows:
        return {}
    return {_safe_sheet_name(sheet_name): (cols, rows)}


def _apply_header_style(ws, header_row: int, headers: Sequence[str]) -> None:
    if Font is None or PatternFill is None or Alignment is None:
        return
    for c_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")


def _auto_size(ws, col_count: int) -> None:
    for idx in range(1, col_count + 1):
        col = ws.cell(row=1, column=idx).column_letter
        max_length = 0
        for row in ws.iter_rows(min_col=idx, max_col=idx):
            value = row[0].value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        ws.column_dimensions[col].width = min(max(10, max_length + 2), 60)


def _ensure_table(ws, name_hint: str, header_row: int, data_end_row: int, col_count: int) -> None:
    if data_end_row <= header_row or col_count < 1:
        return

    if ws.tables:
        return

    first_col = ws.cell(row=header_row, column=1).column_letter
    last_col = ws.cell(row=header_row, column=col_count).column_letter
    ref = f"{first_col}{header_row}:{last_col}{data_end_row}"

    safe_name = "tbl_" + "".join(ch if ch.isalnum() else "_" for ch in name_hint)[:20]
    if not safe_name[4:]:
        safe_name = "tbl_sheet"

    if Table is None or TableStyleInfo is None:
        return

    table = Table(displayName=safe_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


async def read_excel(
    path: str,
    sheet_name: Union[str, int, None] = None,
    use_pandas: bool = True,
    max_rows: Optional[int] = None,
    include_formulas: bool = False,
) -> dict[str, Any]:
    """Read content from an Excel file with optional row limits and formula mode."""
    try:
        openpyxl_ok, openpyxl_err = _openpyxl_ready()
        if not openpyxl_ok and not (use_pandas and pd is not None and not include_formulas):
            return {"success": False, "error": openpyxl_err}

        is_valid, error, _ = validate_path(path)
        if not is_valid:
            return {"success": False, "error": error}

        file_path = Path(path).expanduser().resolve()
        if not file_path.exists():
            return {"success": False, "error": f"Dosya bulunamadi: {file_path.name}"}

        file_size = file_path.stat().st_size
        if file_size > MAX_FILE_SIZE:
            return {
                "success": False,
                "error": f"Dosya cok buyuk ({file_size} bytes). Limit: {MAX_FILE_SIZE} bytes",
            }

        def _read_openpyxl() -> tuple[Any, Any, List[str]]:
            data_only = not include_formulas
            wb = load_workbook(str(file_path), data_only=data_only)

            if sheet_name is None:
                ws = wb.active
            elif isinstance(sheet_name, int):
                if sheet_name < 0 or sheet_name >= len(wb.sheetnames):
                    raise ValueError("sheet index gecersiz")
                ws = wb[wb.sheetnames[sheet_name]]
            elif sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            if max_rows is not None:
                rows = rows[: max(0, int(max_rows))]
            return rows, f"{len(rows)} satir", wb.sheetnames

        def _read_pandas() -> tuple[Any, Any, List[str]]:
            target_sheet = 0 if sheet_name is None else sheet_name
            df = pd.read_excel(str(file_path), sheet_name=target_sheet)

            if isinstance(df, dict):
                result_data: Dict[str, Any] = {}
                summary: Dict[str, Any] = {}
                for s, d in df.items():
                    current = d.head(max_rows) if max_rows is not None else d
                    result_data[s] = current.to_dict(orient="records")
                    summary[s] = f"{len(current)} satir, {len(current.columns)} sutun"
                return result_data, summary, list(df.keys())

            current = df.head(max_rows) if max_rows is not None else df
            result_data = current.to_dict(orient="records")
            summary = f"{len(current)} satir, {len(current.columns)} sutun"
            sheet_label = [str(sheet_name if sheet_name is not None else "active")]
            return result_data, summary, sheet_label

        loop = asyncio.get_event_loop()
        if use_pandas and pd is not None and not include_formulas:
            data, summary, sheets = await loop.run_in_executor(None, _read_pandas)
        else:
            data, summary, sheets = await loop.run_in_executor(None, _read_openpyxl)

        row_count = len(data) if not isinstance(data, dict) else sum(len(v) for v in data.values())
        return {
            "success": True,
            "path": str(file_path),
            "data": data,
            "summary": summary,
            "sheets": sheets,
            "row_count": row_count,
            "size_bytes": file_size,
        }
    except Exception as e:
        logger.error(f"Read Excel error: {e}")
        return {"success": False, "error": str(e)}


async def write_excel(
    path: str | None = None,
    data: Any = None,
    headers: Any = None,
    sheet_name: str = "Sheet1",
    append: bool = False,
    multi_sheet: bool = False,
    create_table: bool = True,
    freeze_header: bool = True,
    autofilter: bool = True,
) -> dict[str, Any]:
    """Write Excel files with optional multi-sheet payload and table generation."""
    try:
        openpyxl_ok, openpyxl_err = _openpyxl_ready()
        if not openpyxl_ok:
            return {"success": False, "error": openpyxl_err}

        if not path:
            path = str(DESKTOP / "tablo.xlsx")

        file_path = Path(path).expanduser().resolve()
        if file_path.suffix.lower() != ".xlsx":
            file_path = file_path.with_suffix(".xlsx")

        is_valid, error, _ = validate_path(str(file_path))
        if not is_valid:
            return {"success": False, "error": error}

        payloads = _sheet_payloads(data=data, headers=headers, sheet_name=sheet_name, multi_sheet=multi_sheet)
        if not payloads:
            return {"success": False, "error": "Yazilacak veri bulunamadi (data bos)."}

        def _write() -> tuple[bool, int, List[str]]:
            if append and file_path.exists():
                wb = load_workbook(str(file_path))
            else:
                wb = Workbook()
                if wb.active and wb.active.max_row == 1 and wb.active.max_column == 1 and wb.active["A1"].value is None:
                    wb.remove(wb.active)

            total_rows = 0
            touched_sheets: List[str] = []

            for s_name, (cols, rows) in payloads.items():
                ws = wb[s_name] if s_name in wb.sheetnames else wb.create_sheet(title=s_name)
                touched_sheets.append(s_name)

                existing_headers: List[str] = []
                has_header = ws.max_row >= 1 and any(ws.cell(1, c).value for c in range(1, ws.max_column + 1))
                if has_header:
                    existing_headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]

                if has_header and append and existing_headers:
                    final_cols = [h for h in existing_headers if h]
                    if not final_cols:
                        final_cols = cols
                else:
                    final_cols = cols

                if not has_header:
                    _apply_header_style(ws, 1, final_cols)
                    start_row = 2
                else:
                    start_row = ws.max_row + 1 if append else ws.max_row + 1

                for r_idx, row_dict in enumerate(rows, start_row):
                    for c_idx, header in enumerate(final_cols, 1):
                        val = row_dict.get(header, "")
                        cell = ws.cell(row=r_idx, column=c_idx, value=val)
                        if isinstance(val, str) and val.startswith("="):
                            cell.value = val

                last_data_row = start_row + len(rows) - 1
                total_rows += len(rows)

                if freeze_header:
                    ws.freeze_panes = "A2"

                if autofilter:
                    ws.auto_filter.ref = ws.dimensions

                _auto_size(ws, len(final_cols))

                if create_table and not append:
                    _ensure_table(ws, s_name, 1, last_data_row, len(final_cols))

            file_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(file_path))
            return True, total_rows, touched_sheets

        loop = asyncio.get_event_loop()
        wrote, row_count, touched_sheets = await loop.run_in_executor(None, _write)
        if not wrote:
            return {"success": False, "error": "Excel dosyasi yazilamadi."}

        if not file_path.exists():
            return {
                "success": False,
                "error": "WRITE_FAILED: Dosya diskte bulunamadi.",
                "error_code": "FILE_NOT_FOUND",
            }

        file_size = file_path.stat().st_size
        if file_size < 1000:
            return {
                "success": False,
                "error": (
                    "WRITE_POSTCHECK_FAILED: Excel dosyasi boyutu supheli sekilde kucuk "
                    f"({file_size} bytes)."
                ),
                "error_code": "WRITE_POSTCHECK_FAILED",
            }

        return {
            "success": True,
            "path": str(file_path),
            "row_count": row_count,
            "size_bytes": file_size,
            "sheets": touched_sheets,
        }
    except Exception as e:
        logger.error(f"Write Excel error: {e}")
        return {"success": False, "error": str(e)}


def _apply_query(df, query_payload: Dict[str, Any]):
    working = df.copy()

    filters = query_payload.get("filters") or []
    for rule in filters:
        col = rule.get("column")
        op = str(rule.get("op", "=="))
        val = rule.get("value")
        if col not in working.columns:
            continue
        if op == "==":
            working = working[working[col] == val]
        elif op == "!=":
            working = working[working[col] != val]
        elif op == ">":
            working = working[working[col] > val]
        elif op == ">=":
            working = working[working[col] >= val]
        elif op == "<":
            working = working[working[col] < val]
        elif op == "<=":
            working = working[working[col] <= val]
        elif op.lower() == "contains":
            working = working[working[col].astype(str).str.contains(str(val), na=False)]

    selected_cols = query_payload.get("columns")
    if selected_cols:
        keep = [c for c in selected_cols if c in working.columns]
        if keep:
            working = working[keep]

    sort_rules = query_payload.get("sort_by")
    if sort_rules:
        if isinstance(sort_rules, str):
            sort_rules = [{"column": sort_rules, "ascending": True}]
        cols = [s.get("column") for s in sort_rules if s.get("column") in working.columns]
        asc = [bool(s.get("ascending", True)) for s in sort_rules if s.get("column") in working.columns]
        if cols:
            working = working.sort_values(by=cols, ascending=asc)

    group_by = query_payload.get("group_by")
    aggregations = query_payload.get("aggregations")
    grouped_records = None
    if group_by and aggregations:
        gb_cols = [c for c in group_by if c in working.columns]
        agg_cols = {k: v for k, v in aggregations.items() if k in working.columns}
        if gb_cols and agg_cols:
            grouped = working.groupby(gb_cols, dropna=False).agg(agg_cols).reset_index()
            grouped_records = grouped.to_dict(orient="records")

    top_n = query_payload.get("top_n")
    if isinstance(top_n, int) and top_n > 0:
        working = working.head(top_n)

    return working, grouped_records


async def analyze_excel_data(path: str, query: Any) -> dict[str, Any]:
    """Analyze Excel data with filter/group/sort primitives and summaries."""
    try:
        openpyxl_ok, openpyxl_err = _openpyxl_ready()
        if not openpyxl_ok and pd is None:
            return {"success": False, "error": openpyxl_err}

        is_valid, error, _ = validate_path(path)
        if not is_valid:
            return {"success": False, "error": error}

        file_path = Path(path).expanduser().resolve()
        if not file_path.exists():
            return {"success": False, "error": f"Dosya bulunamadi: {file_path.name}"}

        query_payload: Dict[str, Any] = {}
        if isinstance(query, dict):
            query_payload = query
        elif isinstance(query, str) and query.strip():
            raw = query.strip()
            if raw.startswith("{"):
                try:
                    query_payload = json.loads(raw)
                except Exception:
                    query_payload = {"mode": raw.lower()}
            else:
                query_payload = {"mode": raw.lower()}

        if pd is None:
            reader = await read_excel(str(file_path), use_pandas=False, max_rows=200)
            if not reader.get("success"):
                return reader
            return {
                "success": True,
                "analysis": {
                    "mode": "fallback",
                    "summary": reader.get("summary"),
                    "row_count": reader.get("row_count"),
                    "sheets": reader.get("sheets", []),
                },
            }

        def _analyze() -> Dict[str, Any]:
            sheet = query_payload.get("sheet", 0)
            df = pd.read_excel(str(file_path), sheet_name=sheet)

            mode = str(query_payload.get("mode", "summary")).lower()
            if mode in {"summary", "describe", "stats"} and not any(
                query_payload.get(k) for k in ("filters", "group_by", "sort_by", "columns", "top_n")
            ):
                return {
                    "mode": "summary",
                    "columns": list(df.columns),
                    "row_count": int(len(df)),
                    "stats": df.describe(include="all").fillna("").to_dict(),
                    "null_counts": df.isnull().sum().to_dict(),
                }

            working, grouped_records = _apply_query(df, query_payload)
            payload = {
                "mode": "query",
                "columns": list(working.columns),
                "row_count": int(len(working)),
                "records": working.to_dict(orient="records"),
                "null_counts": working.isnull().sum().to_dict(),
            }
            if grouped_records is not None:
                payload["grouped"] = grouped_records
            return payload

        loop = asyncio.get_event_loop()
        analysis = await loop.run_in_executor(None, _analyze)
        return {"success": True, "analysis": analysis}
    except Exception as e:
        logger.error(f"analyze_excel_data error: {e}")
        return {"success": False, "error": str(e)}
